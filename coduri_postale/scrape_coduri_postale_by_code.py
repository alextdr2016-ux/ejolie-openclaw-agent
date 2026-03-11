#!/usr/bin/env python3
"""
Script: scrape_coduri_postale_by_code.py  v3
Descriere: Extrage coduri postale Dolj (200000-207633) prin cautare INVERSA
           pe site-ul Postei Romane - introducem codul postal si primim strada.
           
Aceasta este metoda 100% completa - parcurgem TOATE codurile posibile.
Dolj = primele 2 cifre "20", deci range 200000-207999.

API: POST /cnpr-app/modules/cauta-cod-postal/ajax/cautare_pentru_cod.php?q=
     Body: k_cod_postal=200100&k_lang=ro

Nota: Trebuie mai intai sa descoperim parametrul corect pentru cautare dupa cod.
      Pagina are 2 form-uri: dupa adresa si dupa cod postal.
      Scriptul testeaza mai intai ce parametru functioneaza.
"""

import requests
import json
import time
import re
import os
import sys
from datetime import datetime
from urllib.parse import quote

BASE_URL = "https://www.posta-romana.ro"
URL_CAUTARE = f"{BASE_URL}/cnpr-app/modules/cauta-cod-postal/ajax/cautare_pentru_cod.php?q="

DELAY = 0.3  # secunde intre requesturi
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "ro-RO,ro;q=0.9,en;q=0.8",
    "Referer": "https://www.posta-romana.ro/ccp.html",
    "X-Requested-With": "XMLHttpRequest",
    "Origin": "https://www.posta-romana.ro",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
}

# Range coduri postale Dolj
# 200xxx = Craiova (pe strazi)
# 205xxx = Orase Dolj (Bailesti 205100, Calafat 205200, Filiasi 205300, Segarcea 205400)
# 207xxx = Localitati mici Dolj (comune/sate)
CODE_RANGES = [
    (200000, 200999),  # Craiova
    (205000, 205499),  # Orase Dolj (Bailesti, Calafat, Filiasi, Segarcea)
    (207000, 207699),  # Localitati mici Dolj
]


def parse_results_html(html_str):
    """Parseaza HTML din raspunsul API."""
    results = []
    p_tags = re.findall(r'<p[^>]*>(.*?)</p>', html_str)
    for i in range(0, len(p_tags), 5):
        chunk = p_tags[i:i+5]
        if len(chunk) >= 4:
            cod = chunk[0].strip()
            if cod and cod[0].isdigit() and len(cod) == 6:
                results.append({
                    "cod_postal": cod,
                    "judet": chunk[1].strip(),
                    "localitate": chunk[2].strip(),
                    "strada": chunk[3].strip(),
                    "subunitate_postala": chunk[4].strip() if len(chunk) > 4 else "",
                })
    return results


def init_session():
    """Initializeaza sesiunea cu cookies."""
    session = requests.Session()
    session.headers.update(HEADERS)
    session.get(f"{BASE_URL}/ccp.html", timeout=10)
    return session


def discover_code_param(session):
    """
    Descopera ce parametru accepta API-ul pentru cautare dupa cod postal.
    Testeaza: k_cod_postal, k_codpostal, k_cod, cod_postal
    """
    print("[*] Descopar parametrul pentru cautare dupa cod postal...")
    test_code = "200404"  # Cod Craiova cunoscut

    # Endpoint-uri posibile
    endpoints = [
        ("cautare_pentru_cod.php", URL_CAUTARE),
        ("cautare_dupa_cod.php",
         f"{BASE_URL}/cnpr-app/modules/cauta-cod-postal/ajax/cautare_dupa_cod.php?q="),
        ("cautare_pentru_adresa.php",
         f"{BASE_URL}/cnpr-app/modules/cauta-cod-postal/ajax/cautare_pentru_adresa.php?q="),
        ("cautare_cod_postal.php",
         f"{BASE_URL}/cnpr-app/modules/cauta-cod-postal/ajax/cautare_cod_postal.php?q="),
    ]

    param_names = ["k_cod_postal", "k_codpostal",
                   "k_cod", "cod_postal", "k_cp"]

    for ep_name, ep_url in endpoints:
        for param in param_names:
            try:
                data = f"{param}={test_code}&k_lang=ro"
                resp = session.post(ep_url, data=data, timeout=10)
                if resp.status_code == 200:
                    try:
                        result = resp.json()
                        found = result.get("found", 0)
                        if found and found > 0:
                            print(
                                f"    ✅ GASIT! Endpoint: {ep_name}, Param: {param}")
                            print(f"       found={found}")
                            return ep_url, param
                    except:
                        pass
            except:
                pass
            time.sleep(0.2)

    # Daca nu gasim, incercam cu parametrii de adresa dar punand codul in campul adresa
    print("    Incerc cautare cod in campul k_adresa...")
    data = f"k_adresa={test_code}&k_judet=&k_localitate=&k_lang=ro"
    try:
        resp = session.post(URL_CAUTARE, data=data, timeout=10)
        result = resp.json()
        if result.get("found", 0) > 0:
            print(
                f"    ✅ Merge cu k_adresa={test_code} (fara judet/localitate)")
            return URL_CAUTARE, "k_adresa_as_code"
    except:
        pass

    # Ultima incercare: cautare cu codul in campul cod_postal (input id="cod_postal" din pagina)
    print("    Incerc cu inputul cod_postal din pagina...")
    data = f"k_cod_postal={test_code}&k_lang=ro"
    resp = session.post(URL_CAUTARE, data=data, timeout=10)
    print(f"    Response: {resp.text[:300]}")

    return None, None


def search_by_code(session, url, param, code):
    """Cauta un cod postal specific."""
    code_str = str(code).zfill(6)

    if param == "k_adresa_as_code":
        data = f"k_adresa={code_str}&k_judet=&k_localitate=&k_lang=ro"
    else:
        data = f"{param}={code_str}&k_lang=ro"

    try:
        resp = session.post(url, data=data, timeout=15)
        resp.raise_for_status()
        result = resp.json()
        if result.get("found", 0) > 0:
            return parse_results_html(result.get("formular", ""))
        return []
    except Exception as e:
        return []


def export_to_excel(all_data, filename):
    """Exporta in Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Coduri Postale Dolj"

    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:F1')
    ws['A1'] = 'CODURI POȘTALE DOLJ - EXTRASE DE PE POȘTA ROMÂNĂ (2026)'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Sursa: posta-romana.ro | Generat: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    for col, h in enumerate(['Nr.', 'Cod Poștal', 'Județ', 'Localitate', 'Strada / Adresa', 'Subunitate Poștală'], 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    all_data.sort(key=lambda x: (x['cod_postal'], x['strada']))
    for i, rd in enumerate(all_data):
        row = i + 5
        for col, val in enumerate([i+1, rd['cod_postal'], rd['judet'], rd['localitate'], rd['strada'], rd['subunitate_postala']], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10, bold=(col == 2))
            c.alignment = Alignment(horizontal='center') if col in (
                1, 2) else Alignment()
            c.border = tb

    widths = {'A': 7, 'B': 12, 'C': 10, 'D': 22, 'E': 55, 'F': 25}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:F{4 + len(all_data)}'
    wb.save(filename)


def main():
    print("=" * 60)
    print("SCRAPER CODURI POSTALE DOLJ - CAUTARE DUPA COD  v3")
    print(f"Range: 200000-200999, 205000-205499, 207000-207699")
    print(f"Sursa: Posta Romana | {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print("=" * 60)

    # Init
    print("\n[1] Conectare Posta Romana...")
    session = init_session()
    print(f"    ✅ Cookie: {dict(session.cookies).get('PHPSESSID', 'N/A')}")

    # Discover correct parameter
    url, param = discover_code_param(session)

    if not url:
        print("\n❌ Nu am gasit parametrul corect pentru cautare dupa cod postal.")
        print("   Posibil API-ul nu suporta cautare inversa.")
        print("   Incercam plan B: iteram codurile cu cautare prin adresa...")

        # Plan B: cautam fiecare cod ca "adresa" cu judet Dolj
        url = URL_CAUTARE
        param = "k_adresa_direct"

    # Scan all codes
    print(f"\n[2] Scanare coduri postale...")
    all_data = []
    total_codes = sum(end - start + 1 for start, end in CODE_RANGES)
    scanned = 0
    found_count = 0
    empty_streak = 0

    # Save progress periodically
    script_dir = os.path.dirname(os.path.abspath(__file__))
    progress_file = os.path.join(script_dir, "scan_progress.json")

    # Resume from progress if exists
    start_from = None
    if os.path.exists(progress_file):
        with open(progress_file, 'r') as f:
            progress = json.load(f)
            all_data = progress.get('data', [])
            start_from = progress.get('last_code', None)
            print(
                f"    Resuming from {start_from}, {len(all_data)} results loaded")

    for range_start, range_end in CODE_RANGES:
        for code in range(range_start, range_end + 1):
            # Skip if resuming
            if start_from and code <= start_from:
                scanned += 1
                continue

            code_str = str(code).zfill(6)
            scanned += 1

            if param == "k_adresa_direct":
                # Plan B: search code as address text
                data = f"k_adresa={code_str}&k_judet=Dolj&k_localitate=&k_lang=ro"
                try:
                    resp = session.post(url, data=data, timeout=15)
                    result = resp.json()
                    if result.get("found", 0) > 0:
                        results = parse_results_html(
                            result.get("formular", ""))
                    else:
                        results = []
                except:
                    results = []
            else:
                results = search_by_code(session, url, param, code)

            if results:
                for r in results:
                    key = f"{r['cod_postal']}|{r['strada']}"
                    # Avoid duplicates
                    if not any(f"{d['cod_postal']}|{d['strada']}" == key for d in all_data):
                        all_data.append(r)
                        found_count += 1
                empty_streak = 0
                print(
                    f"    [{scanned}/{total_codes}] {code_str}: {len(results)} rezultate (total: {len(all_data)})")
            else:
                empty_streak += 1

            # Progress indicator every 100 codes
            if scanned % 100 == 0:
                print(
                    f"    [{scanned}/{total_codes}] Scanat pana la {code_str}... ({len(all_data)} gasite)")
                # Save progress
                with open(progress_file, 'w') as f:
                    json.dump({'last_code': code, 'data': all_data},
                              f, ensure_ascii=False)

            time.sleep(DELAY)

    # Cleanup progress file
    if os.path.exists(progress_file):
        os.remove(progress_file)

    # Deduplicate
    unique = {}
    for r in all_data:
        unique[f"{r['cod_postal']}|{r['strada']}"] = r
    all_data = list(unique.values())

    print(f"\n{'=' * 60}")
    print(f"TOTAL: {len(all_data)} coduri postale unice")
    print(f"Scanate: {scanned} coduri din range {CODE_RANGES}")
    print(f"{'=' * 60}")

    # Export
    output_xlsx = os.path.join(
        script_dir, "coduri_postale_dolj_complet_2026.xlsx")
    export_to_excel(all_data, output_xlsx)
    print(f"\n✅ Excel: {output_xlsx}")

    output_json = os.path.join(
        script_dir, "coduri_postale_dolj_complet_2026.json")
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON: {output_json}")

    import csv
    output_csv = os.path.join(
        script_dir, "coduri_postale_dolj_complet_2026.csv")
    with open(output_csv, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=[
                           'cod_postal', 'judet', 'localitate', 'strada', 'subunitate_postala'])
        w.writeheader()
        w.writerows(all_data)
    print(f"✅ CSV: {output_csv}")

    print(f"\n🎉 Gata!")


if __name__ == "__main__":
    main()

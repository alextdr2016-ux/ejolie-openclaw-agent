#!/usr/bin/env python3
"""
scrape_coduri_postale_IASI_posta_romana.py
Extrage coduri postale Iași prin căutare inversă pe Poșta Română

Endpoint: POST /cnpr-app/modules/cauta-cod-postal/ajax/cautare_cod.php?q=
Body: k_cod_postal=700000&k_lang=ro

Coduri poștale Iași:
  700xxx = Iași municipiu (pe străzi)
  705xxx = Orașe Iași (Pașcani 705200, Hârlău 705300, Tg. Frumos 705300, Podu Iloaiei 705400)
  707xxx = Localități mici (comune/sate)

Range complet: 700000 → 707999 = 8000 coduri
Delay: 0.2s → ~27 minute
Resume: Ctrl+C safe
"""

import requests
import json
import time
import re
import os
import csv
from datetime import datetime

BASE_URL = "https://www.posta-romana.ro"
URL_INIT = f"{BASE_URL}/ccp.html"
URL_COD = f"{BASE_URL}/cnpr-app/modules/cauta-cod-postal/ajax/cautare_cod.php?q="

CODE_START = 700000
CODE_END = 707999
TOTAL_CODES = CODE_END - CODE_START + 1

DELAY = 0.2

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "ro-RO,ro;q=0.9,en;q=0.8",
    "Referer": "https://www.posta-romana.ro/ccp.html",
    "X-Requested-With": "XMLHttpRequest",
    "Origin": "https://www.posta-romana.ro",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
}


def parse_results_html(html_str):
    results = []
    p_tags = re.findall(r'<p[^>]*>(.*?)</p>', html_str)
    for i in range(0, len(p_tags), 5):
        chunk = p_tags[i:i+5]
        if len(chunk) >= 4:
            cod = chunk[0].strip()
            if cod and cod[0].isdigit() and len(cod) == 6:
                results.append({
                    "cod_postal": cod,
                    "judet": chunk[1].strip(),
                    "localitate": chunk[2].strip(),
                    "strada": chunk[3].strip(),
                    "subunitate_postala": chunk[4].strip() if len(chunk) > 4 else "",
                })
    return results


def init_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    s.get(URL_INIT, timeout=10)
    return s


def search_code(session, code):
    code_str = str(code).zfill(6)
    data = f"k_cod_postal={code_str}&k_lang=ro"
    try:
        resp = session.post(URL_COD, data=data, timeout=15)
        resp.raise_for_status()
        result = resp.json()
        if result.get("found", 0) > 0:
            return parse_results_html(result.get("formular", ""))
        return []
    except:
        return []


def export_to_excel(all_data, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Coduri Postale Iași"

    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:F1')
    ws['A1'] = 'CODURI POȘTALE IAȘI — POȘTA ROMÂNĂ 2026'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Range: {CODE_START}-{CODE_END} | cautare_cod.php | {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    for col, h in enumerate(['Nr.', 'Cod Poștal', 'Județ', 'Localitate', 'Strada / Adresa', 'Subunitate Poștală'], 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    all_data.sort(key=lambda x: (x['cod_postal'], x['strada']))
    for i, rd in enumerate(all_data):
        row = i + 5
        for col, val in enumerate([i+1, rd['cod_postal'], rd['judet'], rd['localitate'], rd['strada'], rd['subunitate_postala']], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10, bold=(col == 2))
            c.alignment = Alignment(horizontal='center') if col in (
                1, 2) else Alignment()
            c.border = tb

    for col, w in {'A': 7, 'B': 12, 'C': 10, 'D': 22, 'E': 55, 'F': 25}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:F{4 + len(all_data)}'
    wb.save(filename)


def main():
    sd = os.path.dirname(os.path.abspath(__file__))
    progress_file = os.path.join(sd, "scan_progress_iasi.json")

    print("=" * 65)
    print("  CODURI POSTALE IAȘI — POȘTA ROMÂNĂ (scan cod cu cod)")
    print(f"  Range: {CODE_START} → {CODE_END} ({TOTAL_CODES} coduri)")
    print(f"  Delay: {DELAY}s → ~{int(TOTAL_CODES * DELAY / 60)} min")
    print(f"  {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print("=" * 65)

    print("\n[1] Conectare...")
    session = init_session()
    print(
        f"    Cookie: {dict(session.cookies).get('PHPSESSID', 'N/A')[:12]}...")

    print("[2] Test cod 700100...")
    test = search_code(session, 700100)
    if test:
        print(f"    ✅ {len(test)} rezultate! Ex: {test[0]['strada']}")
    else:
        print("    ❌ Test eșuat. Oprire.")
        return

    # Resume?
    all_data = []
    existing_keys = set()
    resume_from = CODE_START

    if os.path.exists(progress_file):
        try:
            with open(progress_file, 'r', encoding='utf-8') as f:
                progress = json.load(f)
            all_data = progress.get('data', [])
            resume_from = progress.get('last_code', CODE_START) + 1
            existing_keys = set(
                f"{r['cod_postal']}|{r['strada']}" for r in all_data)
            print(
                f"\n    📋 Resume de la {resume_from} ({len(all_data)} deja găsite)")
        except:
            pass

    print(f"\n[3] Scanare {resume_from} → {CODE_END}...\n")
    start_time = time.time()

    try:
        for code in range(resume_from, CODE_END + 1):
            code_str = str(code).zfill(6)
            scanned = code - CODE_START + 1

            results = search_code(session, code)

            if results:
                new = 0
                for r in results:
                    key = f"{r['cod_postal']}|{r['strada']}"
                    if key not in existing_keys:
                        all_data.append(r)
                        existing_keys.add(key)
                        new += 1
                if new > 0:
                    print(f"  ✅ {code_str}: +{new} (total {len(all_data)})")

            if scanned % 200 == 0:
                pct = scanned / TOTAL_CODES * 100
                eta = (CODE_END - code) * DELAY / 60
                print(
                    f"  [{scanned}/{TOTAL_CODES}] {pct:.1f}% | {code_str} | {len(all_data)} găsite | ~{eta:.0f}min rămas")
                with open(progress_file, 'w', encoding='utf-8') as f:
                    json.dump({'last_code': code, 'data': all_data},
                              f, ensure_ascii=False)

            time.sleep(DELAY)

            if scanned % 2000 == 0:
                print(f"  🔄 Re-init sesiune...")
                session = init_session()

    except KeyboardInterrupt:
        print(
            f"\n\n⏸️  Întrerupt! Salvez progres ({len(all_data)} rezultate)...")
        with open(progress_file, 'w', encoding='utf-8') as f:
            json.dump({'last_code': code, 'data': all_data},
                      f, ensure_ascii=False)
        print(f"  Rulează din nou pentru a continua de la {code}.")

    if os.path.exists(progress_file) and code >= CODE_END:
        os.remove(progress_file)

    unique = {}
    for r in all_data:
        unique[f"{r['cod_postal']}|{r['strada']}"] = r
    all_data = list(unique.values())

    elapsed = time.time() - start_time
    print(f"\n{'=' * 65}")
    print(
        f"  REZULTAT: {len(all_data)} intrări | {len(set(r['cod_postal'] for r in all_data))} coduri unice | {len(set(r['localitate'] for r in all_data))} localități")
    print(f"  Timp: {elapsed/60:.1f} min")
    print(f"{'=' * 65}")

    xlsx = os.path.join(sd, "coduri_postale_iasi_posta_romana_2026.xlsx")
    export_to_excel(all_data, xlsx)
    print(f"\n✅ Excel: {xlsx}")

    csv_f = os.path.join(sd, "coduri_postale_iasi_posta_romana_2026.csv")
    with open(csv_f, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=[
                           'cod_postal', 'judet', 'localitate', 'strada', 'subunitate_postala'])
        w.writeheader()
        w.writerows(all_data)
    print(f"✅ CSV: {csv_f}")

    with open(os.path.join(sd, "coduri_postale_iasi_posta_romana_2026.json"), 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON salvat")
    print(f"\n🎉 GATA!")


if __name__ == "__main__":
    main()

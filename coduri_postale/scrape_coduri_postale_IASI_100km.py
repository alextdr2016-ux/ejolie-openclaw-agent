#!/usr/bin/env python3
"""
scrape_coduri_postale_IASI_100km.py
Extrage TOATE codurile postale pentru Iași + raza 100km de pe codul-postal.ro

Județe acoperite (raza 100km de la Iași):
  - Iași (complet)
  - Vaslui (complet) 
  - Botoșani (~90%)
  - Suceava (~40-50% partea de est)
  - Neamț (~50-60% partea de est)
  - Bacău (~20-30% partea de nord)

Strategie: extragem TOATE județele complet (e mai simplu și mai corect
           decât să calculăm distanțe GPS per localitate)

Timp estimat: ~5-8 minute (6 județe × ~300-500 localități fiecare)
"""

import requests
import json
import time
import re
import os
import csv
from datetime import datetime

BASE_URL = "https://www.codul-postal.ro"
DELAY = 0.25

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "ro-RO,ro;q=0.9",
}

# Județele din raza de 100km de la Iași
JUDETE = [
    {"name": "Iași",     "slug": "iasi",     "raza": "complet"},
    {"name": "Vaslui",   "slug": "vaslui",   "raza": "complet"},
    {"name": "Botoșani", "slug": "botosani", "raza": "~90%"},
    {"name": "Suceava",  "slug": "suceava",  "raza": "~40-50%"},
    {"name": "Neamț",    "slug": "neamt",    "raza": "~50-60%"},
    {"name": "Bacău",    "slug": "bacau",    "raza": "~20-30%"},
]


def get_localities(session, judet_slug):
    """Extrage localitățile dintr-un județ."""
    resp = session.get(f"{BASE_URL}/judet/{judet_slug}", timeout=15)
    resp.raise_for_status()
    pattern = r'href="/judet/' + \
        re.escape(judet_slug) + \
        r'/([^"]+)"[^>]*>.*?<span class="loc-name">([^<]+)</span>'
    matches = re.findall(pattern, resp.text, re.DOTALL)
    seen = set()
    locs = []
    for slug, name in matches:
        slug = slug.strip('/')
        name = name.strip()
        if slug and slug not in seen:
            seen.add(slug)
            locs.append({
                'name': name,
                'slug': slug,
                'url': f"{BASE_URL}/judet/{judet_slug}/{slug}",
            })
    return locs


def scrape_locality(session, loc, judet_name):
    """Extrage coduri din tabelul HTML al localității."""
    try:
        resp = session.get(loc['url'], timeout=30)
        resp.raise_for_status()
        html = resp.text
    except Exception as e:
        print(f"    [ERR] {loc['name']}: {e}")
        return []

    results = []

    # Metoda 1: Tabel 3 coloane (orașe mari cu străzi)
    rows = re.findall(
        r'<tr[^>]*>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*</tr>',
        html, re.DOTALL
    )
    for c1, c2, c3 in rows:
        strada = re.sub(r'<[^>]+>', '', c1).strip()
        numere = re.sub(r'<[^>]+>', '', c2).strip()
        cod = re.sub(r'<[^>]+>', '', c3).strip()
        if cod and cod.isdigit() and len(cod) == 6:
            results.append({
                'cod_postal': cod,
                'judet': judet_name,
                'localitate': loc['name'],
                'strada': f"{strada} {numere}".strip() if numere else strada,
                'numere': numere,
            })

    # Metoda 2: Sate mici - extrage DOAR codul principal
    if not results:
        # Prioritate 1: meta tag
        meta = re.search(
            r'<meta\s+name="fact:postal-code"\s+content="[^"]*?(\d{6})', html)
        if meta:
            results.append({
                'cod_postal': meta.group(1), 'judet': judet_name,
                'localitate': loc['name'], 'strada': '', 'numere': '',
            })
        else:
            # Prioritate 2: titlu
            title = re.search(r'<title>[^<]*?(\d{6})[^<]*</title>', html)
            if title:
                results.append({
                    'cod_postal': title.group(1), 'judet': judet_name,
                    'localitate': loc['name'], 'strada': '', 'numere': '',
                })
            else:
                # Prioritate 3: code-card
                card = re.search(
                    r'class="[^"]*code-card[^"]*"[^>]*>.*?(\d{6})', html, re.DOTALL)
                if card:
                    results.append({
                        'cod_postal': card.group(1), 'judet': judet_name,
                        'localitate': loc['name'], 'strada': '', 'numere': '',
                    })

    return results


def export_excel(data_per_judet, all_data, filename):
    """Exportă în Excel cu sheet per județ + sheet total."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    hdrs = ['Nr.', 'Cod Poștal', 'Județ',
            'Localitate', 'Strada', 'Numere/Blocuri']

    # Culori per județ
    judet_colors = {
        'Iași': 'D6E4F0',
        'Vaslui': 'E2EFDA',
        'Botoșani': 'FFF2CC',
        'Suceava': 'FCE4D6',
        'Neamț': 'EDEDED',
        'Bacău': 'E4DFEC',
    }

    def write_sheet(ws, title, data):
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = Font(name='Arial', bold=True, size=13, color='2F5496')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:F2')
        ws['A2'] = f'Sursa: codul-postal.ro | {datetime.now().strftime("%d.%m.%Y %H:%M")}'
        ws['A2'].font = Font(name='Arial', size=10,
                             italic=True, color='666666')
        ws['A2'].alignment = Alignment(horizontal='center')
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb
        data.sort(key=lambda x: (x['localitate'],
                  x['strada'], x['cod_postal']))
        for i, rd in enumerate(data):
            row = i + 5
            fill_color = judet_colors.get(rd['judet'], 'FFFFFF')
            fill = PatternFill('solid', fgColor=fill_color)
            for col, val in enumerate([i+1, rd['cod_postal'], rd['judet'], rd['localitate'], rd['strada'], rd.get('numere', '')], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name='Arial', size=10, bold=(col == 2))
                c.alignment = Alignment(horizontal='center') if col in (
                    1, 2) else Alignment()
                c.border = tb
                c.fill = fill
        for col, w in {'A': 7, 'B': 12, 'C': 12, 'D': 22, 'E': 45, 'F': 30}.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A5'
        ws.auto_filter.ref = f'A4:F{4+len(data)}'

    # Sheet 1: TOTAL toate județele
    ws_all = wb.active
    ws_all.title = "Total 6 Județe"
    write_sheet(
        ws_all, 'CODURI POȘTALE — IAȘI + 100km RAZĂ (6 JUDEȚE)', all_data)

    # Sheet per județ
    for judet_info in JUDETE:
        jname = judet_info['name']
        jdata = data_per_judet.get(jname, [])
        if jdata:
            ws = wb.create_sheet(jname)
            write_sheet(
                ws, f'CODURI POȘTALE — JUDEȚUL {jname.upper()} ({judet_info["raza"]})', jdata)

    # Sheet sumar
    ws_sum = wb.create_sheet("Sumar")
    ws_sum['A1'] = 'SUMAR PE JUDEȚE ȘI LOCALITĂȚI'
    ws_sum['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')

    # Sumar pe județe
    ws_sum['A3'] = 'Județ'
    ws_sum['B3'] = 'Localități'
    ws_sum['C3'] = 'Intrări'
    ws_sum['D3'] = 'Coduri unice'
    ws_sum['E3'] = 'Raza 100km'
    for col in range(1, 6):
        ws_sum.cell(row=3, column=col).font = hf
        ws_sum.cell(row=3, column=col).fill = hfill
        ws_sum.cell(row=3, column=col).border = tb

    row = 4
    total_locs = total_entries = total_codes = 0
    for ji in JUDETE:
        jdata = data_per_judet.get(ji['name'], [])
        nlocs = len(set(r['localitate'] for r in jdata))
        ncodes = len(set(r['cod_postal'] for r in jdata))
        fill = PatternFill(
            'solid', fgColor=judet_colors.get(ji['name'], 'FFFFFF'))
        ws_sum.cell(row=row, column=1, value=ji['name']).border = tb
        ws_sum.cell(row=row, column=1).fill = fill
        ws_sum.cell(row=row, column=2, value=nlocs).border = tb
        ws_sum.cell(row=row, column=2).alignment = Alignment(
            horizontal='center')
        ws_sum.cell(row=row, column=3, value=len(jdata)).border = tb
        ws_sum.cell(row=row, column=3).alignment = Alignment(
            horizontal='center')
        ws_sum.cell(row=row, column=4, value=ncodes).border = tb
        ws_sum.cell(row=row, column=4).alignment = Alignment(
            horizontal='center')
        ws_sum.cell(row=row, column=5, value=ji['raza']).border = tb
        total_locs += nlocs
        total_entries += len(jdata)
        total_codes += ncodes
        row += 1

    # Total row
    bf = Font(name='Arial', bold=True, size=11)
    ws_sum.cell(row=row, column=1, value='TOTAL').font = bf
    ws_sum.cell(row=row, column=1).border = tb
    ws_sum.cell(row=row, column=2, value=total_locs).font = bf
    ws_sum.cell(row=row, column=2).border = tb
    ws_sum.cell(row=row, column=3, value=total_entries).font = bf
    ws_sum.cell(row=row, column=3).border = tb
    ws_sum.cell(row=row, column=4, value=total_codes).font = bf
    ws_sum.cell(row=row, column=4).border = tb

    for col, w in {'A': 15, 'B': 12, 'C': 10, 'D': 14, 'E': 14}.items():
        ws_sum.column_dimensions[col].width = w

    wb.save(filename)


def main():
    sd = os.path.dirname(os.path.abspath(__file__))

    print("=" * 65)
    print("  CODURI POSTALE IAȘI + 100km RAZĂ")
    print(f"  Județe: {', '.join(j['name'] for j in JUDETE)}")
    print(f"  Sursa: codul-postal.ro")
    print(f"  {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print("=" * 65)

    s = requests.Session()
    s.headers.update(HEADERS)

    all_data = []
    data_per_judet = {}

    for ji in JUDETE:
        jname = ji['name']
        jslug = ji['slug']

        print(f"\n{'='*50}")
        print(f"  JUDEȚUL {jname.upper()} ({ji['raza']})")
        print(f"{'='*50}")

        print(f"  Extrag localități...")
        locs = get_localities(s, jslug)
        print(f"  {len(locs)} localități găsite")

        if not locs:
            print(f"  ❌ Eroare la {jname}!")
            continue

        judet_data = []
        for i, loc in enumerate(locs):
            results = scrape_locality(s, loc, jname)
            judet_data.extend(results)

            if results and len(results) > 1:
                print(f"    [{i+1}/{len(locs)}] {loc['name']}: {len(results)}")

            time.sleep(DELAY)

        # Dedup per județ
        u = {}
        for r in judet_data:
            u[f"{r['cod_postal']}|{r['localitate']}|{r['strada']}"] = r
        judet_data = list(u.values())

        data_per_judet[jname] = judet_data
        all_data.extend(judet_data)

        nlocs = len(set(r['localitate'] for r in judet_data))
        ncodes = len(set(r['cod_postal'] for r in judet_data))
        print(
            f"\n  ✅ {jname}: {len(judet_data)} intrări | {ncodes} coduri | {nlocs} localități")

    # Dedup total
    u = {}
    for r in all_data:
        u[f"{r['cod_postal']}|{r['judet']}|{r['localitate']}|{r['strada']}"] = r
    all_data = list(u.values())

    total_codes = len(set(r['cod_postal'] for r in all_data))
    total_locs = len(set(r['localitate'] for r in all_data))

    print(f"\n{'='*65}")
    print(f"  REZULTAT FINAL")
    print(
        f"  Total: {len(all_data)} intrări | {total_codes} coduri | {total_locs} localități")
    for ji in JUDETE:
        jd = data_per_judet.get(ji['name'], [])
        nc = len(set(r['cod_postal'] for r in jd))
        print(f"    {ji['name']:12s}: {len(jd):5d} intrări | {nc:4d} coduri")
    print(f"{'='*65}")

    # Export
    xlsx = os.path.join(sd, "coduri_postale_IASI_100km_2026.xlsx")
    export_excel(data_per_judet, all_data, xlsx)
    print(f"\n✅ Excel: {xlsx}")

    csv_f = os.path.join(sd, "coduri_postale_IASI_100km_2026.csv")
    with open(csv_f, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(
            f, fieldnames=['cod_postal', 'judet', 'localitate', 'strada', 'numere'])
        w.writeheader()
        w.writerows(sorted(all_data, key=lambda x: (
            x['judet'], x['localitate'], x['cod_postal'])))
    print(f"✅ CSV: {csv_f}")

    with open(os.path.join(sd, "coduri_postale_IASI_100km_2026.json"), 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON salvat")

    print(f"\n🎉 GATA!")


if __name__ == "__main__":
    main()

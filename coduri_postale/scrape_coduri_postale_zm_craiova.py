#!/usr/bin/env python3
"""
Script: scrape_coduri_postale_zm_craiova.py  v2
Descriere: Extrage codurile postale pe strazi de pe site-ul Postei Romane
           pentru toate localitatile din Zona Metropolitana Craiova.
           
Sursa: https://www.posta-romana.ro/ccp.html
API Endpoints (reverse-engineered):
  - POST /cnpr-app/modules/cauta-cod-postal/ajax/cauta_orase.php?q=
    Body: k_judet=Dolj&k_lang=ro
  - POST /cnpr-app/modules/cauta-cod-postal/ajax/cautare_pentru_cod.php?q=
    Body: k_adresa=X&k_judet=Dolj&k_localitate=Craiova&k_lang=ro
    
IMPORTANT: API-ul necesita cuvinte reale (nu litere singure).
Strategie: cautam cu lista de nume de strazi/bulevarde cunoscute.

v2: Fix parametri k_judet/k_localitate/k_adresa/k_lang + strategie cuvinte
"""

import requests
import json
import time
import re
import os
from datetime import datetime
from urllib.parse import quote

BASE_URL = "https://www.posta-romana.ro"
URL_LOCALITATI = f"{BASE_URL}/cnpr-app/modules/cauta-cod-postal/ajax/cauta_orase.php?q="
URL_CAUTARE = f"{BASE_URL}/cnpr-app/modules/cauta-cod-postal/ajax/cautare_pentru_cod.php?q="
JUDET = "Dolj"
LOCALITATI_CU_STRAZI = ["Craiova", "Filiași", "Segarcea"]

LOCALITATI_ZM = [
    "Craiova", "Filiași", "Segarcea",
    "Almăj", "Beharca", "Bogea", "Cotofenii din Față", "Moșneni", "Șitoaia",
    "Brădești", "Brădeștii Bătrâni", "Meteu", "Piscani", "Răcarii de Jos", "Tatomirești",
    "Breasta", "Cotu", "Crovna", "Făget", "Obedin", "Roșieni", "Valea Lungului",
    "Bucovăț", "Cârligei", "Italieni", "Leamna de Jos", "Leamna de Sus", "Palilula", "Sărbătoarea",
    "Calopăr", "Bâzdâna", "Belcinu", "Panaghia", "Sălcuța",
    "Cârcea", "Coșoveni",
    "Ghercești", "Gârlești", "Luncșoru", "Ungureni", "Ungurenii Mici",
    "Işalnița", "Izvoare",
    "Malu Mare", "Ghindeni", "Preajba",
    "Mischii", "Călinești", "Gogoșești", "Mlecănești", "Motoci", "Urechești",
    "Murgași", "Gaia", "Picăturile", "Rupturile", "Velești",
    "Pielești", "Câmpeni", "Lânga",
    "Predești", "Bucicani", "Cârstovani", "Frasin", "Milovan", "Pleșoi", "Predeștii Mici",
    "Șimnicu de Sus", "Albești", "Cornetu", "Deleni", "Dudovicești", "Florești",
    "Izvor", "Jieni", "Leșile", "Milești", "Românești",
    "Teasc", "Secui",
    "Terpezița", "Căciulatu", "Căruia", "Floran", "Lazu",
    "Țuglui", "Jiul", "Unirea",
    "Vârvoru de Jos", "Bujor", "Ciutura", "Criva", "Dobromira", "Drăgoaia", "Gabru", "Vârvor",
    "Vela", "Bucovicior", "Cetățuia", "Desnățui", "Gubaucea", "Segleț", "Suharu", "Știubei",
]

# ~200 cuvinte de cautare care acopera majoritatea strazilor din Craiova
CUVINTE_CAUTARE_CRAIOVA = [
    "Calea Bucuresti", "Calea Severinului", "Calea Unirii", "Dacia",
    "Decebal", "Nicolae Titulescu", "Stirbei Voda", "Carol",
    "1 Mai", "Oltenia", "Romanescu",
    "Albinelor", "Alecsandri", "Alexandru", "Amaradia", "Amurgului", "Aries",
    "Arges", "Armata", "Avram Iancu", "Aviatorilor",
    "Bariera Valcii", "Bechetului", "Bibescu", "Borsec", "Bradului", "Brazda",
    "Brestei", "Bucegi", "Bucovat", "Buftea", "Buzaului", "Barbu",
    "Calarasi", "Calafat", "Calugareni", "Cameliei", "Campeni", "Caracal",
    "Carpati", "Castanilor", "Cernei", "Cimentului", "Ciocarliei", "Closca",
    "Cornitoiu", "Cosuna", "Craiovesti", "Craiovita", "Crangului", "Crisan",
    "Croitori", "Cuza Voda",
    "Danubiu", "Desnatu", "Deva", "Dobrogea", "Dorobanti", "Dunarii",
    "Dimitrie", "Dolj",
    "Ecaterina", "Elena", "Emil", "Energiei", "Eroilor",
    "Fabricii", "Fagului", "Filantropiei", "Florilor", "Frasinului", "Fulger",
    "Garlesti", "Garofitei", "George", "Gheorghe", "Girlesti", "Gorjului", "Gradinari",
    "Haiducilor", "Henri Coanda", "Horea", "Hortensiei", "Hunedoara", "Hydepark",
    "Iancu Jianu", "Iederei", "Independentei", "Industriei", "Ion", "Izvorul Rece",
    "Jianu", "Jiului", "Jiu",
    "Lacului", "Lalelei", "Lapus", "Libertatii", "Liliacului",
    "Lipscani", "Livezilor", "Locomotivei", "Lungulescu",
    "Madona Dudu", "Malinului", "Manzatului", "Maria", "Marinescu",
    "Matasarilor", "Mehedinti", "Meseriasilor", "Mihai Viteazu", "Minerva",
    "Mircea", "Mitropolit", "Moldovei", "Muresului", "Muncii",
    "Nanterre", "Narciselor", "Negru Voda", "Nicolae Balcescu", "Nicolae Iorga", "Nordului", "Nucului",
    "Obedeanu", "Oltet", "Olt", "Operelor", "Orhideelor", "Ostroveni",
    "Pacii", "Padurii", "Parangului", "Parului", "Pasteur", "Pelendava",
    "Petru Rares", "Piata", "Pietii", "Plaiului", "Podului", "Popa Sapca",
    "Popoveni", "Portului", "Potelu", "Prelungirea", "Primaverii", "Progresul", "Putnei", "Putul Rece",
    "Rahovei", "Razboieni", "Recoltei", "Renasterii", "Republicii",
    "Riului", "Rocadei", "Romanesti", "Romania Muncitoare", "Romul",
    "Rovinari", "Rovine",
    "Sacelu", "Sandulescu", "Sararilor", "Savinesti", "Severin",
    "Siretului", "Slatina", "Smeurei", "Socului", "Spahii",
    "Stefan cel Mare", "Stirbei", "Sudului",
    "Tabaci", "Teilor", "Tepes Voda", "Tineretului", "Tismana",
    "Toamnei", "Traian", "Trandafirilor", "Turbinei",
    "Ulmului", "Unirii", "Uranus",
    "Vailor", "Vasile", "Verde", "Viitorului", "Vlad Tepes", "Vulturi",
    "Zambilei", "Zarandului", "Zorilor",
    "Brazda lui Novac", "Craiovita Noua", "Lapus Arges",
    "Valea Rosie", "Sarari", "Ghiseu",
    # Extra - cartiere si zone
    "Balta Verde", "Cernele", "Facai", "Mofleni", "Popoveni", "Simnicu",
    "Barcanesti", "Cioroiasu", "Isalnita", "Preajba",
]

CUVINTE_CAUTARE_ORAS_MIC = [
    "Calea", "Unirii", "Libertatii", "Republicii",
    "Garii", "Stefan", "Nicolae", "Ion", "Mihai",
    "Independentei", "Victoriei", "Eroilor", "Tineretului",
    "Alexandru", "Carol", "Dunarii", "Jiului", "Oltului",
    "Traian", "Decebal", "Avram Iancu",
]

DELAY = 0.4
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "ro-RO,ro;q=0.9,en;q=0.8",
    "Referer": "https://www.posta-romana.ro/ccp.html",
    "X-Requested-With": "XMLHttpRequest",
    "Origin": "https://www.posta-romana.ro",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
}


def parse_results_html(html_str):
    results = []
    p_tags = re.findall(r'<p[^>]*>(.*?)</p>', html_str)
    for i in range(0, len(p_tags), 5):
        chunk = p_tags[i:i+5]
        if len(chunk) >= 4:
            cod = chunk[0].strip()
            if cod and cod[0].isdigit() and len(cod) == 6:
                results.append({
                    "cod_postal": cod,
                    "judet": chunk[1].strip(),
                    "localitate": chunk[2].strip(),
                    "strada": chunk[3].strip(),
                    "subunitate_postala": chunk[4].strip() if len(chunk) > 4 else "",
                })
    return results


def cauta_coduri(judet, localitate, adresa, session):
    data = f"k_adresa={quote(adresa)}&k_judet={quote(judet)}&k_localitate={quote(localitate)}&k_lang=ro"
    try:
        resp = session.post(URL_CAUTARE, data=data, timeout=30)
        resp.raise_for_status()
        result = resp.json()
        if result.get("found", 0) > 0:
            return parse_results_html(result.get("formular", ""))
        return []
    except Exception as e:
        print(f"  [EROARE] {localitate}/{adresa}: {e}")
        return []


def scrape_localitate(judet, localitate, session, cuvinte=None):
    all_results = {}

    # Cautare fara adresa
    results = cauta_coduri(judet, localitate, "", session)
    for r in results:
        all_results[f"{r['cod_postal']}|{r['strada']}"] = r
    time.sleep(DELAY)

    if cuvinte:
        total = len(cuvinte)
        for idx, cuvant in enumerate(cuvinte):
            results = cauta_coduri(judet, localitate, cuvant, session)
            new = sum(
                1 for r in results if f"{r['cod_postal']}|{r['strada']}" not in all_results)
            for r in results:
                all_results[f"{r['cod_postal']}|{r['strada']}"] = r

            if results:
                print(
                    f"    [{idx+1}/{total}] '{cuvant}': {len(results)} gasite, {new} noi → total {len(all_results)}")
            time.sleep(DELAY)

    return list(all_results.values())


def export_to_excel(all_data, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        os.system("pip3 install openpyxl --break-system-packages -q")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Coduri Postale ZM Craiova"

    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    df = Font(name='Arial', size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:F1')
    ws['A1'] = 'CODURI POȘTALE PE STRĂZI - ZONA METROPOLITANĂ CRAIOVA'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Sursa: Poșta Română | Generat: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    for col, h in enumerate(['Nr.', 'Cod Poștal', 'Județ', 'Localitate', 'Strada / Adresa', 'Subunitate Poștală'], 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    all_data.sort(key=lambda x: (x['localitate'], x['strada']))
    for i, rd in enumerate(all_data):
        row = i + 5
        for col, val in enumerate([i+1, rd['cod_postal'], rd['judet'], rd['localitate'], rd['strada'], rd['subunitate_postala']], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10, bold=(col == 2))
            c.alignment = Alignment(horizontal='center') if col in (
                1, 2) else Alignment()
            c.border = tb

    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 30
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:F{4 + len(all_data)}'
    wb.save(filename)
    print(f"\n✅ Excel: {filename} ({len(all_data)} randuri)")


def main():
    print("=" * 60)
    print("SCRAPER CODURI POSTALE - ZONA METROPOLITANA CRAIOVA  v2")
    print(f"Sursa: Posta Romana | {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print("=" * 60)

    session = requests.Session()
    session.headers.update(HEADERS)

    print("\n[1] Conectare...")
    session.get(f"{BASE_URL}/ccp.html", timeout=10)
    session.post(URL_LOCALITATI, data=f"k_judet={JUDET}&k_lang=ro", timeout=10)

    print("[2] Test: 'Calea Bucuresti' in Craiova...")
    test = cauta_coduri(JUDET, "Craiova", "Calea Bucuresti", session)
    if not test:
        print("    ❌ API nu raspunde. Oprire.")
        return
    print(f"    ✅ {len(test)} rezultate. Mergem!")

    all_data = []
    for loc in LOCALITATI_ZM:
        print(f"\n{'='*40}\n[*] {loc}\n{'='*40}")
        cuvinte = CUVINTE_CAUTARE_CRAIOVA if loc == "Craiova" else (
            CUVINTE_CAUTARE_ORAS_MIC if loc in LOCALITATI_CU_STRAZI else None)
        results = scrape_localitate(JUDET, loc, session, cuvinte)
        all_data.extend(results)
        print(f"  → {len(results)} coduri")

    unique = {}
    for r in all_data:
        unique[f"{r['cod_postal']}|{r['strada']}"] = r
    all_data = list(unique.values())

    print(f"\n{'='*60}\nTOTAL: {len(all_data)} coduri postale unice\n{'='*60}")

    sd = os.path.dirname(os.path.abspath(__file__))
    export_to_excel(all_data, os.path.join(
        sd, "coduri_postale_zm_craiova_complet.xlsx"))

    import csv
    with open(os.path.join(sd, "coduri_postale_zm_craiova_complet.csv"), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=[
                           'cod_postal', 'judet', 'localitate', 'strada', 'subunitate_postala'])
        w.writeheader()
        w.writerows(all_data)
    print(f"✅ CSV salvat")

    with open(os.path.join(sd, "coduri_postale_zm_craiova_complet.json"), 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON salvat")
    print(f"\n🎉 Gata! Output in: {sd}")


if __name__ == "__main__":
    main()

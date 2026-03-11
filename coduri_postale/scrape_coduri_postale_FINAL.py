#!/usr/bin/env python3
"""
Script: scrape_coduri_postale_FINAL.py
Descriere: Extrage TOATE codurile postale din judetul Dolj de pe codul-postal.ro
           Site-ul are 37,914 coduri actualizate, listate in tabele HTML simple.

STRATEGIA GENIALA: In loc de API cu rate limiting, folosim un site care
listeaza TOTUL in tabele HTML statice — fara login, fara JS, fara API keys.

URL pattern: https://www.codul-postal.ro/judet/dolj/{localitate}
Fiecare localitate are un tabel HTML cu TOATE strazile si codurile.
Craiova = 1557 randuri pe o singura pagina!

Pas 1: Fetch pagina /judet/dolj → lista cu 387 localitati (link-uri)
Pas 2: Pentru fiecare localitate → fetch pagina → parse tabel HTML
Pas 3: Export Excel + CSV + JSON

Timp estimat: 2-3 minute (387 pagini HTML, ~0.3s fiecare)
"""

import requests
import json
import time
import re
import os
import csv
from datetime import datetime
from html.parser import HTMLParser

BASE_URL = "https://www.codul-postal.ro"
DELAY = 0.3

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ro-RO,ro;q=0.9,en;q=0.8",
}

# Localitati ZM Craiova (pentru marcare in Excel)
ZM_CRAIOVA = {
    "Craiova", "Filiaşi", "Filiași", "Segarcea",
    "Almăj", "Beharca", "Bogea", "Cotofenii din Față", "Cotofenii din Faţă", "Moşneni", "Moșneni", "Şitoaia", "Șitoaia",
    "Brădeşti", "Brădești", "Brădeştii Bătrâni", "Brădeștii Bătrâni", "Meteu", "Piscani", "Răcarii de Jos", "Tatomireşti", "Tatomirești",
    "Breasta", "Cotu", "Crovna", "Făget", "Obedin", "Roşieni", "Roșieni", "Valea Lungului",
    "Bucovăţ", "Bucovăț", "Cârligei", "Italieni", "Leamna de Jos", "Leamna de Sus", "Palilula", "Sărbătoarea",
    "Calopăr", "Bâzdâna", "Belcinu", "Panaghia", "Sălcuţa", "Sălcuța",
    "Cârcea", "Coşoveni", "Coșoveni",
    "Gherceşti", "Ghercești", "Gârleşti", "Gârlești", "Luncşoru", "Luncșoru", "Ungureni", "Ungurenii Mici",
    "Işalniţa", "Ișalnița", "Izvoare",
    "Malu Mare", "Ghindeni", "Preajba",
    "Mischii", "Călineşti", "Călinești", "Gogoşeşti", "Gogoșești", "Mlecăneşti", "Mlecănești", "Motoci", "Urecheşti", "Urechești",
    "Murgaşi", "Murgași", "Gaia", "Picăturile", "Rupturile", "Veleşti", "Velești",
    "Pieleşti", "Pielești", "Câmpeni", "Lânga",
    "Predeşti", "Predești", "Bucicani", "Cârstovani", "Frasin", "Milovan", "Pleşoi", "Pleșoi", "Predeştii Mici", "Predeștii Mici",
    "Şimnicu de Sus", "Șimnicu de Sus", "Albeşti", "Albești", "Cornetu", "Deleni", "Dudoviceşti", "Dudovicești",
    "Floreşti", "Florești", "Izvor", "Jieni", "Leşile", "Leșile", "Mileşti", "Milești", "Româneşti", "Românești",
    "Teasc", "Secui",
    "Terpeziţa", "Terpezița", "Căciulatu", "Căruia", "Floran", "Lazu",
    "Ţuglui", "Țuglui", "Jiul", "Unirea",
    "Vârvoru de Jos", "Bujor", "Ciutura", "Criva", "Dobromira", "Drăgoaia", "Gabru", "Vârvor",
    "Vela", "Bucovicior", "Cetăţuia", "Cetățuia", "Desnăţui", "Desnățui", "Gubaucea", "Segleţ", "Segleț", "Suharu", "Ştiubei", "Știubei",
    "Făcăi", "Mofleni", "Popoveni", "Şimnicu de Jos", "Șimnicu de Jos",
}


def get_localities(session):
    """Extrage lista de localitati din pagina judetului Dolj."""
    resp = session.get(f"{BASE_URL}/judet/dolj", timeout=15)
    resp.raise_for_status()
    html = resp.text

    # Find all links matching /judet/dolj/{slug}
    pattern = r'href="(/judet/dolj/[^"]+)"[^>]*>([^<]*)'
    matches = re.findall(pattern, html)

    localities = []
    seen = set()
    for path, name in matches:
        slug = path.split('/judet/dolj/')[-1].strip('/')
        if slug and slug not in seen and '/' not in slug:
            seen.add(slug)
            localities.append({
                'name': name.strip().split('\n')[0].strip(),
                'slug': slug,
                'url': f"{BASE_URL}{path}",
            })

    return localities


def scrape_locality(session, locality):
    """Extrage codurile postale dintr-o localitate (parseaza tabelul HTML)."""
    try:
        resp = session.get(locality['url'], timeout=15)
        resp.raise_for_status()
        html = resp.text
    except Exception as e:
        print(f"  [EROARE] {locality['name']}: {e}")
        return []

    results = []

    # Parse table rows: <tr><td>Strada</td><td>Numere</td><td>Cod</td></tr>
    row_pattern = r'<tr[^>]*>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*</tr>'
    rows = re.findall(row_pattern, html, re.DOTALL)

    for strada_raw, numere_raw, cod_raw in rows:
        # Clean HTML tags
        strada = re.sub(r'<[^>]+>', '', strada_raw).strip()
        numere = re.sub(r'<[^>]+>', '', numere_raw).strip()
        cod = re.sub(r'<[^>]+>', '', cod_raw).strip()

        # Skip header row
        if cod.lower() in ('cod poștal', 'cod postal', 'cod'):
            continue

        # Validate cod postal (6 digits)
        if cod and len(cod) == 6 and cod.isdigit():
            strada_full = f"{strada} {numere}".strip() if numere else strada
            results.append({
                'cod_postal': cod,
                'judet': 'Dolj',
                'localitate': locality['name'],
                'strada': strada_full,
                'numere': numere,
                'in_zm_craiova': locality['name'] in ZM_CRAIOVA,
            })

    # If no table found (small village with single code), try to find the code differently
    if not results:
        # Look for a single postal code on the page
        single_code = re.findall(
            r'<(?:span|div|p)[^>]*class="[^"]*cod[^"]*"[^>]*>(\d{6})</(?:span|div|p)>', html)
        if not single_code:
            single_code = re.findall(r'>(\d{6})</', html)
            single_code = [c for c in single_code if c.startswith('20')]

        if single_code:
            results.append({
                'cod_postal': single_code[0],
                'judet': 'Dolj',
                'localitate': locality['name'],
                'strada': '',
                'numere': '',
                'in_zm_craiova': locality['name'] in ZM_CRAIOVA,
            })

    return results


def export_to_excel(all_data, zm_data, filename):
    """Exporta in Excel cu 2 sheet-uri: ZM Craiova + Tot Dolj."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    zm_fill = PatternFill('solid', fgColor='E2EFDA')

    # Sheet 1: ZM Craiova
    ws1 = wb.active
    ws1.title = "ZM Craiova"

    ws1.merge_cells('A1:F1')
    ws1['A1'] = 'CODURI POȘTALE - ZONA METROPOLITANĂ CRAIOVA (2026)'
    ws1['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:F2')
    ws1['A2'] = f'Sursa: codul-postal.ro (date actualizate) | Generat: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws1['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws1['A2'].alignment = Alignment(horizontal='center')

    headers = ['Nr.', 'Cod Poștal', 'Județ',
               'Localitate', 'Strada', 'Numere/Blocuri']
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    zm_data.sort(key=lambda x: (x['localitate'], x['strada'], x['cod_postal']))
    for i, rd in enumerate(zm_data):
        row = i + 5
        vals = [i+1, rd['cod_postal'], rd['judet'],
                rd['localitate'], rd['strada'], rd.get('numere', '')]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10, bold=(col == 2))
            c.alignment = Alignment(horizontal='center') if col in (
                1, 2) else Alignment()
            c.border = tb

    for col, w in {'A': 7, 'B': 12, 'C': 10, 'D': 22, 'E': 45, 'F': 35}.items():
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = 'A5'
    ws1.auto_filter.ref = f'A4:F{4 + len(zm_data)}'

    # Sheet 2: Tot Dolj
    ws2 = wb.create_sheet("Tot Dolj")
    ws2.merge_cells('A1:F1')
    ws2['A1'] = 'CODURI POȘTALE - ÎNTREG JUDEȚUL DOLJ (2026)'
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws2['A1'].alignment = Alignment(horizontal='center')

    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    all_data.sort(key=lambda x: (
        x['localitate'], x['strada'], x['cod_postal']))
    for i, rd in enumerate(all_data):
        row = i + 4
        vals = [i+1, rd['cod_postal'], rd['judet'],
                rd['localitate'], rd['strada'], rd.get('numere', '')]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10, bold=(col == 2))
            c.alignment = Alignment(horizontal='center') if col in (
                1, 2) else Alignment()
            c.border = tb
            if rd.get('in_zm_craiova'):
                c.fill = zm_fill

    for col, w in {'A': 7, 'B': 12, 'C': 10, 'D': 22, 'E': 45, 'F': 35}.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = 'A4'
    ws2.auto_filter.ref = f'A3:F{3 + len(all_data)}'

    # Sheet 3: Sumar
    ws3 = wb.create_sheet("Sumar")
    ws3['A1'] = 'SUMAR PE LOCALITĂȚI'
    ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')

    ws3['A3'] = 'Localitate'
    ws3['B3'] = 'Nr. Coduri'
    ws3['C3'] = 'ZM Craiova'
    for col in range(1, 4):
        ws3.cell(row=3, column=col).font = hf
        ws3.cell(row=3, column=col).fill = hfill
        ws3.cell(row=3, column=col).border = tb

    from collections import Counter
    loc_counts = Counter(r['localitate'] for r in all_data)
    for i, (loc, cnt) in enumerate(sorted(loc_counts.items()), 4):
        ws3.cell(row=i, column=1, value=loc).border = tb
        ws3.cell(row=i, column=2, value=cnt).border = tb
        ws3.cell(row=i, column=2).alignment = Alignment(horizontal='center')
        is_zm = 'DA' if loc in ZM_CRAIOVA else ''
        ws3.cell(row=i, column=3, value=is_zm).border = tb
        ws3.cell(row=i, column=3).alignment = Alignment(horizontal='center')
        if is_zm:
            for col in range(1, 4):
                ws3.cell(row=i, column=col).fill = zm_fill

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 14

    wb.save(filename)


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    print("=" * 65)
    print("  CODURI POSTALE DOLJ - EXTRAGERE COMPLETA")
    print(f"  Sursa: codul-postal.ro (37,914 coduri actualizate)")
    print(f"  Metoda: parse HTML tables (fara API, fara rate limit)")
    print(f"  {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print("=" * 65)

    session = requests.Session()
    session.headers.update(HEADERS)

    # Pas 1: Lista localitati
    print("\n[1] Extrag lista localitati Dolj...")
    localities = get_localities(session)
    print(f"    ✅ {len(localities)} localitati gasite")

    if not localities:
        print("    ❌ Nu am gasit localitati. Oprire.")
        return

    # Pas 2: Scrape fiecare localitate
    print(f"\n[2] Extrag coduri postale pentru fiecare localitate...\n")
    all_data = []

    for i, loc in enumerate(localities):
        results = scrape_locality(session, loc)
        all_data.extend(results)

        zm = " [ZM]" if loc['name'] in ZM_CRAIOVA else ""
        if results:
            print(
                f"  [{i+1}/{len(localities)}] {loc['name']}{zm}: {len(results)} coduri")

        time.sleep(DELAY)

    # Deduplicate
    unique = {}
    for r in all_data:
        key = f"{r['cod_postal']}|{r['localitate']}|{r['strada']}"
        unique[key] = r
    all_data = list(unique.values())

    # Separate ZM Craiova
    zm_data = [r for r in all_data if r.get('in_zm_craiova')]

    codes_total = len(set(r['cod_postal'] for r in all_data))
    codes_zm = len(set(r['cod_postal'] for r in zm_data))
    locs_total = len(set(r['localitate'] for r in all_data))
    locs_zm = len(set(r['localitate'] for r in zm_data))

    print(f"\n{'=' * 65}")
    print(f"  REZULTAT FINAL")
    print(
        f"  Tot Dolj:    {len(all_data)} intrari | {codes_total} coduri unice | {locs_total} localitati")
    print(
        f"  ZM Craiova:  {len(zm_data)} intrari | {codes_zm} coduri unice | {locs_zm} localitati")
    print(f"{'=' * 65}")

    # Export Excel
    xlsx = os.path.join(script_dir, "coduri_postale_dolj_COMPLET_2026.xlsx")
    export_to_excel(all_data, zm_data, xlsx)
    print(f"\n✅ Excel: {xlsx}")

    # Export CSV (ZM only)
    csv_zm = os.path.join(script_dir, "coduri_postale_zm_craiova_2026.csv")
    with open(csv_zm, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(
            f, fieldnames=['cod_postal', 'judet', 'localitate', 'strada', 'numere'])
        w.writeheader()
        for r in sorted(zm_data, key=lambda x: (x['localitate'], x['cod_postal'])):
            w.writerow(
                {k: r[k] for k in ['cod_postal', 'judet', 'localitate', 'strada', 'numere']})
    print(f"✅ CSV ZM: {csv_zm}")

    # Export CSV (tot Dolj)
    csv_all = os.path.join(script_dir, "coduri_postale_dolj_2026.csv")
    with open(csv_all, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=[
                           'cod_postal', 'judet', 'localitate', 'strada', 'numere', 'in_zm_craiova'])
        w.writeheader()
        for r in sorted(all_data, key=lambda x: (x['localitate'], x['cod_postal'])):
            w.writerow({k: r[k] for k in ['cod_postal', 'judet',
                       'localitate', 'strada', 'numere', 'in_zm_craiova']})
    print(f"✅ CSV Dolj: {csv_all}")

    # Export JSON
    json_f = os.path.join(script_dir, "coduri_postale_dolj_2026.json")
    with open(json_f, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON: {json_f}")

    print(f"\n🎉 COMPLET! Toate fisierele in: {script_dir}")


if __name__ == "__main__":
    main()

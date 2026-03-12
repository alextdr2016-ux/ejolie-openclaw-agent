#!/usr/bin/env python3
"""
scrape_coduri_postale_IASI.py
Extrage TOATE codurile postale din județul Iași de pe codul-postal.ro
+ marchează localitățile din Zona Metropolitană Iași (28 membri)

Același approach ca scriptul pentru Dolj (FINAL v4):
  - Fetch pagina /judet/iasi → lista localități
  - Pentru fiecare → parse tabel HTML → extrage cod poștal
  - Fix fallback: doar codul principal, nu "localități apropiate"

Apoi merge cu scanul Poșta Română (dacă e disponibil) pentru completare.

ZM Iași (28 membri, actualizat 2023):
  Municipiul Iași + 27 comune:
  Aroneanu, Bârnova, Ciurea, Comarna, Costuleni, Dobrovăț, Golăiești,
  Grajduri, Holboca, Horlești, Lețcani, Miroslava, Mogoșești, Movileni,
  Popricani, Prisăcani, Rediu, Românești, Scânteia, Schitu Duca,
  Tomești, Țigănași, Țuțora, Ungheni, Valea Lupului, Victoria, Voinești
"""

import requests
import json
import time
import re
import os
import csv
from datetime import datetime

BASE_URL = "https://www.codul-postal.ro"
DELAY = 0.3
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "ro-RO,ro;q=0.9",
}

# ZM Iași - 28 membri (cu variante diacritice)
ZM_IASI = {
    # Municipiul
    "Iaşi", "Iași",
    # 27 comune
    "Aroneanu",
    "Bârnova", "Bîrnova",
    "Ciurea",
    "Comarna",
    "Costuleni",
    "Dobrovăţ", "Dobrovăț",
    "Golăieşti", "Golăiești",
    "Grajduri",
    "Holboca",
    "Horleşti", "Horlești",
    "Leţcani", "Lețcani",
    "Miroslava",
    "Mogoşeşti", "Mogoșești",
    "Movileni",
    "Popricani",
    "Prisăcani",
    "Rediu",
    "Româneşti", "Românești",
    "Scânteia", "Scînteia",
    "Schitu Duca",
    "Tomeşti", "Tomești",
    "Ţigănaşi", "Țigănași",
    "Ţuţora", "Țuțora",
    "Ungheni",
    "Valea Lupului",
    "Victoria",
    "Voineşti", "Voinești",
    # Sate aparținătoare comunelor ZM (cele mai importante)
    # Aroneanu
    "Dorobanţ", "Dorobanț", "Şorogari", "Șorogari",
    # Bârnova
    "Cercu", "Pietrăria", "Păun", "Vişan", "Vișan", "Todirel",
    # Ciurea
    "Curături", "Hlincea", "Lunca Cetăţuii", "Lunca Cetățuii", "Picioru Lupului",
    # Holboca
    "Cristeşti", "Cristești", "Dancu", "Orzeni", "Rusenii Noi", "Rusenii Vechi", "Tăuteşti", "Tăutești", "Valea Lungă",
    # Lețcani
    "Bogonos", "Cogeasca",
    # Miroslava
    "Brătuleni", "Ciurbești", "Cornești", "Dancaș", "Găureni", "Horpaz", "Proselnici", "Uricani", "Voroveşti", "Vorovești",
    # Popricani
    "Cuza Vodă", "Mânjeşti", "Mânjeștĭ", "Rediu Aldei", "Vulturi",
    # Rediu
    "Breazu", "Tăuteşti",
    # Tomești
    "Chicerea", "Goruni", "Vlădiceni", "Vladiceni",
    # Valea Lupului
    "Tomeşti", "Tomești",
    # Schitu Duca
    "Dumitreştii Gălăţii",
    # Victoria
    "Frunzeni",
}


def get_localities(session, judet_slug):
    resp = session.get(f"{BASE_URL}/judet/{judet_slug}", timeout=15)
    resp.raise_for_status()
    pattern = r'href="/judet/' + \
        re.escape(judet_slug) + \
        r'/([^"]+)"[^>]*>.*?<span class="loc-name">([^<]+)</span>'
    matches = re.findall(pattern, resp.text, re.DOTALL)
    seen = set()
    locs = []
    for slug, name in matches:
        slug = slug.strip('/')
        name = name.strip()
        if slug and slug not in seen:
            seen.add(slug)
            locs.append({'name': name, 'slug': slug,
                        'url': f"{BASE_URL}/judet/{judet_slug}/{slug}"})
    return locs


def scrape_locality(session, loc, judet_name):
    try:
        resp = session.get(loc['url'], timeout=30)
        resp.raise_for_status()
        html = resp.text
    except Exception as e:
        print(f"    [ERR] {loc['name']}: {e}")
        return []

    results = []

    # Tabel 3 coloane (orașe)
    rows = re.findall(
        r'<tr[^>]*>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*</tr>',
        html, re.DOTALL
    )
    for c1, c2, c3 in rows:
        strada = re.sub(r'<[^>]+>', '', c1).strip()
        numere = re.sub(r'<[^>]+>', '', c2).strip()
        cod = re.sub(r'<[^>]+>', '', c3).strip()
        if cod and cod.isdigit() and len(cod) == 6:
            results.append({
                'cod_postal': cod, 'judet': judet_name, 'localitate': loc['name'],
                'strada': f"{strada} {numere}".strip() if numere else strada,
                'numere': numere, 'in_zm': loc['name'] in ZM_IASI,
            })

    # Sate mici - DOAR codul principal
    if not results:
        meta = re.search(
            r'<meta\s+name="fact:postal-code"\s+content="[^"]*?(\d{6})', html)
        if meta:
            cod = meta.group(1)
        else:
            title = re.search(r'<title>[^<]*?(\d{6})[^<]*</title>', html)
            if title:
                cod = title.group(1)
            else:
                card = re.search(
                    r'class="[^"]*code-card[^"]*"[^>]*>.*?(\d{6})', html, re.DOTALL)
                cod = card.group(1) if card else None

        if cod:
            results.append({
                'cod_postal': cod, 'judet': judet_name, 'localitate': loc['name'],
                'strada': '', 'numere': '', 'in_zm': loc['name'] in ZM_IASI,
            })

    return results


def export_excel(all_data, zm_data, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import Counter

    wb = Workbook()
    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    zm_fill = PatternFill('solid', fgColor='E2EFDA')
    hdrs = ['Nr.', 'Cod Poștal', 'Județ',
            'Localitate', 'Strada', 'Numere/Blocuri']

    def write_sheet(ws, title, subtitle, data):
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = Font(name='Arial', bold=True, size=13, color='2F5496')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:F2')
        ws['A2'] = subtitle
        ws['A2'].font = Font(name='Arial', size=10,
                             italic=True, color='666666')
        ws['A2'].alignment = Alignment(horizontal='center')
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb
        data.sort(key=lambda x: (x['localitate'],
                  x['strada'], x['cod_postal']))
        for i, rd in enumerate(data):
            row = i + 5
            for col, val in enumerate([i+1, rd['cod_postal'], rd['judet'], rd['localitate'], rd['strada'], rd.get('numere', '')], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name='Arial', size=10, bold=(col == 2))
                c.alignment = Alignment(horizontal='center') if col in (
                    1, 2) else Alignment()
                c.border = tb
                if rd.get('in_zm'):
                    c.fill = zm_fill
        for col, w in {'A': 6, 'B': 11, 'C': 8, 'D': 22, 'E': 45, 'F': 30}.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A5'
        ws.auto_filter.ref = f'A4:F{4+len(data)}'

    sub = f'Sursa: codul-postal.ro | {datetime.now().strftime("%d.%m.%Y %H:%M")}'

    ws1 = wb.active
    ws1.title = "ZM Iași"
    write_sheet(
        ws1, 'CODURI POȘTALE — ZONA METROPOLITANĂ IAȘI (28 membri)', sub, zm_data)

    ws2 = wb.create_sheet("Tot Iași")
    write_sheet(ws2, 'CODURI POȘTALE — JUDEȚUL IAȘI COMPLET', sub, all_data)

    # Sumar
    ws3 = wb.create_sheet("Sumar")
    ws3['A1'] = 'SUMAR PE LOCALITĂȚI'
    ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    for col, h in enumerate(['Localitate', 'Intrări', 'Coduri unice', 'ZM Iași'], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font, c.fill, c.border = hf, hfill, tb
    lc = Counter(r['localitate'] for r in all_data)
    lu = {}
    for r in all_data:
        lu.setdefault(r['localitate'], set()).add(r['cod_postal'])
    for i, (loc, cnt) in enumerate(sorted(lc.items()), 4):
        ws3.cell(row=i, column=1, value=loc).border = tb
        ws3.cell(row=i, column=2, value=cnt).border = tb
        ws3.cell(row=i, column=3, value=len(lu.get(loc, set()))).border = tb
        is_zm = 'DA' if loc in ZM_IASI else ''
        ws3.cell(row=i, column=4, value=is_zm).border = tb
        if is_zm:
            for c in range(1, 5):
                ws3.cell(row=i, column=c).fill = zm_fill
    for col, w in {'A': 25, 'B': 10, 'C': 12, 'D': 10}.items():
        ws3.column_dimensions[col].width = w

    wb.save(filename)


def main():
    sd = os.path.dirname(os.path.abspath(__file__))

    print("=" * 60)
    print("  CODURI POSTALE IAȘI + ZM IAȘI — codul-postal.ro")
    print(f"  {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print("=" * 60)

    s = requests.Session()
    s.headers.update(HEADERS)

    print("\n[1] Localități Iași...")
    locs = get_localities(s, 'iasi')
    print(f"    {len(locs)} localități")
    if not locs:
        print("    ❌ Eroare")
        return
    print(f"    Primele 5: {[l['name'] for l in locs[:5]]}")
    zm_count = sum(1 for l in locs if l['name'] in ZM_IASI)
    print(f"    Din ZM Iași: {zm_count}")

    print(f"\n[2] Extragere coduri...\n")
    all_data = []
    for i, loc in enumerate(locs):
        r = scrape_locality(s, loc, 'Iași')
        all_data.extend(r)
        zm = " [ZM]" if loc['name'] in ZM_IASI else ""
        if r:
            print(f"  [{i+1}/{len(locs)}] {loc['name']}{zm}: {len(r)}")
        time.sleep(DELAY)

    # Dedup
    u = {}
    for r in all_data:
        u[f"{r['cod_postal']}|{r['localitate']}|{r['strada']}"] = r
    all_data = list(u.values())
    zm_data = [r for r in all_data if r.get('in_zm')]

    ct = len(set(r['cod_postal'] for r in all_data))
    cz = len(set(r['cod_postal'] for r in zm_data))
    lt = len(set(r['localitate'] for r in all_data))
    lz = len(set(r['localitate'] for r in zm_data))

    print(f"\n{'='*60}")
    print(
        f"  Tot Iași:   {len(all_data)} intrări | {ct} coduri | {lt} localități")
    print(
        f"  ZM Iași:    {len(zm_data)} intrări | {cz} coduri | {lz} localități")
    print(f"{'='*60}")

    xlsx = os.path.join(sd, "coduri_postale_iasi_COMPLET_2026.xlsx")
    export_excel(all_data, zm_data, xlsx)
    print(f"\n✅ Excel: {xlsx}")

    with open(os.path.join(sd, "coduri_postale_iasi_2026.csv"), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=[
                           'cod_postal', 'judet', 'localitate', 'strada', 'numere', 'in_zm'])
        w.writeheader()
        w.writerows(sorted(all_data, key=lambda x: (
            x['localitate'], x['cod_postal'])))
    print("✅ CSV salvat")

    with open(os.path.join(sd, "coduri_postale_iasi_2026.json"), 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print("✅ JSON salvat")
    print(f"\n🎉 GATA!")


if __name__ == "__main__":
    main()

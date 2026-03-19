#!/usr/bin/env python3
"""
scrape_coduri_postale_UNIVERSAL.py v2
Script universal pentru extragere coduri postale per județ.

FIX v2: 
  - max_redirects=10 pe session (previne infinite redirect loop)
  - timeout scăzut la 15s pentru localități problematice
  - skip + log la erori (nu oprește scriptul)
  - retry 1x cu slug normalizat (fără diacritice) la redirect error

UTILIZARE:
  python3 scrape_coduri_postale_UNIVERSAL.py --judet cluj --sursa codul-postal
  python3 scrape_coduri_postale_UNIVERSAL.py --judet brasov --sursa posta-romana
"""

import requests
import json
import time
import re
import os
import csv
import argparse
import unicodedata
from datetime import datetime
from requests.adapters import HTTPAdapter

BASE_CP = "https://www.codul-postal.ro"
BASE_PR = "https://www.posta-romana.ro"
URL_COD = f"{BASE_PR}/cnpr-app/modules/cauta-cod-postal/ajax/cautare_cod.php?q="

HEADERS_CP = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "ro-RO,ro;q=0.9",
}
HEADERS_PR = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "ro-RO,ro;q=0.9,en;q=0.8",
    "Referer": "https://www.posta-romana.ro/ccp.html",
    "X-Requested-With": "XMLHttpRequest",
    "Origin": "https://www.posta-romana.ro",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
}

# ============================================================
# CONFIGURARE JUDEȚE
# ============================================================
JUDETE_CONFIG = {
    "dolj": {
        "name": "Dolj", "slug": "dolj",
        "code_start": 200000, "code_end": 207999,
        "zm_name": "ZM Craiova",
        "zm_members": {
            "Craiova", "Filiaşi", "Filiași", "Segarcea",
            "Almăj", "Beharca", "Bogea", "Cotofenii din Față", "Cotofenii din Faţă", "Moşneni", "Moșneni", "Şitoaia", "Șitoaia",
            "Brădeşti", "Brădești", "Brădeştii Bătrâni", "Brădeștii Bătrâni", "Meteu", "Piscani", "Răcarii de Jos", "Tatomireşti", "Tatomirești",
            "Breasta", "Cotu", "Crovna", "Făget", "Obedin", "Roşieni", "Roșieni", "Valea Lungului",
            "Bucovăţ", "Bucovăț", "Cârligei", "Italieni", "Leamna de Jos", "Leamna de Sus", "Palilula", "Sărbătoarea",
            "Calopăr", "Bâzdâna", "Belcinu", "Panaghia", "Sălcuţa", "Sălcuța", "Cârcea", "Coşoveni", "Coșoveni",
            "Gherceşti", "Ghercești", "Gârleşti", "Gârlești", "Luncşoru", "Luncșoru", "Ungureni", "Ungurenii Mici",
            "Işalniţa", "Ișalnița", "Izvoare", "Malu Mare", "Ghindeni", "Preajba",
            "Mischii", "Călineşti", "Călinești", "Gogoşeşti", "Gogoșești", "Mlecăneşti", "Mlecănești", "Motoci", "Urecheşti", "Urechești",
            "Murgaşi", "Murgași", "Gaia", "Picăturile", "Rupturile", "Veleşti", "Velești",
            "Pieleşti", "Pielești", "Câmpeni", "Lânga",
            "Predeşti", "Predești", "Bucicani", "Cârstovani", "Frasin", "Milovan", "Pleşoi", "Pleșoi", "Predeştii Mici", "Predeștii Mici",
            "Şimnicu de Sus", "Șimnicu de Sus", "Albeşti", "Albești", "Cornetu", "Deleni", "Dudoviceşti", "Dudovicești",
            "Floreşti", "Florești", "Izvor", "Jieni", "Leşile", "Leșile", "Mileşti", "Milești", "Româneşti", "Românești",
            "Teasc", "Secui", "Terpeziţa", "Terpezița", "Căciulatu", "Căruia", "Floran", "Lazu",
            "Ţuglui", "Țuglui", "Jiul", "Unirea",
            "Vârvoru de Jos", "Bujor", "Ciutura", "Criva", "Dobromira", "Drăgoaia", "Gabru", "Vârvor",
            "Vela", "Bucovicior", "Cetăţuia", "Cetățuia", "Desnăţui", "Desnățui", "Gubaucea", "Segleţ", "Segleț", "Suharu", "Ştiubei", "Știubei",
            "Făcăi", "Mofleni", "Popoveni", "Şimnicu de Jos", "Șimnicu de Jos",
            "Almăjel", "Bâlta", "Branişte", "Braniște", "Fratoştița", "Răcarii de Sus", "Uscăci",
        },
    },
    "iasi": {
        "name": "Iași", "slug": "iasi",
        "code_start": 700000, "code_end": 707999,
        "zm_name": "ZM Iași",
        "zm_members": {
            "Iaşi", "Iași", "Aroneanu", "Bârnova", "Bîrnova", "Ciurea", "Comarna", "Costuleni",
            "Dobrovăţ", "Dobrovăț", "Golăieşti", "Golăiești", "Grajduri", "Holboca",
            "Horleşti", "Horlești", "Leţcani", "Lețcani", "Miroslava", "Mogoşeşti", "Mogoșești",
            "Movileni", "Popricani", "Prisăcani", "Rediu", "Româneşti", "Românești",
            "Scânteia", "Schitu Duca", "Tomeşti", "Tomești", "Ţigănaşi", "Țigănași",
            "Ţuţora", "Țuțora", "Ungheni", "Valea Lupului", "Victoria", "Voineşti", "Voinești",
            "Dorobanţ", "Dorobanț", "Şorogari", "Șorogari", "Cercu", "Pietrăria", "Păun",
            "Vişan", "Vișan", "Todirel", "Curături", "Hlincea", "Picioru Lupului",
            "Cristeşti", "Cristești", "Dancu", "Orzeni", "Rusenii Noi", "Rusenii Vechi",
            "Tăuteşti", "Tăutești", "Valea Lungă", "Bogonos", "Cogeasca",
            "Brătuleni", "Ciurbești", "Cornești", "Dancaș", "Găureni", "Horpaz",
            "Proselnici", "Uricani", "Voroveşti", "Vorovești", "Cuza Vodă",
            "Mânjeşti", "Rediu Aldei", "Vulturi", "Breazu",
            "Chicerea", "Goruni", "Vlădiceni", "Vladiceni", "Frunzeni",
            "Lunca Cetăţuii", "Lunca Cetățuii",
        },
    },
    "cluj": {
        "name": "Cluj", "slug": "cluj",
        "code_start": 400000, "code_end": 407999,
        "zm_name": "ZM Cluj-Napoca",
        "zm_members": {
            "Cluj-Napoca", "Aiton", "Apahida", "Baciu", "Bonţida", "Bonțida",
            "Borşa", "Borșa", "Căianu", "Chinteni", "Ciurila", "Cojocna",
            "Feleacu", "Floreşti", "Florești", "Gilău", "Gârbău", "Gîrbău",
            "Jucu", "Petreştii de Jos", "Petreștii de Jos", "Săvădisla",
            "Sânpaul", "Sîmpaul", "Tureni", "Vultureni",
            "Dezmir", "Sânnicoara", "Someşeni", "Someșeni",
            "Luna de Sus", "Gheorgheni", "Lomb", "Sălicea",
        },
    },
    "brasov": {
        "name": "Brașov", "slug": "brasov",
        "code_start": 500000, "code_end": 507999,
        "zm_name": "ZM Brașov",
        "zm_members": {
            "Braşov", "Brașov", "Codlea", "Ghimbav", "Predeal", "Râşnov", "Râșnov",
            "Săcele", "Zărneşti", "Zărnești",
            "Bod", "Bran", "Cristian", "Feldioara", "Hălchiu",
            "Hărman", "Prejmer", "Sânpetru", "Sîmpetru", "Tărlungeni", "Vulcan",
            "Stupini", "Poiana Brașov", "Poiana Braşov", "Timişu de Jos", "Timișu de Jos",
            "Timişu de Sus", "Timișu de Sus", "Colonia Bod", "Crizbav",
        },
    },
    "constanta": {
        "name": "Constanța", "slug": "constanta",
        "code_start": 900000, "code_end": 907999,
        "zm_name": "ZM Constanța",
        "zm_members": {
            "Constanţa", "Constanța",
            "Eforie", "Eforie Nord", "Eforie Sud", "Murfatlar", "Năvodari",
            "Ovidiu", "Techirghiol",
            "23 August", "Agigea", "Corbu", "Costineşti", "Costinești",
            "Cumpăna", "Lumina", "Mihai Kogălniceanu",
            "Poarta Albă", "Tuzla", "Valu lui Traian",
            "Palazu Mare", "Mamaia", "Mamaia-Sat",
        },
    },
    "sibiu": {
        "name": "Sibiu", "slug": "sibiu",
        "code_start": 550000, "code_end": 557999,
        "zm_name": "ZM Sibiu",
        "zm_members": {
            "Sibiu",
            "Şelimbăr", "Șelimbăr", "Şura Mică", "Șura Mică",
            "Roşia", "Roșia", "Şura Mare", "Șura Mare",
            "Ocna Sibiului", "Sadu", "Poplaca",
            "Cisnădie", "Cisnădioară",
            "Cristian", "Gura Râului",
            "Răşinari", "Rășinari",
            "Orlat", "Tălmaciu",
            "Turnişor", "Turnișor", "Bungard",
        },
    },
}


# ============================================================
# HELPER: normalize diacritics for retry
# ============================================================
def remove_diacritics(text):
    """Înlocuiește ăâîșț cu aist pentru URL retry."""
    replacements = {
        'ă': 'a', 'â': 'a', 'î': 'i', 'ș': 's', 'ț': 't',
        'Ă': 'A', 'Â': 'A', 'Î': 'I', 'Ș': 'S', 'Ț': 'T',
        'ş': 's', 'ţ': 't', 'Ş': 'S', 'Ţ': 'T',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


# ============================================================
# FUNCȚII codul-postal.ro
# ============================================================
def cp_create_session():
    s = requests.Session()
    s.headers.update(HEADERS_CP)
    s.max_redirects = 10  # FIX: prevent infinite redirect loops
    return s


def cp_get_localities(session, slug):
    resp = session.get(f"{BASE_CP}/judet/{slug}", timeout=15)
    resp.raise_for_status()
    pattern = r'href="/judet/' + \
        re.escape(slug) + \
        r'/([^"]+)"[^>]*>.*?<span class="loc-name">([^<]+)</span>'
    matches = re.findall(pattern, resp.text, re.DOTALL)
    seen = set()
    locs = []
    for s, name in matches:
        s = s.strip('/')
        name = name.strip()
        if s and s not in seen:
            seen.add(s)
            locs.append({'name': name, 'slug': s,
                        'url': f"{BASE_CP}/judet/{slug}/{s}"})
    return locs


def cp_fetch_page(session, url, loc_name):
    """Fetch cu retry: dacă redirect loop, încearcă fără diacritice."""
    try:
        resp = session.get(url, timeout=15)
        resp.raise_for_status()
        return resp.text
    except requests.exceptions.TooManyRedirects:
        # Retry cu slug normalizat (fără diacritice)
        slug_clean = remove_diacritics(url.split('/')[-1])
        url_retry = '/'.join(url.split('/')[:-1]) + '/' + slug_clean
        try:
            resp = session.get(url_retry, timeout=15)
            resp.raise_for_status()
            return resp.text
        except Exception:
            pass

        # Retry cu slug lowercase fără diacritice
        slug_lower = slug_clean.lower().replace(' ', '-')
        url_retry2 = '/'.join(url.split('/')[:-1]) + '/' + slug_lower
        try:
            resp = session.get(url_retry2, timeout=15)
            resp.raise_for_status()
            return resp.text
        except Exception:
            print(f"    [SKIP] {loc_name}: redirect loop (tried 3 URLs)")
            return None
    except Exception as e:
        print(f"    [ERR] {loc_name}: {e}")
        return None


def cp_scrape_locality(session, loc, judet_name, zm_members):
    html = cp_fetch_page(session, loc['url'], loc['name'])
    if not html:
        return []

    results = []
    rows = re.findall(
        r'<tr[^>]*>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*</tr>', html, re.DOTALL)
    for c1, c2, c3 in rows:
        strada = re.sub(r'<[^>]+>', '', c1).strip()
        numere = re.sub(r'<[^>]+>', '', c2).strip()
        cod = re.sub(r'<[^>]+>', '', c3).strip()
        if cod and cod.isdigit() and len(cod) == 6:
            results.append({
                'cod_postal': cod, 'judet': judet_name, 'localitate': loc['name'],
                'strada': f"{strada} {numere}".strip() if numere else strada,
                'numere': numere, 'in_zm': loc['name'] in zm_members,
                'sursa': 'codul-postal.ro',
            })

    if not results:
        meta = re.search(
            r'<meta\s+name="fact:postal-code"\s+content="[^"]*?(\d{6})', html)
        if meta:
            cod = meta.group(1)
        else:
            title = re.search(r'<title>[^<]*?(\d{6})[^<]*</title>', html)
            if title:
                cod = title.group(1)
            else:
                card = re.search(
                    r'class="[^"]*code-card[^"]*"[^>]*>.*?(\d{6})', html, re.DOTALL)
                cod = card.group(1) if card else None
        if cod:
            results.append({
                'cod_postal': cod, 'judet': judet_name, 'localitate': loc['name'],
                'strada': '', 'numere': '', 'in_zm': loc['name'] in zm_members,
                'sursa': 'codul-postal.ro',
            })
    return results


def run_codul_postal(config):
    s = cp_create_session()

    print(f"\n[1] Localități {config['name']}...")
    locs = cp_get_localities(s, config['slug'])
    print(f"    {len(locs)} localități")
    if not locs:
        print("    ❌ Eroare")
        return []

    zm_count = sum(1 for l in locs if l['name'] in config['zm_members'])
    print(f"    Din {config['zm_name']}: {zm_count}")

    print(f"\n[2] Extragere coduri...\n")
    all_data = []
    skipped = []
    for i, loc in enumerate(locs):
        r = cp_scrape_locality(s, loc, config['name'], config['zm_members'])
        all_data.extend(r)
        zm = " [ZM]" if loc['name'] in config['zm_members'] else ""
        if r and len(r) > 1:
            print(f"  [{i+1}/{len(locs)}] {loc['name']}{zm}: {len(r)}")
        if not r:
            skipped.append(loc['name'])
        time.sleep(0.3)

    if skipped:
        print(
            f"\n  ⚠️  {len(skipped)} localități skip-uite (redirect/eroare):")
        for s_name in skipped[:20]:
            print(f"    - {s_name}")
        if len(skipped) > 20:
            print(f"    ... și încă {len(skipped)-20}")

    u = {}
    for r in all_data:
        u[f"{r['cod_postal']}|{r['localitate']}|{r['strada']}"] = r
    return list(u.values())


# ============================================================
# FUNCȚII Poșta Română
# ============================================================
def pr_parse_html(html_str):
    results = []
    p_tags = re.findall(r'<p[^>]*>(.*?)</p>', html_str)
    for i in range(0, len(p_tags), 5):
        chunk = p_tags[i:i+5]
        if len(chunk) >= 4:
            cod = chunk[0].strip()
            if cod and cod[0].isdigit() and len(cod) == 6:
                results.append({
                    "cod_postal": cod, "judet": chunk[1].strip(),
                    "localitate": chunk[2].strip(), "strada": chunk[3].strip(),
                    "subunitate_postala": chunk[4].strip() if len(chunk) > 4 else "",
                })
    return results


def run_posta_romana(config):
    sd = os.path.dirname(os.path.abspath(__file__))
    progress_file = os.path.join(sd, f"scan_progress_{config['slug']}.json")

    s = requests.Session()
    s.headers.update(HEADERS_PR)
    s.get(f"{BASE_PR}/ccp.html", timeout=10)

    code_start = config['code_start']
    code_end = config['code_end']
    total = code_end - code_start + 1

    print(f"\n[1] Test cod {code_start + 100}...")
    data = f"k_cod_postal={code_start + 100}&k_lang=ro"
    try:
        resp = s.post(URL_COD, data=data, timeout=15)
        result = resp.json()
        if result.get("found", 0) > 0:
            print(f"    ✅ OK")
        else:
            print(f"    ⚠️  Fără rezultat, continuăm...")
    except:
        print(f"    ⚠️  Eroare test, continuăm...")

    all_data = []
    existing_keys = set()
    resume_from = code_start
    if os.path.exists(progress_file):
        try:
            with open(progress_file, 'r', encoding='utf-8') as f:
                progress = json.load(f)
            all_data = progress.get('data', [])
            resume_from = progress.get('last_code', code_start) + 1
            existing_keys = set(
                f"{r['cod_postal']}|{r['strada']}" for r in all_data)
            print(f"    📋 Resume de la {resume_from} ({len(all_data)} deja)")
        except:
            pass

    print(f"\n[2] Scanare {resume_from} → {code_end}...\n")
    start_time = time.time()

    try:
        for code in range(resume_from, code_end + 1):
            code_str = str(code).zfill(6)
            scanned = code - code_start + 1

            try:
                resp = s.post(
                    URL_COD, data=f"k_cod_postal={code_str}&k_lang=ro", timeout=15)
                result = resp.json()
                results = pr_parse_html(result.get("formular", "")) if result.get(
                    "found", 0) > 0 else []
            except:
                results = []

            if results:
                new = 0
                for r in results:
                    key = f"{r['cod_postal']}|{r['strada']}"
                    if key not in existing_keys:
                        r['sursa'] = 'Poșta Română'
                        r['in_zm'] = r['localitate'] in config['zm_members']
                        r['numere'] = ''
                        all_data.append(r)
                        existing_keys.add(key)
                        new += 1
                if new > 0:
                    print(f"  ✅ {code_str}: +{new} (total {len(all_data)})")

            if scanned % 200 == 0:
                pct = scanned / total * 100
                eta = (code_end - code) * 0.2 / 60
                print(
                    f"  [{scanned}/{total}] {pct:.1f}% | {code_str} | {len(all_data)} găsite | ~{eta:.0f}min")
                with open(progress_file, 'w', encoding='utf-8') as f:
                    json.dump({'last_code': code, 'data': all_data},
                              f, ensure_ascii=False)

            time.sleep(0.2)
            if scanned % 2000 == 0:
                s = requests.Session()
                s.headers.update(HEADERS_PR)
                s.get(f"{BASE_PR}/ccp.html", timeout=10)

    except KeyboardInterrupt:
        print(f"\n⏸️  Salvez progres ({len(all_data)})...")
        with open(progress_file, 'w', encoding='utf-8') as f:
            json.dump({'last_code': code, 'data': all_data},
                      f, ensure_ascii=False)

    if os.path.exists(progress_file) and code >= code_end:
        os.remove(progress_file)

    return all_data


# ============================================================
# EXPORT EXCEL
# ============================================================
def export_excel(all_data, zm_data, config, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import Counter

    wb = Workbook()
    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    zm_fill = PatternFill('solid', fgColor='E2EFDA')
    hdrs = ['Nr.', 'Cod Poștal', 'Județ',
            'Localitate', 'Strada', 'Numere', 'Sursa']

    def write_sheet(ws, title, data):
        ws.merge_cells('A1:G1')
        ws['A1'] = title
        ws['A1'].font = Font(name='Arial', bold=True, size=13, color='2F5496')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')
        ws['A2'] = f'{datetime.now().strftime("%d.%m.%Y %H:%M")}'
        ws['A2'].font = Font(name='Arial', size=10,
                             italic=True, color='666666')
        ws['A2'].alignment = Alignment(horizontal='center')
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb
        data.sort(key=lambda x: (x['localitate'],
                  x.get('strada', ''), x['cod_postal']))
        for i, rd in enumerate(data):
            row = i + 5
            for col, val in enumerate([i+1, rd['cod_postal'], rd.get('judet', ''), rd['localitate'], rd.get('strada', ''), rd.get('numere', ''), rd.get('sursa', '')], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name='Arial', size=10, bold=(col == 2))
                c.alignment = Alignment(horizontal='center') if col in (
                    1, 2, 7) else Alignment()
                c.border = tb
                if rd.get('in_zm'):
                    c.fill = zm_fill
        for col, w in {'A': 6, 'B': 11, 'C': 10, 'D': 22, 'E': 45, 'F': 30, 'G': 14}.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A5'
        ws.auto_filter.ref = f'A4:G{4+len(data)}'

    ws1 = wb.active
    ws1.title = config['zm_name'][:31]
    write_sheet(ws1, f'CODURI POȘTALE — {config["zm_name"].upper()}', zm_data)
    ws2 = wb.create_sheet(f"Tot {config['name']}")
    write_sheet(
        ws2, f'CODURI POȘTALE — JUDEȚUL {config["name"].upper()} COMPLET', all_data)

    ws3 = wb.create_sheet("Sumar")
    ws3['A1'] = f'SUMAR {config["name"].upper()}'
    ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    for col, h in enumerate(['Localitate', 'Intrări', 'Coduri', 'ZM'], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font, c.fill, c.border = hf, hfill, tb
    lc = Counter(r['localitate'] for r in all_data)
    lu = {}
    for r in all_data:
        lu.setdefault(r['localitate'], set()).add(r['cod_postal'])
    for i, (loc, cnt) in enumerate(sorted(lc.items()), 4):
        ws3.cell(row=i, column=1, value=loc).border = tb
        ws3.cell(row=i, column=2, value=cnt).border = tb
        ws3.cell(row=i, column=3, value=len(lu.get(loc, set()))).border = tb
        is_zm = 'DA' if loc in config['zm_members'] else ''
        ws3.cell(row=i, column=4, value=is_zm).border = tb
        if is_zm:
            for c in range(1, 5):
                ws3.cell(row=i, column=c).fill = zm_fill
    for col, w in {'A': 25, 'B': 10, 'C': 10, 'D': 6}.items():
        ws3.column_dimensions[col].width = w
    wb.save(filename)


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description='Extrage coduri poștale per județ')
    parser.add_argument('--judet', required=True, choices=list(JUDETE_CONFIG.keys()),
                        help='Județul')
    parser.add_argument('--sursa', required=True, choices=['codul-postal', 'posta-romana'],
                        help='Sursa de date')
    args = parser.parse_args()

    config = JUDETE_CONFIG[args.judet]
    sd = os.path.dirname(os.path.abspath(__file__))
    sursa_tag = 'cp' if args.sursa == 'codul-postal' else 'pr'

    print("=" * 65)
    print(f"  CODURI POSTALE {config['name'].upper()} — {args.sursa}")
    if args.sursa == 'posta-romana':
        print(f"  Range: {config['code_start']} → {config['code_end']}")
    print(f"  {config['zm_name']}")
    print(f"  {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print("=" * 65)

    if args.sursa == 'codul-postal':
        all_data = run_codul_postal(config)
    else:
        all_data = run_posta_romana(config)

    zm_data = [r for r in all_data if r.get('in_zm')]
    ct = len(set(r['cod_postal'] for r in all_data))
    cz = len(set(r['cod_postal'] for r in zm_data))
    lt = len(set(r['localitate'] for r in all_data))
    lz = len(set(r['localitate'] for r in zm_data))

    print(f"\n{'='*65}")
    print(
        f"  Tot {config['name']}: {len(all_data)} intrări | {ct} coduri | {lt} localități")
    print(
        f"  {config['zm_name']}: {len(zm_data)} intrări | {cz} coduri | {lz} localități")
    print(f"{'='*65}")

    xlsx = os.path.join(
        sd, f"coduri_postale_{args.judet}_{sursa_tag}_2026.xlsx")
    export_excel(all_data, zm_data, config, xlsx)
    print(f"\n✅ Excel: {xlsx}")

    csv_f = os.path.join(
        sd, f"coduri_postale_{args.judet}_{sursa_tag}_2026.csv")
    with open(csv_f, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=[
                           'cod_postal', 'judet', 'localitate', 'strada', 'numere', 'in_zm', 'sursa'], extrasaction='ignore')
        w.writeheader()
        w.writerows(sorted(all_data, key=lambda x: (
            x['localitate'], x['cod_postal'])))
    print(f"✅ CSV: {csv_f}")

    with open(os.path.join(sd, f"coduri_postale_{args.judet}_{sursa_tag}_2026.json"), 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON salvat")
    print(f"\n🎉 GATA!")


if __name__ == "__main__":
    main()

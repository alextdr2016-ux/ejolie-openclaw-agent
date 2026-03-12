#!/usr/bin/env python3
"""
Script: Localități în raza de X km de un oraș (Overpass API / OpenStreetMap)
===========================================================================
Folosește Overpass API pentru a extrage TOATE localitățile din România
aflate într-o rază specificată de un punct central.

Tipuri de localități extrase:
  - city     = municipiu / oraș mare
  - town     = oraș mic
  - village  = sat / centru de comună
  - hamlet   = cătun / sat mic

Folosire:
    python3 localitati_overpass.py                          # default: Iași, 100km
    python3 localitati_overpass.py --oras Bucuresti --raza 50
    python3 localitati_overpass.py --lat 47.1585 --lon 27.6014 --raza 150
    python3 localitati_overpass.py --format csv             # doar CSV, fără Excel
    python3 localitati_overpass.py --tip city,town          # doar orașe, fără sate

Deployment pe EC2:
    Locație: ~/ejolie-openclaw-agent/ejolie-sales/scripts/localitati_overpass.py
    pip3 install openpyxl requests

Autor: Alex Tudor | Data: Martie 2026
"""

import argparse
import csv
import json
import math
import os
import sys
import time
from datetime import datetime

try:
    import requests
except ImportError:
    print("EROARE: pip3 install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("ATENȚIE: openpyxl nu e instalat, se generează doar CSV. pip3 install openpyxl")


# ============================================================
# CONFIGURARE DEFAULT
# ============================================================
OVERPASS_URL = "https://overpass-api.de/api/interpreter"

# Orașe predefinite (pentru --oras)
ORASE_PREDEFINITE = {
    "iasi":       (47.1585, 27.6014),
    "bucuresti":  (44.4268, 26.1025),
    "cluj":       (46.7712, 23.6236),
    "timisoara":  (45.7489, 21.2087),
    "constanta":  (44.1598, 28.6348),
    "brasov":     (45.6427, 25.5887),
    "craiova":    (44.3302, 23.7949),
    "galati":     (45.4353, 28.0078),
    "sibiu":      (45.7983, 24.1256),
    "suceava":    (47.6514, 26.2556),
    "bacau":      (46.5670, 26.9146),
    "oradea":     (47.0465, 21.9189),
    "piatra neamt": (46.9275, 26.3708),
    "botosani":   (47.7487, 26.6694),
    "vaslui":     (46.6407, 27.7276),
    "roman":      (46.9168, 26.9264),
    "pascani":    (47.2487, 26.7178),
    "barlad":     (46.2319, 27.6714),
    "husi":       (46.6714, 28.0600),
    "focsani":    (45.6967, 27.1833),
}

# Mapping tipuri OSM → româna
TIP_MAPPING = {
    "city":    "Municipiu/Oraș mare",
    "town":    "Oraș",
    "village": "Sat/Comună",
    "hamlet":  "Cătun",
}


# ============================================================
# FORMULA HAVERSINE
# ============================================================
def haversine(lat1, lon1, lat2, lon2):
    """Distanța în km între 2 puncte GPS."""
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat/2)**2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon/2)**2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))


# ============================================================
# OVERPASS QUERY
# ============================================================
def build_overpass_query(lat, lon, raza_m, tipuri):
    """
    Construiește query-ul Overpass QL.

    Caută noduri cu place=city/town/village/hamlet
    într-o rază de X metri de punctul central.
    """
    # Construim union de query-uri per tip
    queries = ""
    for tip in tipuri:
        queries += f'  node["place"="{tip}"](around:{raza_m},{lat},{lon});\n'

    query = f"""
[out:json][timeout:120];
(
{queries});
out body;
"""
    return query.strip()


def fetch_overpass(lat, lon, raza_km, tipuri, max_retries=3):
    """
    Execută query-ul Overpass și returnează lista de localități.
    Include retry logic pentru rate limiting (429).
    """
    raza_m = int(raza_km * 1000)
    query = build_overpass_query(lat, lon, raza_m, tipuri)

    print(
        f"  Query Overpass: raza {raza_km}km ({raza_m}m), tipuri: {', '.join(tipuri)}")
    print(f"  URL: {OVERPASS_URL}")

    for attempt in range(1, max_retries + 1):
        try:
            print(f"  Încercare {attempt}/{max_retries}...")
            resp = requests.post(
                OVERPASS_URL,
                data={"data": query},
                timeout=120,
                headers={"User-Agent": "ejolie-localitati-script/1.0"}
            )

            if resp.status_code == 429:
                wait = 30 * attempt
                print(f"  ⏳ Rate limited (429). Aștept {wait}s...")
                time.sleep(wait)
                continue

            if resp.status_code == 504:
                print(f"  ⏳ Timeout server (504). Aștept 10s și reîncerc...")
                time.sleep(10)
                continue

            resp.raise_for_status()
            data = resp.json()

            elements = data.get("elements", [])
            print(f"  ✓ Primit {len(elements)} elemente de la Overpass")
            return elements

        except requests.exceptions.Timeout:
            print(f"  ⏳ Timeout client. Aștept 10s...")
            time.sleep(10)
        except requests.exceptions.RequestException as e:
            print(f"  ✗ Eroare: {e}")
            if attempt < max_retries:
                time.sleep(5)

    print("  ✗ Toate încercările au eșuat.")
    return []


# ============================================================
# PROCESARE REZULTATE
# ============================================================
def process_results(elements, center_lat, center_lon):
    """
    Procesează elementele Overpass și returnează lista de localități
    cu distanța calculată, sortată după distanță.
    """
    results = []
    seen = set()  # deduplicare după nume+tip

    for el in elements:
        if el.get("type") != "node":
            continue

        tags = el.get("tags", {})
        name = tags.get("name", "").strip()
        if not name:
            continue

        lat = el.get("lat", 0)
        lon = el.get("lon", 0)
        place_type = tags.get("place", "unknown")

        # Deduplicare
        key = f"{name}_{place_type}_{round(lat, 2)}_{round(lon, 2)}"
        if key in seen:
            continue
        seen.add(key)

        dist = haversine(center_lat, center_lon, lat, lon)

        # Extrage informații suplimentare din tag-uri
        population = tags.get("population", "")
        try:
            population = int(population) if population else 0
        except (ValueError, TypeError):
            population = 0

        county = tags.get("addr:county", tags.get("is_in:county", ""))
        postal = tags.get("addr:postcode", tags.get("postal_code", ""))
        name_ro = tags.get("name:ro", name)

        results.append({
            "nume": name_ro,
            "nume_osm": name,
            "judet": county,
            "tip_osm": place_type,
            "tip": TIP_MAPPING.get(place_type, place_type),
            "lat": round(lat, 6),
            "lon": round(lon, 6),
            "distanta_km": round(dist, 1),
            "populatie": population,
            "cod_postal": postal,
            "osm_id": el.get("id", ""),
        })

    results.sort(key=lambda x: x["distanta_km"])
    return results


# ============================================================
# EXPORT CSV
# ============================================================
def export_csv(results, filename, center_name, raza_km):
    """Salvează în CSV."""
    fieldnames = ["nr", "nume", "judet", "tip", "lat", "lon",
                  "distanta_km", "populatie", "cod_postal", "osm_id"]

    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for i, r in enumerate(results, 1):
            row = {k: r.get(k, "") for k in fieldnames}
            row["nr"] = i
            writer.writerow(row)

    print(f"  ✓ CSV: {filename} ({len(results)} rânduri)")


# ============================================================
# EXPORT XLSX
# ============================================================
def export_xlsx(results, filename, center_name, raza_km):
    """Salvează în Excel formatat profesional."""
    if not HAS_OPENPYXL:
        print("  ⚠ openpyxl lipsă, skip Excel.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = f"Localitati {raza_km}km {center_name}"

    # Stiluri
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center",
                          vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    city_fill = PatternFill("solid", fgColor="E2EFDA")
    town_fill = PatternFill("solid", fgColor="FFF2CC")

    # Titlu
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    c.value = f"Localități în raza de {raza_km} km de {center_name} — {len(results)} rezultate"
    c.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = f"Sursă: OpenStreetMap (Overpass API) | Generat: {ts}"
    c.font = Font(name="Arial", size=9, italic=True, color="666666")
    c.alignment = Alignment(horizontal="center")

    # Headere
    headers = ["Nr.", "Localitate", "Județ", "Tip", "Latitudine", "Longitudine",
               "Distanța (km)", "Populație", "Cod poștal"]
    widths = [6, 28, 16, 20, 13, 13, 14, 13, 12]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 25

    # Date
    for idx, r in enumerate(results, 1):
        row = idx + 4
        vals = [idx, r["nume"], r["judet"], r["tip"], r["lat"], r["lon"],
                r["distanta_km"], r["populatie"] if r["populatie"] > 0 else "",
                r["cod_postal"]]

        is_city = r["tip_osm"] == "city"
        is_town = r["tip_osm"] == "town"

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name="Arial", size=10, bold=(is_city or is_town))
            cell.border = border
            cell.alignment = center if col in (
                1, 5, 6, 7, 8, 9) else left_align

            if is_city:
                cell.fill = city_fill
            elif is_town:
                cell.fill = town_fill
            elif idx % 2 == 0:
                cell.fill = alt_fill

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{4 + len(results)}"

    # === Sheet 2: Statistici ===
    ws2 = wb.create_sheet("Statistici")
    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "Statistici per județ și tip"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")

    # Per județ
    stat_headers = ["Județ", "Nr. Localități",
                    "Municipii/Orașe mari", "Orașe", "Sate/Comune"]
    for i, h in enumerate(stat_headers, 1):
        cell = ws2.cell(row=3, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws2.column_dimensions[get_column_letter(i)].width = [
            16, 16, 20, 10, 14][i-1]

    judete = {}
    for r in results:
        j = r["judet"] or "(necunoscut)"
        if j not in judete:
            judete[j] = {"total": 0, "city": 0,
                         "town": 0, "village": 0, "hamlet": 0}
        judete[j]["total"] += 1
        judete[j][r["tip_osm"]] = judete[j].get(r["tip_osm"], 0) + 1

    row = 4
    for j in sorted(judete.keys()):
        s = judete[j]
        vals = [j, s["total"], s.get("city", 0), s.get("town", 0),
                s.get("village", 0) + s.get("hamlet", 0)]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.border = border
            cell.alignment = center
        row += 1

    # Total
    tot_row = row
    ws2.cell(row=tot_row, column=1, value="TOTAL").font = Font(
        name="Arial", bold=True)
    ws2.cell(row=tot_row, column=2, value=len(results)
             ).font = Font(name="Arial", bold=True)
    for col in range(1, 6):
        ws2.cell(row=tot_row, column=col).border = border
        ws2.cell(row=tot_row, column=col).alignment = center
        ws2.cell(row=tot_row, column=col).fill = PatternFill(
            "solid", fgColor="FFF2CC")

    wb.save(filename)
    print(f"  ✓ Excel: {filename}")


# ============================================================
# EXPORT JSON (pentru OpenClaw)
# ============================================================
def export_json(results, filename):
    """Salvează rezultatele ca JSON (util pentru OpenClaw/Telegram)."""
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"  ✓ JSON: {filename}")


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="Extrage localități într-o rază de un punct central (Overpass API / OSM)"
    )
    parser.add_argument("--oras", type=str, default="iasi",
                        help="Oraș predefinit (iasi, bucuresti, cluj, etc.)")
    parser.add_argument("--lat", type=float, default=None,
                        help="Latitudine punct central (override --oras)")
    parser.add_argument("--lon", type=float, default=None,
                        help="Longitudine punct central (override --oras)")
    parser.add_argument("--raza", type=int, default=100,
                        help="Raza în km (default: 100)")
    parser.add_argument("--tip", type=str, default="city,town,village,hamlet",
                        help="Tipuri localități: city,town,village,hamlet (default: toate)")
    parser.add_argument("--format", type=str, default="all",
                        choices=["csv", "xlsx", "json", "all"],
                        help="Format export (default: all)")
    parser.add_argument("--output-dir", type=str, default=".",
                        help="Director output (default: curent)")
    parser.add_argument("--telegram", action="store_true",
                        help="Trimite rezultatul pe Telegram după generare")

    args = parser.parse_args()

    # Determină coordonatele
    if args.lat and args.lon:
        center_lat, center_lon = args.lat, args.lon
        center_name = f"({args.lat}, {args.lon})"
    else:
        oras_key = args.oras.lower().strip()
        if oras_key not in ORASE_PREDEFINITE:
            print(f"✗ Orașul '{args.oras}' nu e în lista predefinită.")
            print(
                f"  Disponibile: {', '.join(sorted(ORASE_PREDEFINITE.keys()))}")
            print(f"  Alternativ, folosește --lat și --lon.")
            sys.exit(1)
        center_lat, center_lon = ORASE_PREDEFINITE[oras_key]
        center_name = args.oras.capitalize()

    tipuri = [t.strip() for t in args.tip.split(",")]
    raza_km = args.raza

    # Prefix fișiere
    safe_name = center_name.lower().replace(" ", "_").replace(
        "(", "").replace(")", "").replace(",", "")
    prefix = f"localitati_{raza_km}km_{safe_name}"

    print("=" * 60)
    print(f"  LOCALITĂȚI ÎN RAZA DE {raza_km} KM DE {center_name.upper()}")
    print(f"  Coordonate: {center_lat}°N, {center_lon}°E")
    print(f"  Tipuri: {', '.join(tipuri)}")
    print("=" * 60)

    # 1. Fetch Overpass
    print(f"\n[1/3] Interogare Overpass API...")
    elements = fetch_overpass(center_lat, center_lon, raza_km, tipuri)

    if not elements:
        print("\n✗ Nu s-au primit date de la Overpass. Verifică conexiunea.")
        sys.exit(1)

    # 2. Procesare
    print(f"\n[2/3] Procesare rezultate...")
    results = process_results(elements, center_lat, center_lon)
    print(f"  Total localități unice: {len(results)}")

    if not results:
        print("✗ Nicio localitate găsită.")
        sys.exit(1)

    # 3. Export
    print(f"\n[3/3] Export fișiere...")
    os.makedirs(args.output_dir, exist_ok=True)

    csv_path = os.path.join(args.output_dir, f"{prefix}.csv")
    xlsx_path = os.path.join(args.output_dir, f"{prefix}.xlsx")
    json_path = os.path.join(args.output_dir, f"{prefix}.json")

    if args.format in ("csv", "all"):
        export_csv(results, csv_path, center_name, raza_km)
    if args.format in ("xlsx", "all"):
        export_xlsx(results, xlsx_path, center_name, raza_km)
    if args.format in ("json", "all"):
        export_json(results, json_path)

    # Rezumat
    judete = {}
    for r in results:
        j = r["judet"] or "(necunoscut)"
        judete[j] = judete.get(j, 0) + 1

    tip_counts = {}
    for r in results:
        t = r["tip"]
        tip_counts[t] = tip_counts.get(t, 0) + 1

    print(f"\n{'=' * 60}")
    print(f"  REZUMAT")
    print(f"  Total localități: {len(results)}")
    print(f"  Per tip: {', '.join(f'{t}: {c}' for t,
          c in sorted(tip_counts.items()))}")
    if judete:
        print(f"  Județe: {', '.join(f'{j} ({c})' for j,
              c in sorted(judete.items()))}")
    if results:
        print(
            f"  Cea mai apropiată: {results[0]['nume']} ({results[0]['distanta_km']} km)")
        print(
            f"  Cea mai depărtată: {results[-1]['nume']} ({results[-1]['distanta_km']} km)")
    print(f"{'=' * 60}")

    # Telegram (dacă cerut)
    if args.telegram:
        send_telegram(results, center_name, raza_km, xlsx_path)

    return results


def send_telegram(results, center_name, raza_km, xlsx_path):
    """Trimite rezumatul pe Telegram și atașează Excel-ul."""
    try:
        # Citește config Telegram din .env sau direct
        chat_id = os.getenv("TELEGRAM_CHAT_ID", "44151343")

        # Rezumat text
        tip_counts = {}
        for r in results:
            t = r["tip"]
            tip_counts[t] = tip_counts.get(t, 0) + 1

        msg = f"📍 *Localități {raza_km}km de {center_name}*\n"
        msg += f"Total: *{len(results)}* localități\n"
        for t, c in sorted(tip_counts.items()):
            msg += f"  • {t}: {c}\n"

        print(f"\n  📱 Trimitere Telegram...")

        # Trimite via OpenClaw CLI
        os.system(
            f'openclaw message send --channel telegram --target {chat_id} "{msg}"')

        if os.path.exists(xlsx_path):
            os.system(
                f'openclaw message send --channel telegram --target {chat_id} --media {xlsx_path}')
            print(f"  ✓ Trimis pe Telegram")
    except Exception as e:
        print(f"  ⚠ Eroare Telegram: {e}")


if __name__ == "__main__":
    main()

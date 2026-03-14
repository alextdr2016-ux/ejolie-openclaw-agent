#!/usr/bin/env python3
"""
export_fashiondays.py v3 - Export produse Ejolie+Artista pentru Fashion Days
Scrie datele direct in template-ul oficial descarcat din Fashion Days.
NU creeaza fisier nou - pastreaza structura, sheet-urile, validarile originale.

Rulare: python3 export_fashiondays.py --template template_original.xlsx [--limit 10] [--dry-run]
"""

import json
import os
import sys
import argparse
import time
import copy
import requests
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BARCODE_PATH = os.path.join(SCRIPT_DIR, 'barcode_ejolie_map.json')

ENV_PATH = os.path.join(SCRIPT_DIR, '..', '.env')
API_KEY = None
if os.path.exists(ENV_PATH):
    with open(ENV_PATH) as f:
        for line in f:
            if line.startswith('EJOLIE_API_KEY='):
                API_KEY = line.strip().split(
                    '=', 1)[1].strip().strip('"').strip("'")
if not API_KEY:
    API_KEY = os.environ.get(
        'EJOLIE_API_KEY', 'N9komxWU3aclwDHyrXfLjJdBA6ZRTs')

API_BASE = 'https://ejolie.ro/api/'
HEADERS = {'User-Agent': 'Mozilla/5.0'}
PAGE_SIZE = 50
START_ROW = 6  # Template data starts at row 6

# ============================================================
# MAPPINGS Extended → Fashion Days
# ============================================================
COLOR_MAP = {
    'Alb': 'Alb', 'Albastru': 'Albastru', 'Albastru deschis': 'Albastru deschis',
    'Albastru inchis': 'Albastru inchis', 'Albastru petrol': 'Albastru',
    'Animal print': 'Multicolor', 'Aramiu': 'Auriu', 'Argintiu': 'Argintiu',
    'Auriu': 'Auriu', 'Bej': 'Bej', 'Bleumarin': 'Bleumarin', 'Bordo': 'Visiniu',
    'Caramel': 'Maro', 'Ciocolatiu': 'Maro', 'Corai': 'Corai', 'Crem': 'Bej',
    'Galben': 'Galben', 'Gri': 'Gri', 'Kaki': 'Kaki', 'Lavanda': 'Mov',
    'Lila': 'Lila', 'Maro': 'Maro', 'Mov': 'Mov', 'Multicolor': 'Multicolor',
    'Negru': 'Negru', 'Nude': 'Nude', 'Olive': 'Verde', 'Piersica': 'Roz',
    'Portocaliu': 'Portocaliu', 'Pudra': 'Roz', 'Rosu': 'Rosu', 'Roz': 'Roz',
    'Roz prafuit': 'Roz', 'Somon': 'Roz', 'Turcoaz': 'Turcoaz', 'Verde': 'Verde',
    'Verde inchis': 'Verde inchis', 'Verde lime': 'Verde', 'Verde mint': 'Verde',
    'Vernil': 'Verde', 'Visiniu': 'Visiniu', 'floral': 'Multicolor', 'Fucsia': 'Roz',
}
BASE_COLOR_MAP = {
    'Alb': 'Alb', 'Albastru': 'Albastru', 'Albastru deschis': 'Albastru',
    'Albastru inchis': 'Albastru', 'Albastru petrol': 'Albastru',
    'Animal print': 'Multicolor', 'Aramiu': 'Auriu', 'Argintiu': 'Argintiu',
    'Auriu': 'Auriu', 'Bej': 'Bej', 'Bleumarin': 'Albastru', 'Bordo': 'Rosu',
    'Caramel': 'Maro', 'Ciocolatiu': 'Maro', 'Corai': 'Roz', 'Crem': 'Bej',
    'Galben': 'Galben', 'Gri': 'Gri', 'Kaki': 'Verde', 'Lavanda': 'Mov',
    'Lila': 'Mov', 'Maro': 'Maro', 'Mov': 'Mov', 'Multicolor': 'Multicolor',
    'Negru': 'Negru', 'Nude': 'Bej', 'Olive': 'Verde', 'Piersica': 'Roz',
    'Portocaliu': 'Portocaliu', 'Pudra': 'Roz', 'Rosu': 'Rosu', 'Roz': 'Roz',
    'Roz prafuit': 'Roz', 'Somon': 'Roz', 'Turcoaz': 'Albastru', 'Verde': 'Verde',
    'Verde inchis': 'Verde', 'Verde lime': 'Verde', 'Verde mint': 'Verde',
    'Vernil': 'Verde', 'Visiniu': 'Rosu', 'floral': 'Multicolor', 'Fucsia': 'Roz',
}
MATERIAL_MAP = {
    'Acryl': 'Acril', 'Barbie': 'Sintetic', 'Brocart': 'Sintetic', 'Bumbac': 'Bumbac',
    'Catifea': 'Catifea', 'Casmir': 'Lana', 'Crep': 'Sintetic', 'Dantela': 'Dantela',
    'Jerseu': 'Tricot', 'Lana': 'Lana', 'Lycra': 'Elastan', 'Licra': 'Elastan',
    'Matase': 'Matase', 'Neopren': 'Sintetic', 'Organza': 'Sintetic', 'Paiete': 'Sintetic',
    'Piele ecologica': 'Piele ecologica', 'Poliester': 'Poliester', 'Satin': 'Satin',
    'Sifon': 'Sintetic', 'Stofa': 'Lana', 'Tafta': 'Sintetic', 'Tricot': 'Tricot',
    'Tul': 'Tul', 'Tweed': 'Lana', 'Velur': 'Catifea', 'Voal': 'Sintetic',
}
LUNGIME_MAP = {'Lungi': 'Lunga', 'Medii': 'Midi', 'Scurte': 'Scurta'}
CROIALA_MAP = {
    'In clini': 'Evazat', 'Lejer': 'Lejer', 'Mulat': 'Bodycon', 'Peplum': 'Cambrata',
    'Petrecuta': 'Petrecuta', 'Plisat': 'Plisata', 'Pliuri': 'Plisata',
    'Volane': 'Evazat', 'in A': 'Evazat',
}
STIL_MAP = {
    'Asimetrica': 'Elegant', 'Birou': 'Casual', 'Casual': 'Casual',
    'Casual-Elegant': 'Casual', 'Casual-Office': 'Casual',
    'Cu crapatura': 'Elegant', 'De ocazie': 'De ocazie',
    'De seara': 'Elegant', 'Eleganta': 'Elegant', 'Sport': 'Casual',
    'Lunga sirena': 'Elegant',
}
SIZE_MAP = {
    'S': 'S INTL', 'M': 'M INTL', 'L': 'L INTL', 'XL': 'XL INTL',
    'XXL': '2XL INTL', '2XL': '2XL INTL', 'XXXL': '3XL INTL', '3XL': '3XL INTL',
    'XS': 'XS INTL', '4XL': '4XL INTL', '5XL': '5XL INTL',
    'S/M': 'S INTL', 'M/L': 'M INTL', 'L/XL': 'L INTL',
    'ONE SIZE': 'One Size INTL', 'UNIC': 'One Size INTL',
    '34': '34 EU', '36': '36 EU', '38': '38 EU', '40': '40 EU',
    '42': '42 EU', '44': '44 EU', '46': '46 EU', '48': '48 EU',
    '50': '50 EU', '52': '52 EU',
}

# Column positions in Template sheet (1-indexed)
COL = {
    'part_number': 1, 'vendor_ext_id': 2, 'ean': 3,
    'sale_price': 4, 'original_sale_price': 5, 'vat_rate': 6,
    'status': 7, 'offer_currency': 8, 'stock': 9,
    'handling_time': 10, 'lead_time': 11, 'warranty': 12,
    'offer_properties': 13, 'min_sale_price': 14, 'max_sale_price': 15,
    'name': 16, 'brand': 17, 'description': 18, 'url': 19,
    'source_language': 20, 'main_image_url': 21,
    'other_image_url1': 22, 'other_image_url2': 23,
    'other_image_url3': 24, 'other_image_url4': 25, 'other_image_url5': 26,
    'family_id': 27, 'family_name': 28, 'family_type': 29,
    'size_original': 30, 'size_converted': 31,
    'culoare': 32, 'material': 33, 'colectie': 34,
    'stil': 35, 'lungime': 36, 'lungime_maneca': 37,
    'imprimeu': 38, 'detalii': 39, 'sistem_inchidere': 40,
    'culoare_baza': 41, 'captuseala': 42, 'decolteu': 43,
    'buzunare': 44, 'exterior': 45, 'barete': 46,
    'material_garnituri': 47, 'talie': 48, 'pentru': 49,
    'croiala': 50, 'linie_brand': 51, 'marime_convertita': 52, 'fit': 53,
    'safety_info': 54, 'manufacturer_name': 55,
    'manufacturer_address': 56, 'manufacturer_email': 57,
    'responsible_name': 58, 'responsible_address': 59,
    'responsible_email': 60,
}


def fetch_all_products():
    all_products = {}
    page = 1
    while True:
        url = f'{API_BASE}?produse&apikey={API_KEY}&pagina={page}&limit={PAGE_SIZE}'
        try:
            r = requests.get(url, headers=HEADERS, timeout=60)
            data = r.json()
        except Exception as e:
            print(f'  Error page {page}: {e}')
            break
        if not data or (isinstance(data, dict) and data.get('eroare')):
            break
        if isinstance(data, dict):
            count = 0
            for pid, pdata in data.items():
                if isinstance(pdata, dict) and 'id_produs' in pdata:
                    all_products[pid] = pdata
                    count += 1
            print(f'  Page {page}: {count} products')
            if count < PAGE_SIZE:
                break
        else:
            break
        page += 1
        time.sleep(0.3)
    return all_products


def load_barcode_map():
    with open(BARCODE_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def build_product_barcode_map(barcode_map):
    product_barcodes = {}
    for barcode, pid in barcode_map.items():
        pid = str(pid)
        if pid not in product_barcodes:
            product_barcodes[pid] = []
        product_barcodes[pid].append(barcode)
    return product_barcodes


def get_spec(product, spec_name):
    specs = product.get('specificatii', [])
    if isinstance(specs, list):
        for s in specs:
            if isinstance(s, dict) and s.get('nume', '').lower() == spec_name.lower():
                vals = s.get('valoare', [])
                if isinstance(vals, list) and vals:
                    return vals[0]
    return ''


def map_val(value, mapping, default=''):
    if not value:
        return default
    if value in mapping:
        return mapping[value]
    for k, v in mapping.items():
        if k.lower() == value.lower():
            return v
    return default or value


def extract_color_from_name(name):
    colors_map = {
        'negru': 'Negru', 'neagra': 'Negru', 'alb': 'Alb', 'alba': 'Alb',
        'rosu': 'Rosu', 'rosie': 'Rosu', 'verde': 'Verde', 'albastru': 'Albastru',
        'albastra': 'Albastru', 'roz': 'Roz', 'mov': 'Mov', 'galben': 'Galben',
        'gri': 'Gri', 'maro': 'Maro', 'bej': 'Bej', 'crem': 'Crem',
        'bordo': 'Bordo', 'fucsia': 'Fucsia', 'turcoaz': 'Turcoaz', 'corai': 'Corai',
        'nude': 'Nude', 'auriu': 'Auriu', 'argintiu': 'Argintiu', 'bleumarin': 'Bleumarin',
        'lila': 'Lila', 'lavanda': 'Lavanda', 'kaki': 'Kaki', 'portocaliu': 'Portocaliu',
        'somon': 'Somon', 'visiniu': 'Visiniu', 'multicolor': 'Multicolor',
    }
    name_lower = name.lower()
    for pattern, color in colors_map.items():
        if pattern in name_lower:
            return color
    return ''


def generate_rows(products, barcode_map, allowed_brands, limit=None):
    product_barcodes = build_product_barcode_map(barcode_map)
    rows = []
    stats = {'processed': 0, 'no_barcode': 0, 'no_stock': 0, 'no_brand': 0}

    for pid, product in products.items():
        brand_data = product.get('brand', {})
        brand_name = brand_data.get('nume', '') if isinstance(
            brand_data, dict) else str(brand_data)

        if allowed_brands and brand_name.lower() not in allowed_brands:
            stats['no_brand'] += 1
            continue

        pid = str(product.get('id_produs', pid))
        name = product.get('nume', '')
        if not name:
            continue

        optiuni = product.get('optiuni', {})
        sizes_with_stock = []
        if isinstance(optiuni, dict):
            for opt_id, opt_data in optiuni.items():
                if isinstance(opt_data, dict):
                    size_name = opt_data.get('nume_optiune', '')
                    try:
                        stock = int(opt_data.get('stoc_fizic', 0))
                    except (ValueError, TypeError):
                        stock = 0
                    if stock > 0 and size_name:
                        sizes_with_stock.append(
                            {'name': size_name, 'stock': stock})

        if not sizes_with_stock:
            stats['no_stock'] += 1
            continue

        barcodes = product_barcodes.get(pid, [])
        if not barcodes:
            stats['no_barcode'] += 1
            continue

        code = product.get('cod_produs', f'FBR-{pid}')
        description = product.get('descriere', '')
        url = product.get('link', '')
        cota_tva = product.get('cota_tva', 19)

        images = product.get('imagini', [])
        if not images:
            main_img = product.get('imagine', '')
            images = [main_img] if main_img else []

        price_str = product.get('pret_discount', '0') or '0'
        try:
            price_with_vat = float(str(price_str).replace(' RON', '').strip())
        except ValueError:
            price_with_vat = 0
        if price_with_vat <= 0:
            price_str = product.get('pret', '0') or '0'
            try:
                price_with_vat = float(
                    str(price_str).replace(' RON', '').strip())
            except ValueError:
                price_with_vat = 0
        if price_with_vat <= 0:
            continue

        price_no_vat = round(price_with_vat / (1 + cota_tva / 100), 2)
        min_price = round(price_no_vat * 0.75, 2)
        max_price = round(price_no_vat * 2.0, 2)

        culoare = get_spec(product, 'Culoare') or extract_color_from_name(name)
        material = get_spec(product, 'Material')
        lungime = get_spec(product, 'Lungime')
        croi = get_spec(product, 'Croi')
        stil = get_spec(product, 'Stil')

        fd_culoare = map_val(culoare, COLOR_MAP, 'Negru')
        fd_base_color = map_val(culoare, BASE_COLOR_MAP, 'Negru')
        fd_material = map_val(material, MATERIAL_MAP, 'Poliester')
        fd_lungime = map_val(lungime, LUNGIME_MAP, '')
        fd_croiala = map_val(croi, CROIALA_MAP, '')
        fd_stil = map_val(stil, STIL_MAP, 'Elegant')

        for size_idx, size_info in enumerate(sizes_with_stock):
            ean = barcodes[size_idx] if size_idx < len(
                barcodes) else barcodes[-1]
            fd_size = SIZE_MAP.get(
                size_info['name'].upper().strip(), size_info['name'] + ' EU')

            row = {
                'part_number': code,
                'vendor_ext_id': int(pid),
                'ean': ean,
                'sale_price': price_no_vat,
                'original_sale_price': None,
                'vat_rate': cota_tva,
                'status': 1,
                'offer_currency': 'RON',
                'stock': size_info['stock'],
                'handling_time': 1,
                'lead_time': 14,
                'warranty': 0,
                'offer_properties': None,
                'min_sale_price': min_price,
                'max_sale_price': max_price,
                'name': name[:255],
                'brand': brand_name,
                'description': description,
                'url': url,
                'source_language': 'ro_RO',
                'main_image_url': images[0] if images else '',
                'other_image_url1': images[1] if len(images) > 1 else None,
                'other_image_url2': images[2] if len(images) > 2 else None,
                'other_image_url3': images[3] if len(images) > 3 else None,
                'other_image_url4': images[4] if len(images) > 4 else None,
                'other_image_url5': images[5] if len(images) > 5 else None,
                'family_id': int(pid),
                'family_name': name[:100],
                'family_type': 'Marime',
                'size_original': fd_size,
                'size_converted': fd_size,
                'culoare': fd_culoare,
                'material': fd_material,
                'colectie': '2026',
                'stil': fd_stil,
                'lungime': fd_lungime,
                'lungime_maneca': None,
                'imprimeu': 'Uni',
                'detalii': None,
                'sistem_inchidere': None,
                'culoare_baza': fd_base_color,
                'captuseala': None,
                'decolteu': None,
                'buzunare': None,
                'exterior': None,
                'barete': None,
                'material_garnituri': None,
                'talie': None,
                'pentru': 'Femei',
                'croiala': fd_croiala,
                'linie_brand': brand_name,
                'marime_convertita': fd_size,
                'fit': None,
                'safety_info': None,
                'manufacturer_name': 'FABREX SRL',
                'manufacturer_address': 'Chiajna, str. Rezervelor nr. 64B, jud. Ilfov, Romania',
                'manufacturer_email': 'alextdr2016@gmail.com',
                'responsible_name': None,
                'responsible_address': None,
                'responsible_email': None,
            }
            rows.append(row)

        stats['processed'] += 1
        if limit and stats['processed'] >= limit:
            break

    print(f"\nResults:")
    print(f"  Products processed: {stats['processed']}")
    print(f"  Rows generated: {len(rows)} (one per size)")
    print(f"  Skipped (wrong brand): {stats['no_brand']}")
    print(f"  Skipped (no stock): {stats['no_stock']}")
    print(f"  Skipped (no barcode): {stats['no_barcode']}")
    return rows


def write_to_template(rows, template_path, output_path):
    """Write rows into the official Fashion Days template, preserving everything"""
    print(f"\nLoading official template: {template_path}")
    wb = load_workbook(template_path)
    ws = wb['Template']

    print(f"  Template sheets: {wb.sheetnames}")
    print(f"  Writing {len(rows)} rows starting at row {START_ROW}")

    for row_idx, row_data in enumerate(rows):
        excel_row = START_ROW + row_idx
        for key, col_num in COL.items():
            value = row_data.get(key)
            if value is not None:
                ws.cell(row=excel_row, column=col_num, value=value)

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"  Template preserved: all sheets, validations, formatting intact")
    print(
        f"  Data rows written: {len(rows)} (rows {START_ROW} to {START_ROW + len(rows) - 1})")


def main():
    parser = argparse.ArgumentParser(
        description='Export produse pentru Fashion Days (scrie in template oficial)')
    parser.add_argument('--template', required=True,
                        help='Calea catre template-ul oficial descarcat din Fashion Days')
    parser.add_argument('--brand', default='ejolie,artista',
                        help='Brand filter, comma-separated')
    parser.add_argument('--limit', type=int, default=None,
                        help='Limit number of products')
    parser.add_argument('--dry-run', action='store_true',
                        help='Only show stats')
    parser.add_argument('--output', default=None,
                        help='Output file path (default: fashiondays_filled.xlsx)')
    args = parser.parse_args()

    if not os.path.exists(args.template):
        print(f"ERROR: Template file not found: {args.template}")
        sys.exit(1)

    output_path = args.output or os.path.join(
        SCRIPT_DIR, 'fashiondays_filled.xlsx')

    print("=" * 60)
    print("Fashion Days Export v3 - Scrie in template oficial")
    print("=" * 60)

    allowed_brands = [b.strip().lower() for b in args.brand.split(',')]
    print(f"\nBrands: {', '.join(allowed_brands)}")
    print(f"Template: {args.template}")

    print(f"\nFetching products from Extended API...")
    products = fetch_all_products()
    print(f"  Total: {len(products)} products")

    print(f"\nLoading barcode map from: {BARCODE_PATH}")
    barcode_map = load_barcode_map()
    print(f"  Total: {len(barcode_map)} barcodes")

    rows = generate_rows(products, barcode_map,
                         allowed_brands, limit=args.limit)

    if args.dry_run:
        print("\n[DRY RUN] No file written.")
        if rows:
            print(f"\nSample row:")
            for key in ['part_number', 'vendor_ext_id', 'ean', 'sale_price', 'vat_rate',
                        'stock', 'name', 'brand', 'size_original', 'culoare', 'material']:
                print(f"  {key}: {rows[0].get(key)}")
        return

    write_to_template(rows, args.template, output_path)
    print("\nDone!")


if __name__ == '__main__':
    main()

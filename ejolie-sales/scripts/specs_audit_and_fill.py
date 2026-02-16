#!/usr/bin/env python3
"""
EJOLIE.RO - Audit & Auto-Fill Product Specifications
=====================================================
Script care:
1. Exporta produsele din Extended API (folosind aceeasi logica ca export_trendyol.py)
2. Identifica produsele fara specificatii completate
3. Foloseste GPT pentru a genera specificatii lipsa
4. Genereaza fisier Excel pentru import in Extended

Specificatii tracked: Culoare, Material, Lungime, Croi, Stil, Model

Utilizare:
  python3 specs_audit_and_fill.py --audit          # Doar audit
  python3 specs_audit_and_fill.py --fill --limit 5  # Completeaza cu GPT (test 5)
  python3 specs_audit_and_fill.py --fill             # Completeaza toate

Autor: Claude AI pentru Alex Tudor
Data: 16 Februarie 2026
"""

import os
import sys
import json
import time
import argparse
import urllib.request
from pathlib import Path

# ============================================================
# CONFIGURARE - folosim aceleasi variabile ca celelalte scripturi
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Incarca .env manual


def load_dotenv(path):
    if not os.path.exists(path):
        return
    with open(path) as f:
        for line in f:
            line = line.strip()
            if '=' in line and not line.startswith('#'):
                k, v = line.split('=', 1)
                os.environ.setdefault(
                    k.strip(), v.strip().strip('"').strip("'"))


# Cauta .env
for env_path in [
    os.path.join(SCRIPT_DIR, '..', '.env'),
    os.path.join(SCRIPT_DIR, '.env'),
    os.path.expanduser('~/ejolie-openclaw-agent/ejolie-sales/.env'),
]:
    if os.path.exists(env_path):
        load_dotenv(env_path)
        print(f"📁 .env din: {env_path}")
        break

# Aceleasi variabile ca in export_trendyol.py si utils.py
API_KEY = os.environ.get("EJOLIE_API_KEY", "")
API_URL = os.environ.get("EJOLIE_API_URL", "https://ejolie.ro/api/")
OPENAI_KEY = os.environ.get("OPENAI_API_KEY", "")
GPT_MODEL = "gpt-4o-mini"

# Specificatiile pe care le tracked
SPEC_NAMES = ["Culoare", "Material", "Lungime", "Croi", "Stil", "Model"]

# Valorile VALIDE pentru fiecare specificatie (exact cum sunt in Extended)
VALID_VALUES = {
    "Culoare": [
        "Alb", "Albastru", "Albastru deschis", "Albastru inchis", "Albastru petrol",
        "Animal print", "Aramiu", "Argintiu", "Auriu", "Bej", "Bleumarin",
        "Bordo", "Caramel", "Ciocolatiu", "Corai", "Crem", "Galben", "Gri",
        "Kaki", "Lavanda", "Lila", "Maro", "Mov", "Multicolor", "Negru",
        "Nude", "Olive", "Piersica", "Portocaliu", "Pudra", "Rosu", "Roz",
        "Somon", "Turcoaz", "Verde", "Verde inchis", "Verde lime", "Verde mint",
        "Vernil", "Visiniu", "floral"
    ],
    "Material": [
        "Acryl", "Barbie", "Brocart", "Bumbac", "Catifea", "Casmir", "Crep",
        "Dantela", "Jerseu", "Lana", "Lycra", "Matase", "Neopren", "Organza",
        "Paiete", "Piele ecologica", "Poliester", "Satin", "Sifon", "Stofa",
        "Tafta", "Tricot", "Tul", "Tweed", "Velur", "Voal"
    ],
    "Lungime": ["Lungi", "Medii", "Scurte"],
    "Croi": [
        "In clini", "Lejer", "Mulat", "Peplum", "Petrecuta",
        "Plisat", "Pliuri", "Volane", "in A"
    ],
    "Stil": [
        "Asimetrica", "Birou", "Casual", "Casual-Elegant", "Casual-Office",
        "Cu crapatura", "De ocazie", "De seara", "Eleganta", "Sport"
    ],
    "Model": [
        "Accesorizata la baza gatului", "Aplicatii 3D", "Bretele reglabile",
        "Broderie", "Brosa", "Buzunare", "Cambrat", "Captuseala",
        "Centura inclusa", "Cordon", "Cu buzunare", "Cu captuseala",
        "Cu cordon", "Cu fermoar", "Cu funda", "Cu gluga", "Cu nasturi",
        "Decolteu in V", "Fara maneci", "Maneci lungi", "Maneci scurte",
        "Rochie camasa", "Umeri goi", "Un umar gol"
    ]
}


# ============================================================
# API FUNCTIONS - exact ca in export_trendyol.py
# ============================================================

def api_fetch(url):
    """Fetch URL cu User-Agent (asa functioneaza Extended API)"""
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    resp = urllib.request.urlopen(req, timeout=180)
    return json.loads(resp.read().decode("utf-8"))


def fetch_product_ids(brand="ejolie"):
    """Pas 1: Preia lista de ID-uri produse (ca export_trendyol.py)"""
    url = f"{API_URL}?produse&brand={brand}&apikey={API_KEY}"
    print(f"   URL: {url[:80]}...")
    data = api_fetch(url)

    if isinstance(data, list):
        # API returneaza lista de ID-uri sau lista de produse
        if data and isinstance(data[0], dict):
            ids = [str(p.get('id', p.get('id_produs', ''))) for p in data]
        else:
            ids = [str(x) for x in data]
    elif isinstance(data, dict):
        ids = list(data.keys())
    else:
        ids = []

    return ids


def fetch_products_batch(ids):
    """Pas 2: Preia detalii produse in batch (ca export_trendyol.py)"""
    products = []
    batch_size = 20

    for i in range(0, len(ids), batch_size):
        batch = ids[i:i+batch_size]
        url = f"{API_URL}?produse&id_produse={','.join(batch)}&apikey={API_KEY}"

        try:
            data = api_fetch(url)
            if isinstance(data, list):
                products.extend(data)
            elif isinstance(data, dict):
                products.extend(data.values())
            print(
                f"   Batch {i//batch_size + 1}: {len(batch)} IDs → {len(products)} total")
        except Exception as e:
            print(f"   ❌ Eroare batch {i//batch_size + 1}: {e}")

        time.sleep(0.5)

    return products


# ============================================================
# SPECS ANALYSIS
# ============================================================

def analyze_specs(product):
    """Analizeaza ce specificatii are/nu are un produs"""
    specs = product.get('specificatii', [])

    existing = {}
    if isinstance(specs, list):
        for spec in specs:
            name = spec.get('nume', '')
            values = spec.get('valoare', [])
            if values and values[0]:
                existing[name] = values
    elif isinstance(specs, dict):
        for name, values in specs.items():
            if values:
                existing[name] = values if isinstance(
                    values, list) else [values]

    missing = [s for s in SPEC_NAMES if s not in existing]
    return existing, missing


# ============================================================
# GPT - exact ca in export_trendyol.py (urllib, nu openai lib)
# ============================================================

def gpt_generate_specs(product_name, product_description, missing_specs):
    """Foloseste GPT pentru a genera specificatiile lipsa"""
    if not OPENAI_KEY:
        return {}

    valid_options = ""
    for spec in missing_specs:
        if spec in VALID_VALUES:
            valid_options += f"\n{spec} - alege DOAR din: {', '.join(VALID_VALUES[spec])}"

    prompt = f"""Esti expert fashion e-commerce. Analizeaza produsul si completeaza specificatiile lipsa.

PRODUS: {product_name}
DESCRIERE: {str(product_description)[:500]}

SPECIFICATII DE COMPLETAT (alege EXACT din valorile date):
{valid_options}

REGULI:
- Alege DOAR valori din listele de mai sus
- Pentru Culoare: extrage din numele produsului
- Lungime: Lungi=maxi/lunga, Medii=midi, Scurte=mini
- Daca nu esti sigur, pune cea mai probabila valoare

Raspunde DOAR JSON, fara alte texte:
{{{', '.join(f'"{s}": "valoare"' for s in missing_specs)}}}"""

    url = "https://api.openai.com/v1/chat/completions"
    payload = json.dumps({
        "model": GPT_MODEL,
        "messages": [
            {"role": "system", "content": "Raspunzi DOAR in JSON valid. Esti expert fashion."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.3,
        "max_tokens": 200
    }).encode("utf-8")

    req = urllib.request.Request(url, data=payload, headers={
        "Content-Type": "application/json",
        "Authorization": f"Bearer {OPENAI_KEY}"
    })

    try:
        resp = urllib.request.urlopen(req, timeout=30)
        result = json.loads(resp.read().decode("utf-8"))
        text = result["choices"][0]["message"]["content"].strip()
        text = text.replace('```json', '').replace('```', '').strip()
        specs = json.loads(text)

        # Valideaza valorile
        validated = {}
        for name, value in specs.items():
            if name in VALID_VALUES:
                if value in VALID_VALUES[name]:
                    validated[name] = value
                else:
                    closest = find_closest(value, VALID_VALUES[name])
                    if closest:
                        validated[name] = closest
        return validated
    except Exception as e:
        print(f"      ❌ GPT eroare: {e}")
        return {}


def find_closest(value, valid_list):
    """Gaseste cea mai apropiata valoare"""
    vl = value.lower().strip()
    for v in valid_list:
        if v.lower() == vl:
            return v
    for v in valid_list:
        if vl in v.lower() or v.lower() in vl:
            return v
    return None


# ============================================================
# EXCEL GENERATION
# ============================================================

def generate_excel(results, output_path, mode="audit"):
    """Genereaza Excel cu raport + fisier import"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # === Sheet 1: Raport Audit ===
    ws = wb.active
    ws.title = "Raport Audit"

    headers = ["ID", "Cod Produs", "Nume Produs",
               "Categorie"] + SPEC_NAMES + ["Lipsa", "Status"]

    hfill = PatternFill(start_color="4472C4",
                        end_color="4472C4", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    green = PatternFill(start_color="C6EFCE",
                        end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C",
                         end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE",
                      end_color="FFC7CE", fill_type="solid")

    for row_idx, prod in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=prod.get('id', ''))
        ws.cell(row=row_idx, column=2, value=prod.get('cod_produs', ''))
        ws.cell(row=row_idx, column=3, value=prod.get('nume', '')[:50])
        ws.cell(row=row_idx, column=4, value=prod.get('categorie', ''))

        existing = prod.get('specs_existing', {})
        new_specs = prod.get('specs_new', {})
        missing = prod.get('specs_missing', [])

        for i, spec_name in enumerate(SPEC_NAMES):
            val = ''
            fill = red

            if spec_name in existing:
                val = ', '.join(existing[spec_name]) if isinstance(
                    existing[spec_name], list) else existing[spec_name]
                fill = green
            elif spec_name in new_specs:
                val = new_specs[spec_name]
                fill = yellow
            else:
                val = '❌ LIPSA'

            cell = ws.cell(row=row_idx, column=5 + i, value=val)
            cell.fill = fill

        ws.cell(row=row_idx, column=11, value=len(missing))
        ws.cell(row=row_idx, column=12,
                value="✅ OK" if not missing else f"⚠️ {len(missing)} lipsa")

    # === Sheet 2: Import Extended (doar produsele cu specs noi) ===
    if mode == "fill":
        ws2 = wb.create_sheet("Import Extended")

        import_headers = [
            "Nume produs", "Descriere", "Categorie", "Brand",
            "Optiune 1", "Optiune 2", "Optiune 3", "Optiune 4", "Optiune 5",
            "Furnizor", "Pret vanzare", "Pret intrare", "Adaos %", "Discount %",
            "Moneda", "Cod produs", "Stoc", "Stoc fizic", "Greutate (KG)"
        ]
        for i in range(1, 16):
            import_headers.append(f"Imagine {i}")
        for spec in SPEC_NAMES:
            import_headers.append(spec)

        ifill = PatternFill(start_color="70AD47",
                            end_color="70AD47", fill_type="solid")
        for col, h in enumerate(import_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = ifill
            cell.font = hfont

        row_idx = 2
        for prod in results:
            new_specs = prod.get('specs_new', {})
            if not new_specs:
                continue

            ws2.cell(row=row_idx, column=1, value=prod.get('nume', ''))
            ws2.cell(row=row_idx, column=4, value=prod.get('brand', ''))
            ws2.cell(row=row_idx, column=16, value=prod.get('cod_produs', ''))

            # Specificatii - coloana 35+ (dupa 19 cols + 15 imagini)
            spec_start = 20 + 15  # col 35
            for i, spec_name in enumerate(SPEC_NAMES):
                # Pune existent sau nou
                existing = prod.get('specs_existing', {})
                if spec_name in existing:
                    val = existing[spec_name][0] if isinstance(
                        existing[spec_name], list) else existing[spec_name]
                elif spec_name in new_specs:
                    val = new_specs[spec_name]
                else:
                    val = ''
                ws2.cell(row=row_idx, column=spec_start + i, value=val)

            row_idx += 1

    # Auto-width
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            sheet.column_dimensions[col[0].column_letter].width = min(
                max_len + 2, 40)

    wb.save(output_path)
    print(f"\n📊 Excel salvat: {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description='Ejolie.ro - Audit & Fill Specs')
    parser.add_argument('--audit', action='store_true',
                        help='Audit specificatii')
    parser.add_argument('--fill', action='store_true',
                        help='Completeaza cu GPT')
    parser.add_argument('--limit', type=int, default=0,
                        help='Limita produse (0=toate)')
    parser.add_argument('--brand', type=str, default='ejolie', help='Brand')
    parser.add_argument('--in-stock', action='store_true', default=True,
                        help='Doar produse cu stoc > 0 (default: True)')
    parser.add_argument('--all-stock', action='store_true',
                        help='Include si produse fara stoc')
    parser.add_argument('--output', type=str, default='',
                        help='Output Excel path')
    args = parser.parse_args()

    if not args.audit and not args.fill:
        args.audit = True

    print("=" * 60)
    print("🔍 EJOLIE.RO - Audit & Fill Product Specifications")
    print("=" * 60)

    if not API_KEY:
        print("❌ EJOLIE_API_KEY nu e setat in .env!")
        sys.exit(1)

    print(f"✅ API Key: ...{API_KEY[-6:]}")
    if OPENAI_KEY:
        print(f"✅ OpenAI Key: ...{OPENAI_KEY[-6:]}")

    # Pas 1: Preia ID-uri produse
    print(f"\n📦 Pas 1: Preiau ID-uri produse (brand={args.brand})...")
    ids = fetch_product_ids(args.brand)
    print(f"   Gasite: {len(ids)} ID-uri")

    if not ids:
        print("❌ Nu am gasit produse! Verifica API key si brand.")
        sys.exit(1)

    # Pas 2: Preia detalii produse in batch
    print(f"\n📦 Pas 2: Preiau detalii produse...")
    products = fetch_products_batch(ids)
    print(f"   Total produse cu detalii: {len(products)}")

    # Pas 2.5: Filtreaza dupa stoc
    if not args.all_stock:
        before = len(products)
        filtered = []
        for p in products:
            stoc = p.get('stoc', '')
            stoc_fizic = p.get('stoc_fizic', 0)

            # Verifica stoc: "in stoc", "In stoc" sau stoc_fizic > 0
            has_stock = False
            if isinstance(stoc_fizic, (int, float)) and stoc_fizic > 0:
                has_stock = True
            elif isinstance(stoc_fizic, str) and stoc_fizic.isdigit() and int(stoc_fizic) > 0:
                has_stock = True
            elif isinstance(stoc, str) and 'stoc' in stoc.lower() and 'fara' not in stoc.lower() and 'indisponibil' not in stoc.lower():
                has_stock = True

            if has_stock:
                filtered.append(p)

        products = filtered
        print(f"   🏷️ Filtru stoc > 0: {before} → {len(products)} produse")

    # Aplica limita
    if args.limit > 0:
        products = products[:args.limit]
        print(f"   ⚡ Limitat la {args.limit} produse")

    # Pas 3: Analizeaza specificatii
    print(f"\n🔎 Pas 3: Analizez specificatii...")

    results = []
    stats = {"complete": 0, "missing": 0}
    missing_per_spec = {s: 0 for s in SPEC_NAMES}

    for i, prod in enumerate(products):
        name = prod.get('nume', prod.get('nume_scurt', 'N/A'))
        cod = prod.get('cod_produs', '')
        pid = prod.get('id', prod.get('id_produs', ''))
        brand = prod.get('brand', '')
        if isinstance(brand, dict):
            brand = brand.get('nume', '')

        categorie = ''
        cats = prod.get('categorii', [])
        if isinstance(cats, list) and cats:
            categorie = cats[0].get('nume', '') if isinstance(
                cats[0], dict) else str(cats[0])
        elif isinstance(cats, dict):
            first = list(cats.values())[0] if cats else {}
            categorie = first.get('nume', '') if isinstance(
                first, dict) else str(first)

        existing, missing = analyze_specs(prod)

        result = {
            'id': pid,
            'cod_produs': cod,
            'nume': name,
            'brand': brand,
            'categorie': categorie,
            'descriere': prod.get('continut', prod.get('descriere', '')),
            'specs_existing': existing,
            'specs_missing': missing,
            'specs_new': {}
        }

        if missing:
            stats["missing"] += 1
            for s in missing:
                missing_per_spec[s] += 1

            if args.fill and OPENAI_KEY:
                print(
                    f"   [{i+1}/{len(products)}] 🤖 {name[:40]}... → {', '.join(missing)}")
                new_specs = gpt_generate_specs(
                    name, result['descriere'], missing)
                result['specs_new'] = new_specs
                if new_specs:
                    print(f"      ✅ {new_specs}")
                time.sleep(0.5)
            else:
                if i < 10 or i % 50 == 0:
                    print(
                        f"   [{i+1}/{len(products)}] ⚠️ {name[:40]}... → lipsa: {', '.join(missing)}")
        else:
            stats["complete"] += 1

        results.append(result)

    # Pas 4: Raport
    print("\n" + "=" * 60)
    print("📊 RAPORT FINAL")
    print("=" * 60)
    print(f"   Total produse: {len(results)}")
    print(f"   ✅ Complete: {stats['complete']}")
    print(f"   ⚠️ Cu lipsa: {stats['missing']}")

    if stats['missing'] > 0:
        print(f"\n   Per specificatie:")
        for spec in SPEC_NAMES:
            cnt = missing_per_spec[spec]
            pct = (cnt / len(results)) * 100 if results else 0
            bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
            print(f"   {spec:12s}: {cnt:4d} lipsa ({pct:5.1f}%) {bar}")

    if args.fill:
        filled = sum(1 for r in results if r['specs_new'])
        print(f"\n   🤖 GPT a completat: {filled} produse")

    # Pas 5: Excel
    output = args.output or os.path.expanduser(
        f'~/ejolie_specs_{"fill" if args.fill else "audit"}.xlsx'
    )
    generate_excel(results, output, "fill" if args.fill else "audit")

    print(f"\n🎯 Gata! Fisierul: {output}")
    if args.fill:
        print(f"   Importa in Extended: Manager → Actualizari → Import produse")
    print()


if __name__ == '__main__':
    main()

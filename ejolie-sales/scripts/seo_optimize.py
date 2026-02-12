#!/usr/bin/env python3
"""Generate SEO meta title + description suggestions using GPT"""

import os, json, argparse, urllib.request, time

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SEO_CACHE = os.path.join(SCRIPT_DIR, "seo_cache.json")
STOCK_CACHE = os.path.join(SCRIPT_DIR, "stock_cache.json")

# Load env
env_path = os.path.join(SCRIPT_DIR, ".env")
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            if '=' in line and not line.startswith('#'):
                k, v = line.strip().split('=', 1)
                os.environ[k] = v

OPENAI_KEY = os.environ.get("OPENAI_API_KEY", "")


def call_gpt(prompt, model="gpt-4o-mini"):
    """Call OpenAI API"""
    url = "https://api.openai.com/v1/chat/completions"
    data = json.dumps({
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.7,
        "max_tokens": 500,
    }).encode("utf-8")
    
    req = urllib.request.Request(url, data=data, headers={
        "Authorization": f"Bearer {OPENAI_KEY}",
        "Content-Type": "application/json",
    })
    
    resp = urllib.request.urlopen(req, timeout=30)
    result = json.loads(resp.read().decode("utf-8"))
    return result["choices"][0]["message"]["content"].strip()


def generate_seo_suggestions(max_products=None):
    """Generate SEO suggestions for products with low scores"""
    
    if not os.path.exists(SEO_CACHE):
        print("❌ SEO cache lipsă! Rulează mai întâi seo_audit.py")
        return []
    
    with open(SEO_CACHE) as f:
        audit = json.load(f)
    
    results = []
    count = 0
    
    for r in sorted(audit["results"], key=lambda x: x["score"]):
        seo = r.get("seo_data", {})
        meta_title = seo.get("meta_title", "")
        meta_desc = seo.get("meta_description", "")
        
        # Only optimize products with short meta title or missing meta desc
        needs_fix = len(meta_title) < 45 or len(meta_desc) < 100
        if not needs_fix:
            continue
        
        count += 1
        if max_products and count > max_products:
            break
        
        print(f"  🤖 [{count}] {r['name'][:50]}...", end=" ", flush=True)
        
        prompt = f"""Ești expert SEO pentru ejolie.ro, magazin online de rochii elegante din România.
Generează meta title și meta description OPTIMIZATE SEO.

Produs: {r['name']}
URL: {r['link']}
Meta title actual: {meta_title}
Meta description actual: {meta_desc[:200]}

REGULI STRICTE:
1. Meta title: OBLIGATORIU între 50-60 caractere (numără exact!)
   - Format: "[Tip] [Nume] [Culoare] | Ejolie.ro"  
   - Exemplu: "Rochie de Seară Elysia Neagră cu Trenă | Ejolie.ro" (51 chars)
   - Exemplu: "Rochie Elegantă Tea Verde de Ocazie | Ejolie.ro" (48 chars)
   - NU adăuga categorii lungi după | doar "Ejolie.ro"
   - MAXIM 60 caractere! Dacă depășești, scurtează.

2. Meta description: OBLIGATORIU între 135-155 caractere (numără exact!)
   - Include: beneficiu + keyword + call-to-action
   - Exemplu bun: "Rochie de seară Elysia neagră cu trenă din voal. Croială sirenă care modelează silueta. Comandă online cu livrare rapidă în toată România."
   - Folosește: "Comandă online", "Livrare rapidă", "Colecție nouă 2026"

3. Scrie în ROMÂNĂ cu diacritice (ă, â, î, ș, ț)
4. NU repeta cuvinte
5. Fii specific la tip: rochie de seară, rochie cocktail, rochie de ocazie, sacou, compleu

Răspunde EXACT în formatul:
META_TITLE: [titlul aici]
META_DESC: [descrierea aici]"""
        
        try:
            response = call_gpt(prompt)
            
            new_title = ""
            new_desc = ""
            for line in response.split("\n"):
                if line.startswith("META_TITLE:"):
                    new_title = line.replace("META_TITLE:", "").strip()
                elif line.startswith("META_DESC:"):
                    new_desc = line.replace("META_DESC:", "").strip()
            
            print(f"✅ ({len(new_title)}c / {len(new_desc)}c)")
            
            results.append({
                "name": r["name"],
                "cod": r.get("cod", ""),
                "link": r["link"],
                "score": r["score"],
                "meta_title_actual": meta_title,
                "meta_title_sugerat": new_title,
                "meta_desc_actual": meta_desc,
                "meta_desc_sugerat": new_desc,
            })
        except Exception as e:
            print(f"❌ {e}")
        
        time.sleep(1)  # Rate limit
    
    return results


def export_xlsx(results, output=None):
    """Export suggestions to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        os.system("pip3 install openpyxl --break-system-packages -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    if output is None:
        output = "/home/ubuntu/seo_suggestions_ejolie.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sugestii SEO"
    
    headers = ["Produs", "Scor", "Meta Title Actual", "Chars", "Meta Title Sugerat", "Chars", "Meta Desc Actual", "Chars", "Meta Desc Sugerat", "Chars", "URL"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    
    for row_idx, r in enumerate(results, 2):
        data = [
            r["name"],
            r["score"],
            r["meta_title_actual"],
            len(r["meta_title_actual"]),
            r["meta_title_sugerat"],
            len(r["meta_title_sugerat"]),
            r["meta_desc_actual"][:100],
            len(r["meta_desc_actual"]),
            r["meta_desc_sugerat"],
            len(r["meta_desc_sugerat"]),
            r["link"],
        ]
        
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Color code character counts
            if col == 4:  # actual title chars
                if val < 30:
                    cell.fill = red_fill
                elif val < 45:
                    cell.fill = yellow_fill
            if col == 6:  # suggested title chars
                if 50 <= val <= 60:
                    cell.fill = green_fill
            if col == 8:  # actual desc chars
                if val < 70:
                    cell.fill = red_fill
                elif val < 120:
                    cell.fill = yellow_fill
            if col == 10:  # suggested desc chars
                if 120 <= val <= 155:
                    cell.fill = green_fill
    
    widths = [35, 7, 30, 6, 35, 6, 35, 6, 40, 6, 45]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    
    wb.save(output)
    print(f"\n✅ Excel salvat: {output}")
    return output


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=10, help="Max products to optimize")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()
    
    print(f"🤖 Generare sugestii SEO (max {args.limit} produse)...")
    results = generate_seo_suggestions(max_products=args.limit)
    
    if not results:
        print("✅ Toate produsele au SEO ok!")
        return
    
    print(f"\n📊 {len(results)} sugestii generate")
    export_xlsx(results, output=args.output)


if __name__ == "__main__":
    main()

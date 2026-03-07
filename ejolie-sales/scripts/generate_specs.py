#!/usr/bin/env python3
"""
generate_specs.py — Generează specificații lipsă pentru produse ejolie.ro cu Gemini Vision.

Citește products_missing_specs.json (output scan_specs.py).
Trimite imaginea + numele produsului la Gemini Vision.
Mapează valorile la ID-uri Extended (fuzzy matching).
Completează DOAR specs lipsă (nu suprascrie cele existente).
Salvează generated_specs.json + raport Excel.
Trimite raport text + Excel pe Telegram.

Usage:
    python3 generate_specs.py --id 12345       # Test pe 1 produs
    python3 generate_specs.py --limit 1        # Primul din lista incomplete
    python3 generate_specs.py                  # Toate produsele incomplete
    python3 generate_specs.py --dry-run        # Arată ce ar genera, fără salvare
    python3 generate_specs.py --no-telegram    # Fără raport Telegram
"""

import json
import os
import sys
import time
import argparse
import requests
import base64
from pathlib import Path
from dotenv import load_dotenv

# ── Paths ──────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
ENV_PATH = SCRIPT_DIR / '..' / '.env'
MISSING_PATH = SCRIPT_DIR / 'products_missing_specs.json'
FEED_PATH = SCRIPT_DIR / 'product_feed.json'
OUTPUT_PATH = SCRIPT_DIR / 'generated_specs.json'
EXCEL_PATH = SCRIPT_DIR / 'generated_specs_report.xlsx'

# ── Load .env ──────────────────────────────────────────────────────
load_dotenv(ENV_PATH)
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
EJOLIE_API_KEY = os.getenv('EJOLIE_API_KEY')
API_URL = os.getenv('EJOLIE_API_URL', 'https://ejolie.ro/api/')
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
TELEGRAM_CHAT_ID = os.getenv('TELEGRAM_CHAT_ID')

if not GEMINI_API_KEY:
    print("❌ GEMINI_API_KEY nu e setat în .env")
    sys.exit(1)

# ── Gemini API ─────────────────────────────────────────────────────
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}"

# ── Headers ────────────────────────────────────────────────────────
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
}

# ── Cele 6 specificații ────────────────────────────────────────────
SPEC_NAMES = ['Culoare', 'Material', 'Lungime', 'Croi', 'Stil', 'Model']

# ══════════════════════════════════════════════════════════════════
#  MAPPING-URI SPECIFICAȚII → ID-URI EXTENDED
# ══════════════════════════════════════════════════════════════════

CULOARE_MAP = {
    "Alb": 19, "Albastru": 22, "Albastru deschis": 34, "Albastru inchis": 7460,
    "Albastru petrol": 7481, "Animal print": 7465, "Aramiu": 14494, "Argintiu": 3040,
    "Auriu": 26, "Beige": 31, "Bej": 31, "Bleu": 3150, "Bleumarin": 30,
    "Bleumarin-Inchis": 3056, "Bordo": 27, "Bronz": 3061, "Cappuccino": 14500,
    "Caramiziu": 3054, "Corai": 3156, "Crem": 3051, "Fistic": 14499,
    "Fucsia": 3053, "Galben": 3038, "Grena": 37, "Gri": 33,
    "Gri metalizat": 14495, "Ivoar": 3154, "Kaki": 3052, "Lavanda": 3055,
    "Leopard": 3067, "Lila": 3153, "Magenta": 3157, "Maro": 29,
    "Marsala": 3037, "Mint": 14488, "Mov": 25, "Mov pruna": 14478,
    "Multicolor": 3042, "Mustar": 3096, "Negru": 20, "Nude": 38,
    "Olive": 14477, "Pastel": 7463, "Portocaliu": 3041, "Rosu": 21,
    "Roz": 24, "Roz prafuit": 14498, "Somon": 36, "Turcoaz": 28,
    "Verde": 23, "Verde inchis": 3044, "Verde lime": 3155,
    "Verde mint": 14497, "Vernil": 3151, "Visiniu": 7480, "floral": 3152
}

MATERIAL_MAP = {
    "Acryl": 3092, "Barbie": 46, "Brocart": 45, "Bumbac": 47, "Catifea": 39,
    "Casmir": 3097, "Crep": 3149, "Dantela": 7455, "Elasten": 48,
    "In": 7458, "Licra": 7470, "Lurex": 3057, "Matase": 14479,
    "Neopren": 7461, "Nylon": 3078, "Organza": 7456, "Paiete": 42,
    "Piele eco": 7489, "Poliester": 49, "Satin": 3079,
    "Stofa": 3048, "Tafta": 43, "Tricot": 3039, "Tul": 41,
    "Vascoza": 50, "Voal": 44, "Lana": 3073
}

LUNGIME_MAP = {
    "Lungi": 51, "Medii": 53, "Scurte": 52
}

CROI_MAP = {
    "Baby Doll": 54, "Cambrat": 7462, "Clos": 55, "Conic": 62, "Drept": 56,
    "Evazat": 14485, "In clini": 3046, "Lejer": 59, "Mulat": 57,
    "Peplum": 58, "Petrecuta": 3045, "Plisat": 14491, "Pliuri": 61,
    "Volane": 7471, "in A": 60
}

STIL_MAP = {
    "Asimetrica": 7472, "Birou": 63, "Casual": 64, "Casual-Elegant": 65,
    "Casual-Office": 66, "Cu crapatura": 73, "De ocazie": 67,
    "De seara": 75, "De zi": 68, "Dreapta": 74, "Elegant": 1147, "Eleganta": 1147,
    "Evazata": 70, "Lejera": 69, "Lunga": 71, "Lunga sirena": 72,
    "cu trena": 3047
}

MODEL_MAP = {
    "Accesorizata la baza gatului": 3050, "Aplicatii 3D": 14484, "Bretele reglabile": 7482,
    "Broderie": 77, "Brosa": 7479, "Buzunare": 7485, "Cambrat": 76,
    "Cordon": 7484, "Corset": 14490,
    "Cu decolteu in V": 81, "Cu pliuri": 84, "Cu volane": 83, "Dantela": 40,
    "Decolteu in V": 14481, "Fara maneci": 14501,
    "Fermoar": 14480, "Funda": 14483,
    "Maneca lunga": 7488, "Maneci lungi": 7488, "Maneca scurta": 14489,
    "Maneci 3/4": 7473, "Maneci bufante": 14468, "Maneci evazate": 14496,
    "Nasturi": 7487, "Paiete": 7467, "Plisata": 7466,
    "Umeri goi": 7468, "Cu umerii goi": 7468
}

SPEC_MAPS = {
    "Culoare": CULOARE_MAP,
    "Material": MATERIAL_MAP,
    "Lungime": LUNGIME_MAP,
    "Croi": CROI_MAP,
    "Stil": STIL_MAP,
    "Model": MODEL_MAP
}

# ── Fuzzy Mappings ─────────────────────────────────────────────────
FUZZY_MAP = {
    "Grena": "Bordo", "Roz prafuit": "Roz prafuit",
    "Roz prafu": "Roz prafuit", "Rose": "Roz", "Ivory": "Crem",
    "Ecru": "Crem", "Caramizius": "Caramiziu", "Portolicalie": "Portocaliu",
    "Smarald": "Verde inchis", "Verde smarald": "Verde inchis",
    "Turquoise": "Turcoaz", "Bej": "Beige",
    "Burgundy": "Bordo", "Champagne": "Beige", "Gold": "Auriu",
    "Silver": "Argintiu", "Navy": "Bleumarin", "Coral": "Corai",
    "Negru cu alb": "Negru", "Crem-roze": "Crem",
    "Indigo": "Albastru inchis",
    "Orange": "Portocaliu", "Petrol": "Albastru petrol",
    "Lycra": "Licra", "Tulle": "Tul", "Tull": "Tul",
    "Sifon": "Voal", "Mătase": "Matase", "Pudra": "Roz prafuit",
    "Sport": "Casual", "Elegant": "Eleganta",
    "Piele ecologica": "Piele eco", "Jerseu": "Tricot",
    "Maneci scurte": "Maneca scurta", "Cu nasturi": "Nasturi",
    "Cu fermoar": "Fermoar", "Cu funda": "Funda",
    "Cu cordon": "Cordon", "Cu buzunare": "Buzunare",
    "Sirena": "Lunga sirena", "Scurta": "De zi",
}

SPEC_VALUES = {
    "Culoare": list(CULOARE_MAP.keys()),
    "Material": list(MATERIAL_MAP.keys()),
    "Lungime": list(LUNGIME_MAP.keys()),
    "Croi": list(CROI_MAP.keys()),
    "Stil": list(STIL_MAP.keys()),
    "Model": list(MODEL_MAP.keys())
}


# ══════════════════════════════════════════════════════════════════
#  TELEGRAM
# ══════════════════════════════════════════════════════════════════

def send_telegram(text):
    """Trimite mesaj text pe Telegram. Suportă mesaje lungi."""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("  ⚠️  Telegram: token/chat_id lipsă, skip")
        return

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"

    chunks = []
    while len(text) > 4000:
        split_at = text.rfind('\n', 0, 4000)
        if split_at == -1:
            split_at = 4000
        chunks.append(text[:split_at])
        text = text[split_at:]
    chunks.append(text)

    for chunk in chunks:
        try:
            r = requests.post(url, json={
                'chat_id': TELEGRAM_CHAT_ID,
                'text': chunk,
                'parse_mode': 'HTML'
            }, timeout=15)
            if r.status_code != 200:
                print(f"  ⚠️  Telegram text eroare: {r.status_code}")
            time.sleep(0.3)
        except Exception as e:
            print(f"  ⚠️  Telegram eroare: {e}")


def send_telegram_file(filepath, caption=""):
    """Trimite fișier pe Telegram."""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("  ⚠️  Telegram: token/chat_id lipsă, skip")
        return

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    try:
        with open(filepath, 'rb') as f:
            r = requests.post(url, data={
                'chat_id': TELEGRAM_CHAT_ID,
                'caption': caption[:1024]  # Telegram caption limit
            }, files={
                'document': (os.path.basename(filepath), f)
            }, timeout=30)

        if r.status_code == 200:
            print(f" ✅ Fișier trimis!")
        else:
            print(f" ⚠️  Telegram file eroare: {r.status_code}")
    except Exception as e:
        print(f"  ⚠️  Telegram file eroare: {e}")


def build_telegram_report(results):
    """Construiește raportul Telegram text scurt (sumar)."""
    success_count = sum(1 for r in results if not r.get('error') and not r.get('unmapped'))
    unmapped_count = sum(1 for r in results if r.get('unmapped') and not r.get('error'))
    error_count = sum(1 for r in results if r.get('error'))

    lines = []
    lines.append("📋 <b>RAPORT GENERARE SPECIFICAȚII</b>")
    lines.append("")
    lines.append(f"📊 Total: {len(results)} produse")
    lines.append(f"✅ Succes complet: {success_count}")
    lines.append(f"⚠️ Cu probleme mapping: {unmapped_count}")
    lines.append(f"❌ Erori: {error_count}")
    lines.append("")
    lines.append("📎 Raportul Excel detaliat e atașat mai jos.")
    lines.append("")
    lines.append("👉 Dacă totul e ok, rulează:")
    lines.append("<code>python3 upload_specs.py</code>")

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════
#  EXCEL REPORT
# ══════════════════════════════════════════════════════════════════

def export_excel_report(results):
    """Generează raport Excel cu toate produsele și specs generate."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Specificații Generate"

    # ── Stiluri ────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    new_font = Font(bold=True, color="006100")
    kept_font = Font(color="808080", italic=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ── Header ─────────────────────────────────────────────────────
    headers = ['ID', 'Nume Produs', 'Status', 'Culoare', 'Material', 'Lungime', 'Croi', 'Stil', 'Model', 'Probleme']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # ── Date ───────────────────────────────────────────────────────
    for row_idx, result in enumerate(results, 2):
        pid = result.get('id', '')
        name = result.get('name', '')
        specs_to_add = result.get('specs_to_add', {})
        kept = result.get('kept_existing', [])
        unmapped = result.get('unmapped', [])
        error = result.get('error', '')

        # Status
        if error:
            status = "❌ EROARE"
            row_fill = error_fill
        elif unmapped:
            status = "⚠️ PARȚIAL"
            row_fill = warning_fill
        else:
            status = "✅ OK"
            row_fill = success_fill

        ws.cell(row=row_idx, column=1, value=int(pid) if str(pid).isdigit() else pid).border = thin_border
        ws.cell(row=row_idx, column=2, value=name).border = thin_border

        status_cell = ws.cell(row=row_idx, column=3, value=status)
        status_cell.fill = row_fill
        status_cell.border = thin_border

        # Specs per coloană
        for col_idx, spec_name in enumerate(SPEC_NAMES, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            if error:
                cell.value = error if col_idx == 4 else ""
                continue

            # Verificăm dacă e spec nou generat sau existent păstrat
            if spec_name in specs_to_add:
                values = specs_to_add[spec_name]
                vals_str = ", ".join([v['value'] for v in values])
                cell.value = f"🆕 {vals_str}"
                cell.font = new_font
            elif spec_name in kept:
                cell.value = "✅ (existent)"
                cell.font = kept_font
            else:
                # Verificăm dacă e unmapped
                unmapped_vals = [val for spec, val in unmapped if spec == spec_name]
                if unmapped_vals:
                    cell.value = f"⚠️ {', '.join(unmapped_vals)}"
                    cell.fill = warning_fill
                else:
                    cell.value = ""

        # Coloana Probleme
        problems = []
        if error:
            problems.append(error)
        for spec, val in unmapped:
            problems.append(f"{spec}: '{val}'")

        ws.cell(row=row_idx, column=10, value="; ".join(problems) if problems else "").border = thin_border

    # ── Lățimi coloane ─────────────────────────────────────────────
    ws.column_dimensions['A'].width = 8    # ID
    ws.column_dimensions['B'].width = 45   # Nume
    ws.column_dimensions['C'].width = 14   # Status
    ws.column_dimensions['D'].width = 18   # Culoare
    ws.column_dimensions['E'].width = 18   # Material
    ws.column_dimensions['F'].width = 12   # Lungime
    ws.column_dimensions['G'].width = 14   # Croi
    ws.column_dimensions['H'].width = 18   # Stil
    ws.column_dimensions['I'].width = 25   # Model
    ws.column_dimensions['J'].width = 35   # Probleme

    # ── Sumar (sheet 2) ────────────────────────────────────────────
    ws2 = wb.create_sheet("Sumar")
    ws2.cell(row=1, column=1, value="Statistică").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="Total produse:").font = Font(bold=True)
    ws2.cell(row=3, column=2, value=len(results))
    ws2.cell(row=4, column=1, value="Succes complet:").font = Font(bold=True)
    ws2.cell(row=4, column=2, value=sum(1 for r in results if not r.get('error') and not r.get('unmapped')))
    ws2.cell(row=5, column=1, value="Cu probleme:").font = Font(bold=True)
    ws2.cell(row=5, column=2, value=sum(1 for r in results if r.get('unmapped') and not r.get('error')))
    ws2.cell(row=6, column=1, value="Erori:").font = Font(bold=True)
    ws2.cell(row=6, column=2, value=sum(1 for r in results if r.get('error')))

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 10

    # ── Freeze panes ───────────────────────────────────────────────
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:J{len(results) + 1}"

    # ── Salvare ────────────────────────────────────────────────────
    wb.save(str(EXCEL_PATH))
    print(f"📊 Raport Excel salvat: {EXCEL_PATH.name}")

    return str(EXCEL_PATH)


# ══════════════════════════════════════════════════════════════════
#  API FUNCTIONS
# ══════════════════════════════════════════════════════════════════

def fetch_product_specs(product_id):
    """Fetch specificații pentru un produs via API."""
    if not EJOLIE_API_KEY:
        return None

    url = f"{API_URL}?id_produs={product_id}&apikey={EJOLIE_API_KEY}"

    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        data = r.json()

        if isinstance(data, dict) and str(product_id) in data:
            product_data = data[str(product_id)]
        elif isinstance(data, dict) and len(data) == 1:
            product_data = list(data.values())[0]
        else:
            product_data = data

        specs_raw = product_data.get('specificatii', []) if isinstance(product_data, dict) else []

        specs = {}
        for spec_name in SPEC_NAMES:
            specs[spec_name] = []

        EMPTY_VALUE = 'Fara optiune definita'
        for item in specs_raw:
            name = item.get('nume', '')
            values = item.get('valoare', [])
            if name in specs:
                clean_values = [v for v in values if v and v != EMPTY_VALUE]
                specs[name] = clean_values

        return specs

    except Exception as e:
        print(f"  ❌ API eroare {product_id}: {e}")
        return None


# ══════════════════════════════════════════════════════════════════
#  CORE FUNCTIONS
# ══════════════════════════════════════════════════════════════════

def download_image(url):
    """Descarcă imaginea și returnează bytes + mime type."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()

        content_type = r.headers.get('content-type', 'image/jpeg')
        if 'png' in content_type:
            mime = 'image/png'
        elif 'webp' in content_type:
            mime = 'image/webp'
        else:
            mime = 'image/jpeg'

        return r.content, mime
    except Exception as e:
        print(f"  ❌ Eroare download imagine: {e}")
        return None, None


def call_gemini_vision(image_bytes, mime_type, product_name, missing_specs):
    """Trimite imaginea + prompt la Gemini Vision."""
    image_b64 = base64.b64encode(image_bytes).decode('utf-8')

    specs_needed = []
    for spec in missing_specs:
        values = SPEC_VALUES.get(spec, [])
        values_str = ", ".join(values)
        specs_needed.append(f'- "{spec}": alege din [{values_str}]')

    specs_prompt = "\n".join(specs_needed)

    prompt = f"""Analizează imaginea acestui produs vestimentar și numele: "{product_name}".

Generează DOAR aceste specificații lipsă:
{specs_prompt}

REGULI:
1. Alege DOAR din valorile date între paranteze pătrate pentru fiecare specificație.
2. Pentru "Model" poți alege MULTIPLE valori separate prin virgulă.
3. Pentru celelalte specificații alege O SINGURĂ valoare.
4. IMPORTANT: Dacă numele produsului conține o culoare (ex: "neagra", "rosie", "fucsia", "verde", "alba", "bordo", "albastra", "aurie", "maro", "bej", "crem", "lila", "mov", "portocalie", "kaki", "gri", "turcoaz", "corai"), folosește OBLIGATORIU acea culoare. Numele produsului are prioritate maximă pentru Culoare.
5. Dacă nu ești sigur, alege cea mai probabilă valoare bazat pe imagine și nume.
6. Răspunde DOAR cu JSON valid, fără explicații, fără markdown, fără ```json.

Format răspuns (exemplu):
{{"Culoare": "Negru", "Material": "Poliester", "Lungime": "Lungi", "Croi": "Mulat", "Stil": "De seara", "Model": "Fara maneci, Decolteu in V"}}
"""

    payload = {
        "contents": [{
            "parts": [
                {
                    "inlineData": {
                        "mimeType": mime_type,
                        "data": image_b64
                    }
                },
                {"text": prompt}
            ]
        }],
        "generationConfig": {
            "temperature": 0.1,
            "maxOutputTokens": 500,
            "thinkingConfig": {
                "thinkingBudget": 0
            }
        }
    }

    try:
        r = requests.post(GEMINI_URL, json=payload, timeout=60)
        r.raise_for_status()
        data = r.json()

        candidates = data.get('candidates', [])
        if not candidates:
            print(f"  ❌ Gemini: niciun candidat")
            return None

        parts = candidates[0].get('content', {}).get('parts', [])
        text = ""
        for part in parts:
            if 'text' in part:
                text += part['text']

        if not text.strip():
            print(f"  ❌ Gemini: răspuns gol")
            return None

        text = text.strip()
        if text.startswith("```json"):
            text = text[7:]
        if text.startswith("```"):
            text = text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()

        specs = json.loads(text)
        return specs

    except json.JSONDecodeError as e:
        print(f"  ❌ Gemini JSON invalid: {e}")
        print(f"     Răspuns: {text[:200]}")
        return None
    except requests.exceptions.RequestException as e:
        print(f"  ❌ Gemini API eroare: {e}")
        return None
    except Exception as e:
        print(f"  ❌ Gemini eroare: {e}")
        return None


def map_value_to_id(spec_name, value):
    """Mapează valoare Gemini → ID Extended. 4 niveluri matching."""
    spec_map = SPEC_MAPS.get(spec_name, {})

    if value in spec_map:
        return value, spec_map[value]

    for map_key, map_id in spec_map.items():
        if map_key.lower() == value.lower():
            return map_key, map_id

    fuzzy_value = FUZZY_MAP.get(value)
    if fuzzy_value:
        if fuzzy_value in spec_map:
            return fuzzy_value, spec_map[fuzzy_value]
        for map_key, map_id in spec_map.items():
            if map_key.lower() == fuzzy_value.lower():
                return map_key, map_id

    value_lower = value.lower()
    for map_key, map_id in spec_map.items():
        if value_lower in map_key.lower() or map_key.lower() in value_lower:
            return map_key, map_id

    return None, None


def process_gemini_response(gemini_specs, missing_specs):
    """Procesează răspunsul Gemini și mapează la ID-uri."""
    specs_to_add = {}
    unmapped = []

    for spec_name in missing_specs:
        raw_value = gemini_specs.get(spec_name, "")
        if not raw_value:
            unmapped.append((spec_name, "GOLI — Gemini nu a returnat valoare"))
            continue

        if spec_name == "Model" and "," in str(raw_value):
            values = [v.strip() for v in str(raw_value).split(",")]
            mapped_values = []
            for val in values:
                mapped_name, mapped_id = map_value_to_id(spec_name, val)
                if mapped_id:
                    mapped_values.append({"value": mapped_name, "value_id": mapped_id})
                else:
                    unmapped.append((spec_name, val))
            if mapped_values:
                specs_to_add[spec_name] = mapped_values
        else:
            mapped_name, mapped_id = map_value_to_id(spec_name, str(raw_value))
            if mapped_id:
                specs_to_add[spec_name] = [{"value": mapped_name, "value_id": mapped_id}]
            else:
                unmapped.append((spec_name, raw_value))

    return specs_to_add, unmapped


def load_missing_products():
    """Citește products_missing_specs.json."""
    if not MISSING_PATH.exists():
        print(f"❌ {MISSING_PATH} nu există. Rulează scan_specs.py mai întâi.")
        sys.exit(1)

    with open(MISSING_PATH, 'r', encoding='utf-8') as f:
        products = json.load(f)

    print(f"📦 Încărcat {len(products)} produse cu specs lipsă")
    return products


def load_product_from_feed(product_id):
    """Citește un produs specific din product_feed.json."""
    if not FEED_PATH.exists():
        return None

    with open(FEED_PATH, 'r', encoding='utf-8') as f:
        products = json.load(f)

    for p in products:
        if str(p.get('id')) == str(product_id):
            return p

    return None


def print_product_result(result):
    """Afișează rezultatul generării pentru un produs."""
    print(f"\n{'='*50}")
    print(f"📦 [{result['id']}] {result['name']}")
    print(f"{'='*50}")

    if result.get('kept_existing'):
        print(f"  ✅ Specs existente păstrate: {', '.join(result['kept_existing'])}")

    if result.get('specs_to_add'):
        print(f"  🆕 Specs generate:")
        for spec_name, values in result['specs_to_add'].items():
            vals_str = ", ".join([f"{v['value']} (ID:{v['value_id']})" for v in values])
            print(f"    → {spec_name}: {vals_str}")

    if result.get('unmapped'):
        print(f"  ⚠️  Valori fără mapping:")
        for spec, val in result['unmapped']:
            print(f"    ✗ {spec}: '{val}'")

    if result.get('error'):
        print(f"  ❌ Eroare: {result['error']}")


def main():
    parser = argparse.ArgumentParser(description='Generează specificații lipsă cu Gemini Vision')
    parser.add_argument('--id', type=int, help='ID produs specific')
    parser.add_argument('--limit', type=int, default=0, help='Limită produse (0 = toate)')
    parser.add_argument('--dry-run', action='store_true', help='Arată rezultate fără salvare')
    parser.add_argument('--no-telegram', action='store_true', help='Nu trimite raport pe Telegram')
    args = parser.parse_args()

    # ── Încarcă produse ────────────────────────────────────────────
    if args.id:
        product = load_product_from_feed(args.id)
        if not product:
            print(f"❌ Produs {args.id} nu a fost găsit în product_feed.json")
            sys.exit(1)

        print(f"🎯 Generare specs pentru produs specific: {args.id}")
        current_specs = fetch_product_specs(args.id)
        if current_specs is None:
            print(f"❌ Nu pot fetcha specs pentru {args.id}")
            sys.exit(1)

        missing = [s for s in SPEC_NAMES if not current_specs.get(s)]
        if not missing:
            print(f"✅ Produsul {args.id} are toate 6 specificațiile complete!")
            sys.exit(0)

        products = [{
            'id': product.get('id'),
            'name': product.get('title', product.get('name', '')),
            'image': product.get('image', ''),
            'current_specs': current_specs,
            'missing': missing
        }]
    else:
        products = load_missing_products()

    # ── Filtru --limit ─────────────────────────────────────────────
    if args.limit > 0:
        products = products[:args.limit]
        print(f"📏 Limitat la {args.limit} produse")

    # ── Generare ───────────────────────────────────────────────────
    results = []
    success = 0
    errors = 0
    total = len(products)

    print(f"\n🤖 Generare specs cu Gemini Vision pentru {total} produse...\n")

    for i, product in enumerate(products, 1):
        pid = product.get('id')
        name = product.get('name', '')
        image_url = product.get('image', '')
        missing = product.get('missing', [])

        kept_existing = [s for s in SPEC_NAMES if s not in missing]

        print(f"[{i}/{total}] {pid} - {name[:50]}...")
        print(f"  📷 Download imagine...", end='', flush=True)

        if not image_url:
            print(f" ❌ Fără imagine!")
            result = {
                'id': pid, 'name': name, 'error': 'Fara imagine',
                'specs_to_add': {}, 'kept_existing': kept_existing,
                'unmapped': [], 'missing': missing
            }
            results.append(result)
            errors += 1
            continue

        image_bytes, mime_type = download_image(image_url)
        if not image_bytes:
            print(f" ❌ Download eșuat!")
            result = {
                'id': pid, 'name': name, 'error': 'Download imagine esuat',
                'specs_to_add': {}, 'kept_existing': kept_existing,
                'unmapped': [], 'missing': missing
            }
            results.append(result)
            errors += 1
            continue

        print(f" OK ({len(image_bytes)//1024}KB)")
        print(f"  🤖 Gemini Vision ({len(missing)} specs lipsă)...", end='', flush=True)

        gemini_specs = call_gemini_vision(image_bytes, mime_type, name, missing)
        if not gemini_specs:
            print(f" ❌ Gemini fail!")
            result = {
                'id': pid, 'name': name, 'error': 'Gemini Vision fail',
                'specs_to_add': {}, 'kept_existing': kept_existing,
                'unmapped': [], 'missing': missing
            }
            results.append(result)
            errors += 1
            continue

        print(f" OK")

        specs_to_add, unmapped = process_gemini_response(gemini_specs, missing)

        result = {
            'id': pid,
            'name': name,
            'specs_to_add': specs_to_add,
            'kept_existing': kept_existing,
            'unmapped': unmapped,
            'missing': missing,
            'gemini_raw': gemini_specs
        }
        results.append(result)

        mapped_count = len(specs_to_add)
        total_missing = len(missing)
        if unmapped:
            print(f"  ⚠️  Mapat {mapped_count}/{total_missing}, {len(unmapped)} fără mapping")
        else:
            print(f"  ✅ Mapat {mapped_count}/{total_missing} specs")
            success += 1

        print_product_result(result)

        # Pauză între requesturi (1s)
        if i < total:
            time.sleep(1)

    # ── Statistici ─────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"📊 STATISTICI GENERARE")
    print(f"{'='*60}")
    print(f"  Total produse:           {total}")
    print(f"  ✅ Succes complet:        {success}")
    print(f"  ⚠️  Parțial (cu unmapped): {total - success - errors}")
    print(f"  ❌ Erori:                 {errors}")

    all_unmapped = []
    for r in results:
        for spec, val in r.get('unmapped', []):
            all_unmapped.append(f"{spec}: '{val}'")

    if all_unmapped:
        print(f"\n  ⚠️  Valori fără mapping (de adăugat manual):")
        for item in sorted(set(all_unmapped)):
            print(f"    ✗ {item}")

    print(f"{'='*60}")

    # ── Salvare JSON ───────────────────────────────────────────────
    if not args.dry_run:
        to_save = [r for r in results if r.get('specs_to_add')]

        with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
            json.dump(to_save, f, ensure_ascii=False, indent=2)

        print(f"\n💾 Salvat {len(to_save)} produse cu specs generate în {OUTPUT_PATH.name}")

    # ── Export Excel ───────────────────────────────────────────────
    if not args.dry_run and results:
        print(f"\n📊 Generare raport Excel...", end='', flush=True)
        excel_path = export_excel_report(results)
        print(f" ✅")

    # ── Raport Telegram (text + Excel) ─────────────────────────────
    if not args.no_telegram and not args.dry_run and results:
        print(f"\n📱 Trimit raport pe Telegram...", end='', flush=True)
        report = build_telegram_report(results)
        send_telegram(report)
        print(f" ✅ Text trimis!")

        print(f"📎 Trimit Excel pe Telegram...", end='', flush=True)
        send_telegram_file(excel_path, caption=f"📋 Raport specificații - {len(results)} produse")

    print(f"\n✅ Generare completă!")


if __name__ == '__main__':
    main()
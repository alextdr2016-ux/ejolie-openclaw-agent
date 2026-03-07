#!/usr/bin/env python3
"""
export_trendyol.py v3 - Export ejolie.ro products to Trendyol import template
Column order matches official Trendyol "Evening & Prom Dress" template (51 columns).

Utilizare:
    python3 export_trendyol.py --brand ejolie
    python3 export_trendyol.py --start 12356 --end 12415
    python3 export_trendyol.py --ids 12356,12360,12370
    python3 export_trendyol.py --start 12356 --end 12415 --telegram
"""

import os
import json
import re
import argparse
import time
import sys
from datetime import datetime

import requests
from dotenv import load_dotenv

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, '..', '.env')
if not os.path.exists(ENV_PATH):
    ENV_PATH = os.path.join(SCRIPT_DIR, '.env')
load_dotenv(ENV_PATH)

API_KEY = os.getenv('EJOLIE_API_KEY', '')
API_BASE = os.getenv('EJOLIE_API_URL', 'https://ejolie.ro/api/')
GEMINI_KEY = os.getenv('GEMINI_API_KEY', '')
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', '')
TELEGRAM_CHAT_ID = os.getenv('TELEGRAM_CHAT_ID', '44151343')
HEADERS_API = {'User-Agent': 'Mozilla/5.0'}
MAP_FILE = os.path.join(SCRIPT_DIR, 'barcode_ejolie_map.json')

# ═══════════════════════════════════════════
#  OFFICIAL TRENDYOL TEMPLATE - 51 columns
#  Category: Evening & Prom Dress (543)
# ═══════════════════════════════════════════

TRENDYOL_HEADERS = [
    "Cod de bare",                    # 0
    "Codul modelului",                # 1
    "Marca produsului",               # 2
    "ID categorie",                   # 3
    "Titlu",                          # 4
    "Descriere",                      # 5
    "Preț inițial",                   # 6
    "Preț de vânzare Trendyol",       # 7
    "Stoc",                           # 8
    "Cod stoc",                       # 9
    "Cotă TVA",                       # 10
    "Imagine 1",                      # 11
    "Imagine 2",                      # 12
    "Imagine 3",                      # 13
    "Imagine 4",                      # 14
    "Imagine 5",                      # 15
    "Imagine 6",                      # 16
    "Imagine 7",                      # 17
    "Imagine 8",                      # 18
    "Termendeprelucrare",             # 19
    "Mărime",                         # 20
    "Culoare",                        # 21
    "Persona",                        # 22
    "Tipul de material",              # 23
    "Material",                       # 24
    "Grupa de vârstă",                # 25
    "Potrivire",                      # 26
    "Model",                          # 27
    "Detalii privind durabilitatea",  # 28
    "Lungimea mânecii",              # 29
    "Tipul de material",              # 30
    "Buzunar",                        # 31
    "Lungime",                        # 32
    "Siluetă",                        # 33
    "Tipul de închidere",             # 34
    "Tip de produs",                  # 35
    "Ocazie",                         # 36
    "Numele producătorului",          # 37
    "Culoare",                        # 38
    "Detalii despre produs",          # 39
    "Origine",                        # 40
    "Caracteristici suplimentare",    # 41
    "Compoziția materialului",        # 42
    "Starea curelei",                 # 43
    "Tipul de mânecă",               # 44
    "Colecția",                       # 45
    "Instrucțiuni de îngrijire",      # 46
    "Guler",                          # 47
    "Căptușeală",                     # 48
    "Sex",                            # 49
    "Sezonul",                        # 50
]

# ═══════════════════════════════════════════
#  COLOR MAPPING
# ═══════════════════════════════════════════

TRENDYOL_COLORS = {
    "negru": "Negru", "neagra": "Negru", "negra": "Negru", "black": "Negru",
    "alb": "Alb", "alba": "Alb", "ivory": "Alb",
    "rosu": "Roșu", "rosie": "Roșu",
    "verde": "Verde", "lime": "Verde", "mint": "Verde", "vernil": "Verde",
    "albastru": "Albastru", "albastra": "Albastru", "bleu": "Albastru",
    "galben": "Galben", "galbena": "Galben",
    "roz": "Roz", "fucsia": "Roz", "somon": "Roz",
    "bordo": "Burgundia", "burgundy": "Burgundia", "visiniu": "Burgundia",
    "mov": "Multicolor", "lila": "Multicolor", "lavanda": "Multicolor",
    "gri": "Gri",
    "bej": "Bej", "nude": "Bej", "crem": "Crem",
    "turcoaz": "Turcoaz", "portocaliu": "Portocaliu", "portocalie": "Portocaliu",
    "coral": "Portocaliu", "caramiziu": "Portocaliu",
    "argintiu": "Argintiu", "auriu": "Auriu", "aurie": "Auriu",
    "maro": "Maro", "kaki": "Kaki", "ciocolatiu": "Maro",
    "multicolor": "Multicolor", "imprimeu": "Multicolor", "floral": "Multicolor",
    "bleumarin": "Bleumarin",
}

ALLOWED_SILUETA = ["A-line", "Asimetric", "Bodycon", "Drept",
                   "Sirenă", "Prințesă", "Shift", "Slip", "Oversize", "Peplum", "Wrap"]
ALLOWED_MANECA = ["Bretele spaghetti", "Fără mâneci",
                  "Lung", "Mâneca trei sferturi", "Scurt"]
ALLOWED_GULER = ["Fără bretele", "Gât rotund", "Gât în V",
                 "Guler bărcuță", "Guler cu rever", "Cache-coeur", "Gât înalt", "Un umăr"]
ALLOWED_OCAZIE = ["Absolvire / Bal de absolvire", "Bal", "Casual",
                  "Cocktail", "Elegant", "Elegant / Noapte", "Petrecere", "Seara / zilnic"]
ALLOWED_MATERIAL = ["Amestec poliester", "Catifea", "Crepe", "Cu paiete",
                    "Dantelă", "Fabric lucios", "Satin", "Tul", "Voal", "Altele"]
ALLOWED_SEZON = ["Iarna", "Primăvară / Toamnă",
                 "Toamnă / Iarnă", "Toate anotimpurile", "Vara"]


# ═══════════════════════════════════════════
#  EAN-13 BARCODE
# ═══════════════════════════════════════════

_ean_counter = 0
_existing_barcodes = set()


def init_ean_counter():
    """Încarcă barcodes existente din mapping și setează counter-ul să continue de unde a rămas."""
    global _ean_counter, _existing_barcodes
    if os.path.exists(MAP_FILE):
        with open(MAP_FILE, 'r') as f:
            mapping = json.load(f)
        _existing_barcodes = set(mapping.keys())
        for bc in _existing_barcodes:
            if bc.startswith('200') and len(bc) == 13:
                try:
                    counter = int(bc[3:12])
                    if counter > _ean_counter:
                        _ean_counter = counter
                except ValueError:
                    pass
        print(
            f"  ✓ Barcodes existente: {len(_existing_barcodes)}, counter continuă de la {_ean_counter + 1}")


def generate_ean13():
    global _ean_counter
    _ean_counter += 1
    base = "200" + str(_ean_counter).zfill(9)
    base = base[:12]
    total = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(base))
    check = (10 - (total % 10)) % 10
    barcode = base + str(check)
    # Safety check
    if barcode in _existing_barcodes:
        return generate_ean13()  # recurse until unique
    return barcode


def update_barcode_mapping(new_barcodes):
    """Adaugă barcodes noi în mapping-ul existent."""
    if not new_barcodes:
        return
    mapping = {}
    if os.path.exists(MAP_FILE):
        with open(MAP_FILE, 'r') as f:
            mapping = json.load(f)
    mapping.update(new_barcodes)
    with open(MAP_FILE, 'w') as f:
        json.dump(mapping, f)
    print(
        f"  ✓ Mapping actualizat: +{len(new_barcodes)} barcodes (total: {len(mapping)})")


def extract_color(name):
    words = name.lower().split()
    for word in words:
        if word in TRENDYOL_COLORS:
            return TRENDYOL_COLORS[word]
    name_lower = name.lower()
    for key, val in TRENDYOL_COLORS.items():
        if key in name_lower:
            return val
    return "Multicolor"


# ═══════════════════════════════════════════
#  GEMINI ATTRIBUTE EXTRACTION
# ═══════════════════════════════════════════

def call_gemini(prompt):
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={GEMINI_KEY}"
    data = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": 500, "thinkingConfig": {"thinkingBudget": 0}}
    }
    try:
        resp = requests.post(url, json=data, timeout=30)
        resp.raise_for_status()
        result = resp.json()
        return result['candidates'][0]['content']['parts'][0]['text'].strip()
    except Exception as e:
        print(f"⚠ Gemini error: {e}")
        return ""


def extract_attributes(name, description):
    clean_desc = re.sub(r'<[^>]+>', '', str(description))[:500]
    prompt = f"""Analizează acest produs vestimentar și extrage atributele.

Nume: {name}
Descriere: {clean_desc}

Alege EXACT din valorile permise:
SILUETA: {', '.join(ALLOWED_SILUETA)}
LUNGIME_MANECA: {', '.join(ALLOWED_MANECA)}
GULER: {', '.join(ALLOWED_GULER)}
OCAZIE: {', '.join(ALLOWED_OCAZIE)}
MATERIAL: {', '.join(ALLOWED_MATERIAL)}
TIP_MATERIAL: Knit, Laced, Țesut, Nespecificat
MODEL: Simplu, Cu model floral, Dantelă, Satin, Plain
CAPTUSIT: Căptușit, Fără căptușeală
INCHIDERE: Cu fermoar, Fără închidere, Buton
SEZON: {', '.join(ALLOWED_SEZON)}

Răspunde EXACT:
SILUETA: valoare
LUNGIME_MANECA: valoare
GULER: valoare
OCAZIE: valoare
MATERIAL: valoare
TIP_MATERIAL: valoare
MODEL: valoare
CAPTUSIT: valoare
INCHIDERE: valoare
SEZON: valoare"""

    response = call_gemini(prompt)
    attrs = {}
    for line in response.split("\n"):
        if ":" in line:
            key, val = line.split(":", 1)
            attrs[key.strip().upper()] = val.strip()
    return attrs


# ═══════════════════════════════════════════
#  API FETCH
# ═══════════════════════════════════════════

def fetch_products_by_brand(brand="ejolie"):
    all_products = {}
    page = 1
    print(f"  Fetch produse brand '{brand}'...")
    while True:
        url = f"{API_BASE}?produse&brand={brand}&apikey={API_KEY}&pagina={page}&limit=50"
        try:
            resp = requests.get(url, headers=HEADERS_API, timeout=60)
            data = resp.json()
        except Exception as e:
            print(f"  ✗ Eroare pagina {page}: {e}")
            break
        if not data or not isinstance(data, dict):
            break
        all_products.update(data)
        count = len(data)
        print(f"    Pagina {page}: {count} (total: {len(all_products)})")
        if count < 50:
            break
        page += 1
        time.sleep(0.3)
    return all_products


def fetch_products_by_ids(product_ids):
    all_products = {}
    print(f"  Fetch {len(product_ids)} produse pe ID...")
    for i, pid in enumerate(product_ids):
        url = f"{API_BASE}?id_produs={pid}&apikey={API_KEY}"
        try:
            resp = requests.get(url, headers=HEADERS_API, timeout=30)
            data = resp.json()
            if isinstance(data, dict) and str(pid) in data:
                product = data[str(pid)]
                if product.get('optiuni'):
                    all_products[str(pid)] = product
                else:
                    print(f"    ⚠ ID {pid}: fără opțiuni")
            else:
                print(f"    ✗ ID {pid}: nu există")
        except Exception as e:
            print(f"    ✗ ID {pid}: {e}")
        if (i + 1) % 10 == 0:
            print(f"    ... {i+1}/{len(product_ids)}")
        time.sleep(0.3)
    return all_products


# ═══════════════════════════════════════════
#  EXCEL EXPORT - NEW COLUMN ORDER
# ═══════════════════════════════════════════

def export_trendyol(products, output_path):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Introdu informațiile despre pro"

    red_fill = PatternFill(start_color="FFD9D9",
                           end_color="FFD9D9", fill_type="solid")
    for col, h in enumerate(TRENDYOL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if col <= 11:
            cell.fill = red_fill

    row_idx = 2
    prod_count = 0
    attr_cache = {}
    new_barcodes = {}  # barcode → ejolie_id (for updating mapping)

    for pid, prod in products.items():
        optiuni = prod.get("optiuni", {})
        if not isinstance(optiuni, dict):
            continue

        name = prod.get("nume", "")
        cod = prod.get("cod_produs", "")
        pret = prod.get("pret", "0")
        pret_discount = prod.get("pret_discount", "0")
        descriere = prod.get("descriere", "")
        clean_desc = re.sub(r'<[^>]+>', '', str(descriere))[:2000]
        color = extract_color(name)

        if len(name) > 100:
            print(
                f"    ⚠ Titlu trunchiat: {name[:50]}... ({len(name)} → 100 char)")

        imagini = prod.get("imagini", [])
        if not isinstance(imagini, list):
            imagini = []
        imagine_main = prod.get("imagine", "")
        if imagine_main and imagine_main not in imagini:
            imagini.insert(0, imagine_main)

        # Gemini attributes
        if pid not in attr_cache:
            prod_count += 1
            print(f"  🤖 [{prod_count}] {name[:45]}...", end=" ", flush=True)
            attrs = extract_attributes(name, descriere)
            attr_cache[pid] = attrs
            print("✅")
            time.sleep(0.5)
        attrs = attr_cache[pid]

        try:
            p = float(str(pret).replace(",", ".")) if pret else 0
            pd = float(str(pret_discount).replace(
                ",", ".")) if pret_discount else 0
            sell_price = pd if pd > 0 and pd < p else p
        except:
            p = 0
            sell_price = 0

        brand_data = prod.get("brand", {})
        brand_name = brand_data.get("nume", "Ejolie") if isinstance(
            brand_data, dict) else str(brand_data)

        for oid, opt in optiuni.items():
            if not isinstance(opt, dict):
                continue
            stoc_fizic = int(opt.get("stoc_fizic", 0))
            if stoc_fizic <= 0:
                continue

            size = opt.get("nume_optiune", "")
            try:
                op = float(str(opt.get("pret", pret)).replace(",", "."))
                opd = float(
                    str(opt.get("pret_discount", "0")).replace(",", "."))
                opt_sell = opd if opd > 0 and opd < op else op
            except:
                op = p
                opt_sell = sell_price

            barcode = generate_ean13()
            new_barcodes[barcode] = pid  # track for mapping update

            # 51 columns in EXACT Trendyol template order
            row = [
                barcode,                                          # 0  Cod de bare
                cod or pid,                                       # 1  Codul modelului
                brand_name,                                       # 2  Marca produsului
                543,                                              # 3  ID categorie
                name[:100],                                       # 4  Titlu
                # 5  Descriere
                clean_desc[:5000],
                op,                                               # 6  Preț inițial
                opt_sell,                                         # 7  Preț de vânzare Trendyol
                stoc_fizic,                                       # 8  Stoc
                f"{cod}-{size}",                                  # 9  Cod stoc
                21,                                               # 10 Cotă TVA
            ]
            # Imagini 1-8 (columns 11-18)
            for i in range(8):
                row.append(imagini[i] if i < len(imagini) else "")

            row.extend([
                # 19 Termendeprelucrare (1 zi)
                1,
                size,                                             # 20 Mărime
                color,                                            # 21 Culoare
                "Feminin",                                        # 22 Persona
                # 23 Tipul de material
                attrs.get("TIP_MATERIAL", "Țesut"),
                attrs.get("MATERIAL", "Amestec poliester"),       # 24 Material
                "Adult",                                          # 25 Grupa de vârstă
                # 26 Potrivire
                attrs.get("SILUETA", "A-line"),
                attrs.get("MODEL", "Simplu"),                     # 27 Model
                "Nu",                                             # 28 Detalii privind durabilitatea
                # 29 Lungimea mânecii
                attrs.get("LUNGIME_MANECA", "Fără mâneci"),
                # 30 Tipul de material (2)
                attrs.get("TIP_MATERIAL", "Țesut"),
                "Fără buzunar",                                   # 31 Buzunar
                "",                                               # 32 Lungime
                attrs.get("SILUETA", "A-line"),                   # 33 Siluetă
                # 34 Tipul de închidere
                attrs.get("INCHIDERE", "Cu fermoar"),
                "Simplu",                                         # 35 Tip de produs
                attrs.get("OCAZIE", "Elegant / Noapte"),          # 36 Ocazie
                brand_name,                                       # 37 Numele producătorului
                # 38 Culoare (2)
                color,
                "",                                               # 39 Detalii despre produs
                "RO",                                             # 40 Origine
                "",                                               # 41 Caracteristici suplimentare
                "",                                               # 42 Compoziția materialului
                "Fără centură",                                   # 43 Starea curelei
                # 44 Tipul de mânecă
                attrs.get("LUNGIME_MANECA", "Fără mâneci"),
                "Elegant / Noapte",                               # 45 Colecția
                "",                                               # 46 Instrucțiuni de îngrijire
                attrs.get("GULER", "Gât rotund"),                 # 47 Guler
                # 48 Căptușeală
                attrs.get("CAPTUSIT", "Fără căptușeală"),
                "Femeie",                                         # 49 Sex
                attrs.get("SEZON", "Toate anotimpurile"),         # 50 Sezonul
            ])

            for col, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col, value=val)
            row_idx += 1

    for col in range(1, len(TRENDYOL_HEADERS) + 1):
        import openpyxl.utils
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    total = row_idx - 2
    wb.save(output_path)

    # Actualizează barcode mapping cu noile barcodes
    if new_barcodes:
        update_barcode_mapping(new_barcodes)

    return total, prod_count


# ═══════════════════════════════════════════
#  TELEGRAM
# ═══════════════════════════════════════════

def send_telegram_file(filepath, caption=''):
    if not TELEGRAM_BOT_TOKEN:
        return False
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    try:
        with open(filepath, 'rb') as f:
            resp = requests.post(url,
                                 data={'chat_id': TELEGRAM_CHAT_ID,
                                       'caption': caption, 'parse_mode': 'HTML'},
                                 files={'document': f}, timeout=60)
        return resp.status_code == 200
    except:
        return False


# ═══════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Export ejolie.ro to Trendyol template')
    parser.add_argument("--brand", default=None,
                        help="Export pe brand (ejolie, trendya, artista)")
    parser.add_argument("--start", type=int, help="ID produs start")
    parser.add_argument("--end", type=int, help="ID produs end")
    parser.add_argument("--ids", type=str,
                        help="ID-uri specifice separate prin virgulă")
    parser.add_argument("--output", default=None, help="Fișier output")
    parser.add_argument("--limit", type=int, default=None,
                        help="Limită produse")
    parser.add_argument("--telegram", action='store_true',
                        help="Trimite pe Telegram")
    args = parser.parse_args()

    if not API_KEY:
        print("✗ EJOLIE_API_KEY nu e setat!")
        sys.exit(1)
    if not GEMINI_KEY:
        print("✗ GEMINI_API_KEY nu e setat!")
        sys.exit(1)

    if args.ids:
        product_ids = [int(x.strip()) for x in args.ids.split(',')]
        mode = f"IDs: {len(product_ids)} produse"
    elif args.start and args.end:
        product_ids = list(range(args.start, args.end + 1))
        mode = f"Range: {args.start}-{args.end}"
    elif args.brand:
        product_ids = None
        mode = f"Brand: {args.brand}"
    else:
        print("✗ Specifică --brand, --start/--end, sau --ids")
        sys.exit(1)

    output_path = args.output or os.path.join(
        SCRIPT_DIR, f"trendyol_export_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    )

    print(f"═══ Export Trendyol v3 ═══")
    print(f"  Mod:      {mode}")
    print(f"  Output:   {output_path}")
    print()

    if product_ids:
        products = fetch_products_by_ids(product_ids)
    else:
        products = fetch_products_by_brand(brand=args.brand)

    if args.limit:
        products = dict(list(products.items())[:args.limit])

    if not products:
        print("✗ Niciun produs!")
        sys.exit(1)

    print(f"\n  📦 {len(products)} produse\n")

    # Inițializează counter EAN-13 din barcodes existente
    init_ean_counter()

    total_rows, prod_count = export_trendyol(products, output_path)

    print(f"""
═══ EXPORT COMPLET ═══
  Fișier:    {output_path}
  Rânduri:   {total_rows}
  Produse:   {prod_count} (Gemini)
  Coloane:   51 (template oficial + Sezonul)

  ⬆ UPLOAD: Acțiuni colective → Crearea de produse noi""")

    if args.telegram:
        caption = f"📦 <b>Trendyol Export — {datetime.now().strftime('%d.%m.%Y')}</b>\n✅ {total_rows} rânduri, {prod_count} produse"
        if send_telegram_file(output_path, caption=caption):
            print("  ✓ Trimis pe Telegram!")


if __name__ == "__main__":
    main()

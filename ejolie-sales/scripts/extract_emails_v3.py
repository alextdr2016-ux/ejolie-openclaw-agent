#!/usr/bin/env python3
"""
extract_emails_v3.py — Extrage emailuri + data primei comenzi din ejolie.ro
Chunks de 7 zile, salvează progres, Excel cu email + data.
"""

import json
import os
import sys
import time
import urllib.request
import urllib.parse
from datetime import datetime, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, '..', '.env')
if os.path.exists(ENV_PATH):
    with open(ENV_PATH) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())

API_KEY = os.environ.get('EJOLIE_API_KEY', '')
BASE_URL = 'https://ejolie.ro/api/'
CUTOFF = datetime(2025, 8, 1)
MAX_EMAILS = 999999  # toate
CHUNK_DAYS = 7
PROGRESS_FILE = os.path.join(SCRIPT_DIR, 'email_extract_progress_v3.json')


def fetch_chunk(start_date, end_date, retries=2):
    ds = start_date.strftime('%d-%m-%Y')
    de = end_date.strftime('%d-%m-%Y')
    url = f"{BASE_URL}?comenzi&data_start={ds}&data_end={de}&limit=1000&apikey={API_KEY}"

    for attempt in range(retries + 1):
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Extended API'})
            with urllib.request.urlopen(req, timeout=90) as resp:
                data = json.loads(resp.read().decode('utf-8'))
            return data if isinstance(data, dict) else {}
        except Exception as e:
            if attempt < retries:
                print(f" retry({attempt+1})...", end='', flush=True)
                time.sleep(5)
            else:
                return None
    return None


def save_progress(email_dates, current_date, stats):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump({
            'email_dates': email_dates,  # {email: first_order_date}
            'current_date': current_date.isoformat(),
            'stats': stats,
        }, f)


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            data = json.load(f)
        return (
            data['email_dates'],
            datetime.fromisoformat(data['current_date']),
            data['stats']
        )
    return None


def main():
    if not API_KEY:
        print("✗ EJOLIE_API_KEY lipseste!")
        sys.exit(1)

    progress = load_progress()
    if progress:
        email_dates, current_date, stats = progress
        print(f"▶ Reluare de la {current_date.strftime('%d.%m.%Y')} ({len(email_dates)} emailuri deja)")
    else:
        email_dates = {}  # {email: "DD.MM.YYYY prima comanda"}
        current_date = datetime(2021, 1, 1)
        stats = {'total_orders': 0, 'chunks_ok': 0, 'chunks_fail': 0}

    print(f"═══ Extragere emailuri + date comenzi ejolie.ro ═══")
    print(f"  Perioada: {current_date.strftime('%d.%m.%Y')} → {CUTOFF.strftime('%d.%m.%Y')}")
    print(f"  Chunk: {CHUNK_DAYS} zile | Limită: {MAX_EMAILS}")
    print(flush=True)

    while current_date < CUTOFF:
        chunk_end = min(current_date + timedelta(days=CHUNK_DAYS - 1), CUTOFF - timedelta(days=1))
        label = f"{current_date.strftime('%d.%m')}-{chunk_end.strftime('%d.%m.%Y')}"
        print(f"  [{label}] ", end='', flush=True)

        orders = fetch_chunk(current_date, chunk_end)

        if orders is None:
            print(f"✗ TIMEOUT", flush=True)
            stats['chunks_fail'] += 1
            current_date = chunk_end + timedelta(days=1)
            save_progress(email_dates, current_date, stats)
            time.sleep(3)
            continue

        new_count = 0
        for oid, order in orders.items():
            client = order.get('client', {})
            if isinstance(client, dict):
                email = client.get('email', '').strip().lower()
                if email and '@' in email and len(email) > 5:
                    # Get order date
                    order_date = order.get('data', '')  # format from API
                    if not order_date:
                        order_date = order.get('data_ora', '')[:10] if order.get('data_ora') else ''

                    if email not in email_dates:
                        email_dates[email] = order_date
                        new_count += 1
                    # Keep earliest date
                    # (chunks go chronologically, so first seen = earliest)

        stats['total_orders'] += len(orders)
        stats['chunks_ok'] += 1

        print(f"✓ {len(orders)} comenzi, +{new_count} noi (total: {len(email_dates)})", flush=True)

        current_date = chunk_end + timedelta(days=1)
        save_progress(email_dates, current_date, stats)

        if len(email_dates) >= MAX_EMAILS:
            print(f"\n  ⚠ Limită {MAX_EMAILS} atinsă!")
            break

        time.sleep(0.8)

    # === SAVE RESULTS ===
    print(f"\n═══ REZULTAT ═══")
    print(f"  Total comenzi: {stats['total_orders']}")
    print(f"  Emailuri unice: {len(email_dates)}")
    print(f"  Chunks OK/Fail: {stats['chunks_ok']}/{stats['chunks_fail']}")

    # Sort by date then email
    items = sorted(email_dates.items(), key=lambda x: x[1])
    items = items[:MAX_EMAILS]

    # CSV
    csv_path = os.path.join(SCRIPT_DIR, 'emailuri_comenzi_ejolie.csv')
    with open(csv_path, 'w') as f:
        f.write('email,data_prima_comanda\n')
        for email, date in items:
            f.write(f"{email},{date}\n")
    print(f"  CSV: {csv_path}")

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Emailuri Comenzi"

        hfont = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        hfill = PatternFill('solid', fgColor='4472C4')

        for col, (header, width) in enumerate([('#', 8), ('Email', 45), ('Data Prima Comandă', 20)], 1):
            c = ws.cell(row=1, column=col, value=header)
            c.font = hfont
            c.fill = hfill
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[chr(64 + col)].width = width

        for i, (email, date) in enumerate(items, 1):
            ws.cell(row=i + 1, column=1, value=i)
            ws.cell(row=i + 1, column=2, value=email)
            ws.cell(row=i + 1, column=3, value=date)

        ws.freeze_panes = 'A2'

        xlsx_path = os.path.join(SCRIPT_DIR, 'emailuri_comenzi_ejolie.xlsx')
        wb.save(xlsx_path)
        print(f"  Excel: {xlsx_path}")

        # Copy to /tmp
        import shutil
        shutil.copy2(xlsx_path, '/tmp/emailuri_comenzi_ejolie.xlsx')
        shutil.copy2(csv_path, '/tmp/emailuri_comenzi_ejolie.csv')
        print(f"  Copiat în /tmp/")

    except Exception as e:
        print(f"  ⚠ Excel error: {e}")

    # Cleanup
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)

    print(f"\n✅ DONE — {len(items)} emailuri cu date extrase")


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""Ejolie.ro SEO Audit - scans product pages for SEO issues"""

import os, json, urllib.request, time, re, argparse
from html.parser import HTMLParser

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_FILE = os.path.join(SCRIPT_DIR, "stock_cache.json")
SEO_CACHE = os.path.join(SCRIPT_DIR, "seo_cache.json")


class MetaParser(HTMLParser):
    """Extract SEO elements from HTML"""
    def __init__(self):
        super().__init__()
        self.meta_title = ""
        self.meta_description = ""
        self.h1_tags = []
        self.images = []  # (src, alt)
        self.canonical = ""
        self._in_title = False
        self._in_h1 = False
        self._h1_text = ""

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        if tag == "title":
            self._in_title = True
        elif tag == "h1":
            self._in_h1 = True
            self._h1_text = ""
        elif tag == "meta":
            name = attrs_dict.get("name", "").lower()
            prop = attrs_dict.get("property", "").lower()
            content = attrs_dict.get("content", "")
            if name == "description" or prop == "og:description":
                if not self.meta_description:
                    self.meta_description = content
            if prop == "og:title" and not self.meta_title:
                self.meta_title = content
        elif tag == "img":
            src = attrs_dict.get("src", "")
            alt = attrs_dict.get("alt", "")
            if src:
                self.images.append({"src": src, "alt": alt})
        elif tag == "link":
            if attrs_dict.get("rel") == "canonical":
                self.canonical = attrs_dict.get("href", "")

    def handle_data(self, data):
        if self._in_title:
            self.meta_title += data.strip()
        if self._in_h1:
            self._h1_text += data.strip()

    def handle_endtag(self, tag):
        if tag == "title":
            self._in_title = False
        elif tag == "h1":
            self._in_h1 = False
            if self._h1_text:
                self.h1_tags.append(self._h1_text)


def fetch_page(url, timeout=30):
    """Fetch page HTML"""
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (compatible; EjolieSEOBot/1.0)"
        })
        start = time.time()
        resp = urllib.request.urlopen(req, timeout=timeout)
        html = resp.read().decode("utf-8", errors="ignore")
        load_time = time.time() - start
        return html, load_time, resp.status
    except Exception as e:
        return None, 0, str(e)


def analyze_page(url, prod_name, prod_desc):
    """Analyze a single page for SEO issues"""
    issues = []
    html, load_time, status = fetch_page(url)

    if html is None:
        issues.append({"type": "EROARE", "severity": "HIGH", "msg": f"Pagina nu se încarcă: {status}"})
        return issues, {}

    parser = MetaParser()
    try:
        parser.feed(html)
    except:
        issues.append({"type": "EROARE", "severity": "HIGH", "msg": "HTML parsing failed"})
        return issues, {}

    seo_data = {
        "meta_title": parser.meta_title,
        "meta_description": parser.meta_description,
        "h1_count": len(parser.h1_tags),
        "h1_text": parser.h1_tags[0] if parser.h1_tags else "",
        "images_total": len(parser.images),
        "images_no_alt": len([i for i in parser.images if not i["alt"]]),
        "load_time": round(load_time, 2),
        "canonical": parser.canonical,
    }

    # 1. Meta Title
    if not parser.meta_title:
        issues.append({"type": "META_TITLE", "severity": "HIGH", "msg": "Meta title lipsă"})
    elif len(parser.meta_title) < 30:
        issues.append({"type": "META_TITLE", "severity": "MEDIUM", "msg": f"Meta title prea scurt ({len(parser.meta_title)} chars)"})
    elif len(parser.meta_title) > 65:
        issues.append({"type": "META_TITLE", "severity": "LOW", "msg": f"Meta title prea lung ({len(parser.meta_title)} chars)"})

    # 2. Meta Description
    if not parser.meta_description:
        issues.append({"type": "META_DESC", "severity": "HIGH", "msg": "Meta description lipsă"})
    elif len(parser.meta_description) < 70:
        issues.append({"type": "META_DESC", "severity": "MEDIUM", "msg": f"Meta description prea scurtă ({len(parser.meta_description)} chars)"})
    elif len(parser.meta_description) > 160:
        issues.append({"type": "META_DESC", "severity": "LOW", "msg": f"Meta description prea lungă ({len(parser.meta_description)} chars)"})

    # 3. H1
    if len(parser.h1_tags) == 0:
        issues.append({"type": "H1", "severity": "HIGH", "msg": "H1 lipsă"})
    elif len(parser.h1_tags) > 1:
        issues.append({"type": "H1", "severity": "MEDIUM", "msg": f"Multiple H1 tags ({len(parser.h1_tags)})"})

    # 4. Images without alt
    no_alt = [i for i in parser.images if not i["alt"]]
    if no_alt:
        issues.append({"type": "IMG_ALT", "severity": "MEDIUM", "msg": f"{len(no_alt)}/{len(parser.images)} imagini fără alt tag"})

    # 5. Load time
    if load_time > 5:
        issues.append({"type": "SPEED", "severity": "HIGH", "msg": f"Pagină lentă ({load_time:.1f}s)"})
    elif load_time > 3:
        issues.append({"type": "SPEED", "severity": "MEDIUM", "msg": f"Pagină destul de lentă ({load_time:.1f}s)"})

    # 6. Check page content for product description
    # Look for description section in HTML
    desc_match = re.search(r'(?:DESCRIERE|descriere|description)(.*?)(?:</div>|</section>)', html, re.DOTALL | re.IGNORECASE)
    if desc_match:
        clean_desc = re.sub(r'<[^>]+>', '', desc_match.group(1)).strip()
        if len(clean_desc) < 100:
            issues.append({"type": "CONTENT", "severity": "MEDIUM", "msg": f"Descriere produs prea scurtă ({len(clean_desc)} chars)"})
    else:
        # Fallback: check total text on page
        page_text = re.sub(r'<[^>]+>', '', html)
        page_text = re.sub(r'\s+', ' ', page_text).strip()
        if len(page_text) < 500:
            issues.append({"type": "CONTENT", "severity": "HIGH", "msg": "Pagină cu puțin conținut text"})

    # 7. Product name too short
    if prod_name and len(prod_name) < 15:
        issues.append({"type": "TITLE", "severity": "LOW", "msg": f"Nume produs scurt ({len(prod_name)} chars)"})

    return issues, seo_data


def run_audit(brand=None, limit=None):
    """Run SEO audit on all products"""
    # Load product data from stock cache
    if not os.path.exists(CACHE_FILE):
        print("❌ Stock cache lipsă! Rulează mai întâi stock_cache_update.py")
        return []

    with open(CACHE_FILE) as f:
        cache = json.load(f)

    products = cache.get("products", {})
    print(f"🔍 SEO Audit - {len(products)} produse disponibile")

    results = []
    count = 0

    for pid, prod in products.items():
        if brand and brand.lower() not in prod.get("brand", "").lower():
            continue

        count += 1
        if limit and count > limit:
            break

        name = prod.get("nume", "")
        link = f"https://ejolie.ro/product/{name.lower().replace(' ', '-')}"

        # Try to get link from feed
        feed_path = os.path.join(SCRIPT_DIR, "product_feed.json")
        if os.path.exists(feed_path):
            with open(feed_path) as f:
                feed = json.load(f)
            for fp in feed:
                if fp.get("id") == pid:
                    link = fp.get("link", link)
                    break

        print(f"  🔍 [{count}] {name[:50]}...", end=" ", flush=True)

        issues, seo_data = analyze_page(link, name, prod.get("descriere", ""))

        severity_count = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
        for i in issues:
            severity_count[i["severity"]] += 1

        status = "🔴" if severity_count["HIGH"] > 0 else "🟡" if severity_count["MEDIUM"] > 0 else "🟢"
        print(f"{status} {severity_count['HIGH']}H {severity_count['MEDIUM']}M {severity_count['LOW']}L")

        results.append({
            "id": pid,
            "name": name,
            "cod": prod.get("cod", ""),
            "brand": prod.get("brand", ""),
            "link": link,
            "issues": issues,
            "seo_data": seo_data,
            "score": max(0, 100 - severity_count["HIGH"]*20 - severity_count["MEDIUM"]*10 - severity_count["LOW"]*5),
        })

        time.sleep(0.3)  # Be nice

    # Save results
    audit = {
        "updated": time.strftime("%Y-%m-%d %H:%M:%S"),
        "total_scanned": len(results),
        "results": results,
    }
    with open(SEO_CACHE, "w") as f:
        json.dump(audit, f, ensure_ascii=False, indent=1)

    return results


def export_xlsx(results, output=None):
    """Export audit to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        os.system("pip3 install openpyxl --break-system-packages -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    if output is None:
        output = "/home/ubuntu/seo_audit_ejolie.xlsx"

    wb = openpyxl.Workbook()

    # Colors
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Sheet 1: Sumar
    ws = wb.active
    ws.title = "Sumar Produse"
    headers = ["Produs", "Cod", "Scor SEO", "HIGH", "MEDIUM", "LOW", "Meta Title", "Meta Desc", "H1", "Img fără Alt", "Timp (s)", "URL"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(sorted(results, key=lambda x: x["score"]), 2):
        seo = r.get("seo_data", {})
        high = sum(1 for i in r["issues"] if i["severity"] == "HIGH")
        med = sum(1 for i in r["issues"] if i["severity"] == "MEDIUM")
        low = sum(1 for i in r["issues"] if i["severity"] == "LOW")

        data = [
            r["name"], r["cod"], r["score"], high, med, low,
            seo.get("meta_title", "")[:50],
            seo.get("meta_description", "")[:50],
            seo.get("h1_text", "")[:50],
            seo.get("images_no_alt", 0),
            seo.get("load_time", 0),
            r["link"],
        ]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col == 3:  # Score
                if val >= 80:
                    cell.fill = green_fill
                elif val >= 50:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill
                cell.font = Font(bold=True)

    widths = [40, 15, 10, 7, 9, 7, 35, 35, 35, 12, 8, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Sheet 2: Toate Problemele
    ws2 = wb.create_sheet("Probleme Detaliate")
    headers2 = ["Produs", "Cod", "Tip Problemă", "Severitate", "Detalii", "URL"]

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2
    for r in sorted(results, key=lambda x: x["score"]):
        for issue in r["issues"]:
            data = [r["name"], r["cod"], issue["type"], issue["severity"], issue["msg"], r["link"]]
            for col, val in enumerate(data, 1):
                cell = ws2.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col == 4:
                    if val == "HIGH":
                        cell.fill = red_fill
                    elif val == "MEDIUM":
                        cell.fill = yellow_fill
                    else:
                        cell.fill = green_fill
            row_idx += 1

    widths2 = [40, 15, 15, 12, 50, 50]
    for col, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Sheet 3: Statistici
    ws3 = wb.create_sheet("Statistici")
    total = len(results)
    avg_score = sum(r["score"] for r in results) / total if total else 0
    high_total = sum(1 for r in results for i in r["issues"] if i["severity"] == "HIGH")
    med_total = sum(1 for r in results for i in r["issues"] if i["severity"] == "MEDIUM")
    low_total = sum(1 for r in results for i in r["issues"] if i["severity"] == "LOW")
    no_meta_title = sum(1 for r in results if any(i["type"] == "META_TITLE" and i["severity"] == "HIGH" for i in r["issues"]))
    no_meta_desc = sum(1 for r in results if any(i["type"] == "META_DESC" and i["severity"] == "HIGH" for i in r["issues"]))
    no_h1 = sum(1 for r in results if any(i["type"] == "H1" and i["severity"] == "HIGH" for i in r["issues"]))

    stats = [
        ("📊 SEO AUDIT EJOLIE.RO", ""),
        ("", ""),
        ("Produse scanate", total),
        ("Scor mediu SEO", f"{avg_score:.0f}/100"),
        ("", ""),
        ("🔴 Probleme HIGH", high_total),
        ("🟡 Probleme MEDIUM", med_total),
        ("🟢 Probleme LOW", low_total),
        ("", ""),
        ("Fără Meta Title", no_meta_title),
        ("Fără Meta Description", no_meta_desc),
        ("Fără H1", no_h1),
    ]

    for row_idx, (label, val) in enumerate(stats, 1):
        ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True if row_idx == 1 else False, size=14 if row_idx == 1 else 11)
        ws3.cell(row=row_idx, column=2, value=val)
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 15

    wb.save(output)
    print(f"\n✅ Excel salvat: {output}")
    return output


def print_summary(results):
    """Print text summary"""
    total = len(results)
    avg_score = sum(r["score"] for r in results) / total if total else 0
    high_total = sum(1 for r in results for i in r["issues"] if i["severity"] == "HIGH")
    med_total = sum(1 for r in results for i in r["issues"] if i["severity"] == "MEDIUM")

    print(f"\n📊 SEO AUDIT SUMMARY")
    print("━" * 50)
    print(f"Produse scanate: {total}")
    print(f"Scor mediu: {avg_score:.0f}/100")
    print(f"🔴 HIGH: {high_total} | 🟡 MEDIUM: {med_total}")
    print()

    # Worst products
    worst = sorted(results, key=lambda x: x["score"])[:10]
    print("⚠️ Top 10 produse cu probleme:")
    for r in worst:
        print(f"  {r['score']:3d}/100 | {r['name'][:45]}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--brand", default="ejolie")
    parser.add_argument("--limit", type=int, default=None, help="Limit products to scan")
    parser.add_argument("--format", choices=["text", "xlsx"], default="xlsx")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    print(f"🔍 Starting SEO Audit for brand: {args.brand}")
    results = run_audit(brand=args.brand, limit=args.limit)

    if not results:
        print("❌ No products found")
        return

    print_summary(results)

    if args.format == "xlsx":
        export_xlsx(results, output=args.output)


if __name__ == "__main__":
    main()

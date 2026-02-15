#!/usr/bin/env python3
"""Fill Extended import template with product data from API"""

import os, json, re, argparse
import openpyxl
from openpyxl.styles import Font, Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Color extraction from product names
COLOR_MAP = {
    "neagra": "Negru", "negru": "Negru", "negra": "Negru", "black": "Negru",
    "alba": "Alb", "alb": "Alb", "white": "Alb", "ivory": "Ivory", "ivoar": "Ivory",
    "rosie": "Rosu", "rosu": "Rosu", "red": "Rosu",
    "verde": "Verde", "green": "Verde",
    "albastra": "Albastru", "albastru": "Albastru", "blue": "Albastru",
    "galbena": "Galben", "galben": "Galben",
    "roz": "Roz", "pink": "Roz", "fucsia": "Fucsia",
    "bordo": "Bordo", "burgundy": "Bordo", "visiniu": "Visiniu",
    "mov": "Mov", "lila": "Lila", "purple": "Mov",
    "gri": "Gri", "grey": "Gri", "gray": "Gri",
    "bej": "Bej", "nude": "Nude", "crem": "Crem",
    "turcoaz": "Turcoaz", "coral": "Coral", "portocalie": "Portocaliu",
    "argintiu": "Argintiu", "argintii": "Argintiu", "auriu": "Auriu", "aurie": "Auriu",
    "maro": "Maro", "camel": "Camel", "kaki": "Kaki",
    "multicolor": "Multicolor", "imprimeu": "Imprimeu",
    "bleumarin": "Bleumarin", "navy": "Bleumarin",
    "somon": "Somon", "lavanda": "Lavanda", "mint": "Mint",
}

def extract_color(name):
    """Extract color from product name"""
    words = name.lower().split()
    for word in words:
        if word in COLOR_MAP:
            return COLOR_MAP[word]
    # Try compound: "cu dungi negre" etc
    name_lower = name.lower()
    for key, val in COLOR_MAP.items():
        if key in name_lower:
            return val
    return ""
CACHE_FILE = os.path.join(SCRIPT_DIR, "stock_cache.json")
FEED_FILE = os.path.join(SCRIPT_DIR, "product_feed.json")

# Load env
env_path = os.path.join(SCRIPT_DIR, ".env")
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            if '=' in line and not line.startswith('#'):
                k, v = line.strip().split('=', 1)
                os.environ[k] = v


def load_full_products(brand=None):
    """Load products from cache + feed for full details"""
    import urllib.request, time

    api_key = os.environ.get("EJOLIE_API_KEY", "")
    api_url = os.environ.get("EJOLIE_API_URL", "https://ejolie.ro/api/")

    # Load feed for IDs and links
    with open(FEED_FILE) as f:
        feed = json.load(f)

    # Filter by brand
    product_ids = []
    feed_map = {}
    for p in feed:
        if brand and brand.lower() not in p.get("brand", "").lower():
            continue
        pid = p.get("id", "")
        if pid:
            product_ids.append(pid)
            feed_map[pid] = p

    print(f"📦 {len(product_ids)} produse {brand or 'toate'}")

    # Fetch full details from API in batches
    all_products = {}
    batch_size = 20
    total = len(product_ids)

    for i in range(0, total, batch_size):
        batch = product_ids[i:i+batch_size]
        ids_str = ",".join(batch)
        url = f"{api_url}?produse&id_produse={ids_str}&apikey={api_key}"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            data = json.loads(urllib.request.urlopen(req, timeout=180).read().decode("utf-8"))
            for pid, prod in data.items():
                if not isinstance(prod, dict):
                    continue
                prod["_feed"] = feed_map.get(pid, {})
                all_products[pid] = prod
            done = min(i + batch_size, total)
            print(f"  📡 {done}/{total}...")
        except Exception as e:
            print(f"  ⚠️ Batch error: {e}")
        time.sleep(0.5)

    return all_products


def clean_html(text):
    """Remove HTML tags"""
    if not text:
        return ""
    return re.sub(r'<[^>]+>', '', str(text)).strip()


def export_template(products, output=None, only_in_stock=True):
    """Export products to Extended import template"""
    if output is None:
        output = "/home/ubuntu/ejolie_import_template.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "import-test-1"

    # Headers (matching Extended template)
    headers = [
        "Nume produs", "Descriere", "Categorie", "Brand",
        "Optiune 1", "Optiune 2", "Optiune 3", "Optiune 4", "Optiune 5",
        "Furnizor",
        "Pret vanzare – LEI (cu TVA)", "Pret intrare – LEI (fara sau cu TVA in functie de setare)",
        "Adaos %", "Discount %", "Moneda", "Cod produs",
        "Stoc", "Stoc fizic", "Greutate (KG)",
        "Imagine 1", "Imagine 2", "Imagine 3", "Imagine 4", "Imagine 5",
        "Imagine 6", "Imagine 7", "Imagine 8", "Imagine 9", "Imagine 10",
        "Imagine 11", "Imagine 12", "Imagine 13", "Imagine 14", "Imagine 15",
        "Nume specificatie 1", "Valoare specificatie 1",
        "Nume specificatie 2", "Valoare specificatie 2",
    ]

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_idx = 2

    for pid, prod in products.items():
        optiuni = prod.get("optiuni", {})
        if not isinstance(optiuni, dict):
            optiuni = {}

        # Get categories
        categorii = prod.get("categorii", {})
        cat_names = []
        if isinstance(categorii, dict):
            for cid, cat in categorii.items():
                if isinstance(cat, dict):
                    cat_names.append(cat.get("nume", ""))
        elif isinstance(categorii, list):
            for cat in categorii:
                if isinstance(cat, dict):
                    cat_names.append(cat.get("nume", ""))
        cat_str = ">".join(cat_names) if cat_names else ""

        # Feed data
        feed = prod.get("_feed", {})
        link = feed.get("link", "")

        # Brand
        brand_data = prod.get("brand", {})
        brand_name = brand_data.get("nume", "") if isinstance(brand_data, dict) else str(brand_data)

        # Images
        imagini = prod.get("imagini", [])
        if not isinstance(imagini, list):
            imagini = []

        # Specificatii
        specs = prod.get("specificatii", {})
        spec_list = []
        if isinstance(specs, dict):
            for sid, spec in specs.items():
                if isinstance(spec, dict):
                    name = spec.get("nume", "")
                    vals = spec.get("valoare", [])
                    if isinstance(vals, list):
                        val_str = ", ".join(str(v) for v in vals)
                    else:
                        val_str = str(vals)
                    spec_list.append((name, val_str))
        elif isinstance(specs, list):
            for spec in specs:
                if isinstance(spec, dict):
                    name = spec.get("nume", "")
                    vals = spec.get("valoare", [])
                    if isinstance(vals, list):
                        val_str = ", ".join(str(v) for v in vals)
                    else:
                        val_str = str(vals)
                    spec_list.append((name, val_str))

        # Prices
        pret = prod.get("pret", "0")
        pret_discount = prod.get("pret_discount", "0")

        # Calculate discount %
        try:
            p = float(pret) if pret else 0
            pd = float(pret_discount) if pret_discount else 0
            discount_pct = round((1 - pd/p) * 100, 2) if p > 0 and pd > 0 and pd < p else 0
        except:
            discount_pct = 0

        if optiuni:
            # One row per size option
            for oid, opt in optiuni.items():
                if not isinstance(opt, dict):
                    continue

                stoc_info = opt.get("stoc", "")
                stoc_fizic = opt.get("stoc_fizic", 0)

                if only_in_stock and "In stoc" not in str(stoc_info):
                    continue

                size_name = opt.get("nume_optiune", "")
                opt_pret = opt.get("pret", pret)
                opt_pret_disc = opt.get("pret_discount", pret_discount)

                row = [
                    prod.get("nume", ""),
                    clean_html(prod.get("descriere", "")),
                    cat_str,
                    brand_name,
                    f"Marime:{size_name}" if size_name else "",
                    f"Culoare:{extract_color(prod.get('nume', ''))}" if extract_color(prod.get("nume", "")) else "",
                    "", "", "",  # Optiune 3-5
                    "",  # Furnizor
                    float(opt_pret) if opt_pret else 0,
                    0,  # Pret intrare
                    0,  # Adaos
                    discount_pct,
                    "RON",
                    prod.get("cod_produs", ""),
                    stoc_info,
                    int(stoc_fizic) if stoc_fizic else 0,
                    0,  # Greutate
                ]

                # Add images (up to 15)
                for img_idx in range(15):
                    if img_idx < len(imagini):
                        row.append(imagini[img_idx])
                    else:
                        row.append("")

                # Add color as first spec, then other specs
                color = extract_color(prod.get("nume", ""))
                if color:
                    row.append("Culoare")
                    row.append(color)
                    if len(spec_list) > 0:
                        row.append(spec_list[0][0])
                        row.append(spec_list[0][1])
                    else:
                        row.append("")
                        row.append("")
                else:
                    for spec_idx in range(2):
                        if spec_idx < len(spec_list):
                            row.append(spec_list[spec_idx][0])
                            row.append(spec_list[spec_idx][1])
                        else:
                            row.append("")
                            row.append("")

                for col, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col, value=val)
                row_idx += 1
        else:
            # No sizes - single row
            stoc_info = prod.get("stoc", "")
            stoc_fizic = prod.get("stoc_fizic", 0)

            if only_in_stock and "In stoc" not in str(stoc_info):
                continue

            row = [
                prod.get("nume", ""),
                clean_html(prod.get("descriere", "")),
                cat_str,
                brand_name,
                "",
                f"Culoare:{extract_color(prod.get('nume', ''))}" if extract_color(prod.get("nume", "")) else "",
                "", "", "",  # Optiuni 3-5
                "",  # Furnizor
                float(pret) if pret else 0,
                0,
                0,
                discount_pct,
                "RON",
                prod.get("cod_produs", ""),
                stoc_info,
                int(stoc_fizic) if stoc_fizic else 0,
                0,
            ]

            for img_idx in range(15):
                if img_idx < len(imagini):
                    row.append(imagini[img_idx])
                else:
                    row.append("")

            color = extract_color(prod.get("nume", ""))
            if color:
                row.append("Culoare")
                row.append(color)
                if len(spec_list) > 0:
                    row.append(spec_list[0][0])
                    row.append(spec_list[0][1])
                else:
                    row.append("")
                    row.append("")
            else:
                for spec_idx in range(2):
                    if spec_idx < len(spec_list):
                        row.append(spec_list[spec_idx][0])
                        row.append(spec_list[spec_idx][1])
                    else:
                        row.append("")
                        row.append("")

            for col, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col, value=val)
            row_idx += 1

    # Auto-width for first columns
    widths = [40, 60, 30, 15, 15, 15, 10, 10, 10, 15, 15, 15, 8, 10, 6, 15, 10, 10, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    total_rows = row_idx - 2
    wb.save(output)
    print(f"✅ Template salvat: {output} ({total_rows} rânduri)")
    return output


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--brand", default="ejolie")
    parser.add_argument("--all", action="store_true", help="Include out of stock")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    products = load_full_products(brand=args.brand)
    export_template(products, output=args.output, only_in_stock=not args.all)


if __name__ == "__main__":
    main()

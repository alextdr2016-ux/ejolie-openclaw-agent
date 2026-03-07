#!/usr/bin/env python3
"""
sync_trendyol_stock.py - Sincronizare stoc Trendyol cu ejolie.ro
Citește exportul Trendyol, actualizează coloana Stoc cu date din Extended API,
salvează Excel-ul actualizat și (opțional) trimite pe Telegram.

Utilizare:
    python3 sync_trendyol_stock.py --input Products_07_03_2026.xlsx
    python3 sync_trendyol_stock.py --input Products_07_03_2026.xlsx --telegram
    python3 sync_trendyol_stock.py --input Products_07_03_2026.xlsx --output stoc_actualizat.xlsx --telegram
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# === CONFIG ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, '..', '.env')
if not os.path.exists(ENV_PATH):
    ENV_PATH = os.path.join(SCRIPT_DIR, '.env')
load_dotenv(ENV_PATH)

API_KEY = os.getenv('EJOLIE_API_KEY')
API_BASE = 'https://ejolie.ro/api/'
HEADERS = {'User-Agent': 'Mozilla/5.0'}

TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
TELEGRAM_CHAT_ID = os.getenv('TELEGRAM_CHAT_ID', '44151343')

PAGE_SIZE = 50


# ═══════════════════════════════════════════
#  FETCH ALL PRODUCTS FROM EJOLIE API
# ═══════════════════════════════════════════

def fetch_all_products():
    """Fetch toate produsele din ejolie.ro API cu paginare."""
    all_products = {}
    page = 1

    print("  Fetch produse din ejolie.ro API...")
    while True:
        url = f"{API_BASE}?produse&apikey={API_KEY}&pagina={page}&limit={PAGE_SIZE}"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=60)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            print(f"  ✗ Eroare pagina {page}: {e}")
            break

        if not data or (isinstance(data, dict) and not data):
            break

        # API returnează dict cu ID-uri ca key-uri
        if isinstance(data, dict):
            products = data
        elif isinstance(data, list):
            products = {str(p.get('id_produs', '')): p for p in data if isinstance(p, dict)}
        else:
            break

        if not products:
            break

        all_products.update(products)
        count = len(products)
        print(
            f"    Pagina {page}: {count} produse (total: {len(all_products)})")

        if count < PAGE_SIZE:
            break

        page += 1
        time.sleep(0.3)

    print(f"  ✓ Total produse: {len(all_products)}")
    return all_products


def fetch_product_details(product_id):
    """Fetch detalii produs individual (fallback)."""
    url = f"{API_BASE}?id_produs={product_id}&apikey={API_KEY}"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        if str(product_id) in data:
            return data[str(product_id)]
    except:
        pass
    return None


# ═══════════════════════════════════════════
#  BUILD STOCK LOOKUP
# ═══════════════════════════════════════════

def build_stock_lookup(products):
    """
    Construiește un dict de lookup:
      cod_produs → {
        'id': id_produs,
        'nume': nume,
        'sizes': { '36': stoc_fizic, '38': stoc_fizic, ... },
        'total_stock': total
      }
    """
    lookup = {}

    for pid, product in products.items():
        cod = product.get('cod_produs', '').strip()
        if not cod:
            continue

        sizes = {}
        total_stock = 0
        optiuni = product.get('optiuni', {})
        for opt in optiuni.values():
            marime = str(opt.get('nume_optiune', '')).strip()
            stoc = int(opt.get('stoc_fizic', 0))
            sizes[marime] = stoc
            total_stock += stoc

        # Adaugăm și stocul din furnizor dacă optiuni e gol
        if not optiuni:
            furnizori = product.get('furnizor', [])
            for f in furnizori:
                opt_id = f.get('optiune_id', '')
                stoc = int(f.get('stoc_numeric', 0))
                total_stock += stoc

        lookup[cod] = {
            'id': pid,
            'nume': product.get('nume', ''),
            'sizes': sizes,
            'total_stock': total_stock,
        }

    return lookup


# ═══════════════════════════════════════════
#  MATCHING LOGIC
# ═══════════════════════════════════════════

def normalize_code(code):
    """Normalizează un cod eliminând spații și convertind la uppercase."""
    return code.strip().upper() if code else ''


def find_stock_for_row(trendyol_code, trendyol_size, lookup):
    """
    Încearcă să găsească stocul din ejolie pentru un rând Trendyol.

    Strategii de matching (în ordine):
    1. Match exact pe cod_produs
    2. Match pe codul de bază (fără culoare-mărime din suffix)
    3. Match pe prefix FBR- cu variante

    Returns: (stoc_fizic, match_type, matched_code) sau (None, 'no_match', None)
    """
    trendyol_size = str(trendyol_size).strip()

    # Strategy 1: Exact match
    if trendyol_code in lookup:
        product = lookup[trendyol_code]
        if trendyol_size in product['sizes']:
            return product['sizes'][trendyol_size], 'exact', trendyol_code
        # Mărimea nu e găsită - produsul există dar fără acea mărime = stoc 0
        return 0, 'exact_no_size', trendyol_code

    # Strategy 2: Case-insensitive match
    trendyol_upper = normalize_code(trendyol_code)
    for cod, product in lookup.items():
        if normalize_code(cod) == trendyol_upper:
            if trendyol_size in product['sizes']:
                return product['sizes'][trendyol_size], 'case_insensitive', cod
            return 0, 'case_no_size', cod

    # Strategy 3: Base code extraction
    # ART2715-lila-40 → try ART2715
    # 7722-caramizie-38 → try 7722
    # EJ77318-Negru-Buline-36 → try EJ77318
    base_code = extract_base_code(trendyol_code)
    if base_code and base_code != trendyol_code:
        if base_code in lookup:
            product = lookup[base_code]
            if trendyol_size in product['sizes']:
                return product['sizes'][trendyol_size], 'base_code', base_code
            return 0, 'base_no_size', base_code

        # Case insensitive base
        base_upper = normalize_code(base_code)
        for cod, product in lookup.items():
            if normalize_code(cod) == base_upper:
                if trendyol_size in product['sizes']:
                    return product['sizes'][trendyol_size], 'base_case', cod
                return 0, 'base_case_no_size', cod

    # Strategy 4: Fuzzy - caută cod care CONȚINE codul Trendyol sau invers
    for cod, product in lookup.items():
        cod_clean = normalize_code(cod)
        if trendyol_upper in cod_clean or cod_clean in trendyol_upper:
            if len(min(trendyol_upper, cod_clean)) >= 4:  # minim 4 char match
                if trendyol_size in product['sizes']:
                    return product['sizes'][trendyol_size], 'fuzzy', cod
                return 0, 'fuzzy_no_size', cod

    return None, 'no_match', None


def extract_base_code(code):
    """
    Extrage codul de bază dintr-un cod Trendyol care poate avea culoare-mărime.

    ART2715-lila-40 → ART2715
    7722-caramizie-38 → 7722
    EJ77318-Negru-Buline-36 → EJ77318
    FBR-B25280 → FBR-B25280 (nu modifică FBR- coduri)
    3049-neagra-44 → 3049
    964205-verde-40 → 964205
    """
    if not code:
        return code

    # FBR- codes sunt deja clean
    if code.startswith('FBR-'):
        return code

    # Pattern: BASE-culoare-marime sau BASE-culoare
    # Culori cunoscute în română
    colors = [
        'neagra', 'negru', 'alba', 'alb', 'rosie', 'rosu', 'albastra', 'albastru',
        'verde', 'mov', 'lila', 'fucsia', 'roz', 'rozprafuit', 'roze', 'bleu',
        'bordo', 'maro', 'bej', 'galbena', 'galben', 'portocalie', 'portocaliu',
        'caramizie', 'caramiziu', 'ivoar', 'turcoaz', 'rozdungi', 'verdedeschis',
        'multicolor', 'auriu', 'argintiu', 'crem', 'somon', 'gri',
    ]

    parts = code.split('-')
    if len(parts) >= 2:
        # Verificăm dacă al doilea segment e o culoare
        second = parts[1].lower()
        for color in colors:
            if second == color or second.startswith(color):
                return parts[0]

    # ART codes: ART2715-lila-40 → ART2715
    if code.startswith('ART') and len(parts) >= 2:
        return parts[0]

    return code


# ═══════════════════════════════════════════
#  EXCEL UPDATE
# ═══════════════════════════════════════════

def update_trendyol_excel(input_path, output_path, lookup):
    """Actualizează coloana Stoc din Excel-ul Trendyol cu date din ejolie.ro."""
    wb = load_workbook(input_path)
    ws = wb['Produse']

    # Găsim indexurile coloanelor
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[val] = col

    cod_col = headers.get('Codul modelului')
    size_col = headers.get('Mărime')
    stoc_col = headers.get('Stoc')
    name_col = headers.get('Numele produsului')

    if not all([cod_col, size_col, stoc_col]):
        print("  ✗ Coloane lipsă în Excel!")
        return None

    # Stiluri pentru celulele modificate
    changed_fill = PatternFill('solid', fgColor='C6EFCE')  # Verde deschis
    zero_fill = PatternFill('solid', fgColor='FFC7CE')     # Roșu deschis
    no_match_fill = PatternFill('solid', fgColor='FFEB9C')  # Galben

    stats = {
        'total': 0, 'matched': 0, 'no_match': 0,
        'stock_changed': 0, 'stock_zero': 0, 'stock_same': 0,
        'match_types': {}
    }

    print(f"\n  Actualizare stoc ({ws.max_row - 1} rânduri)...\n")

    for row in range(2, ws.max_row + 1):
        trendyol_code = str(
            ws.cell(row=row, column=cod_col).value or '').strip()
        trendyol_size = str(
            ws.cell(row=row, column=size_col).value or '').strip()
        old_stock = ws.cell(row=row, column=stoc_col).value
        name = str(ws.cell(row=row, column=name_col).value or '')

        if not trendyol_code:
            continue

        stats['total'] += 1

        stoc, match_type, matched_code = find_stock_for_row(
            trendyol_code, trendyol_size, lookup)

        # Track match types
        stats['match_types'][match_type] = stats['match_types'].get(
            match_type, 0) + 1

        if stoc is not None:
            stats['matched'] += 1
            old_val = int(old_stock) if old_stock else 0

            # Update stock
            ws.cell(row=row, column=stoc_col, value=stoc)

            if stoc == 0:
                ws.cell(row=row, column=stoc_col).fill = zero_fill
                stats['stock_zero'] += 1
                if old_val != stoc:
                    print(
                        f"    ⚠ STOC 0: {name} (mărime {trendyol_size}) | era: {old_val} → 0")
            elif stoc != old_val:
                ws.cell(row=row, column=stoc_col).fill = changed_fill
                stats['stock_changed'] += 1
                print(
                    f"    ✎ {name} (mărime {trendyol_size}) | {old_val} → {stoc} [{match_type}]")
            else:
                stats['stock_same'] += 1
        else:
            stats['no_match'] += 1
            ws.cell(row=row, column=stoc_col).fill = no_match_fill
            print(
                f"    ? NO MATCH: {trendyol_code} | {name} (mărime {trendyol_size})")

    wb.save(output_path)
    return stats


# ═══════════════════════════════════════════
#  TELEGRAM
# ═══════════════════════════════════════════

def send_telegram_file(filepath, caption=''):
    if not TELEGRAM_BOT_TOKEN:
        print("  ⚠ TELEGRAM_BOT_TOKEN nu e setat în .env")
        return False
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    try:
        with open(filepath, 'rb') as f:
            resp = requests.post(url,
                                 data={'chat_id': TELEGRAM_CHAT_ID,
                                       'caption': caption, 'parse_mode': 'HTML'},
                                 files={'document': f}, timeout=60)
        return resp.status_code == 200
    except Exception as e:
        print(f"  ✗ Telegram eroare: {e}")
        return False


# ═══════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Sincronizare stoc Trendyol cu ejolie.ro')
    parser.add_argument('--input', type=str, required=True,
                        help='Fișier Excel Trendyol export')
    parser.add_argument('--output', type=str, default=None,
                        help='Fișier output (default: stoc_trendyol_DATA.xlsx)')
    parser.add_argument('--telegram', action='store_true',
                        help='Trimite pe Telegram')
    parser.add_argument('--cache', type=str, default=None,
                        help='Folosește cache JSON în loc de API (product_feed.json)')
    args = parser.parse_args()

    if not API_KEY:
        print("✗ EJOLIE_API_KEY nu e setat în .env!")
        sys.exit(1)

    if not os.path.exists(args.input):
        print(f"✗ Fișierul {args.input} nu există!")
        sys.exit(1)

    output_path = args.output or os.path.join(
        SCRIPT_DIR, f"stoc_trendyol_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    )

    print(f"═══ Sync Stoc Trendyol ═══")
    print(f"  Input:    {args.input}")
    print(f"  Output:   {output_path}")
    print(f"  Telegram: {'Da' if args.telegram else 'Nu'}")
    print()

    # === STEP 1: Fetch produse ejolie ===
    if args.cache and os.path.exists(args.cache):
        print(f"  Citesc din cache: {args.cache}")
        with open(args.cache, 'r') as f:
            cache_data = json.load(f)
        # Cache poate fi lista sau dict
        if isinstance(cache_data, list):
            products = {}
            for p in cache_data:
                pid = str(p.get('id', p.get('id_produs', '')))
                products[pid] = p
        else:
            products = cache_data
        print(f"  ✓ {len(products)} produse din cache")
    else:
        products = fetch_all_products()

    if not products:
        print("  ✗ Niciun produs găsit!")
        sys.exit(1)

    # === STEP 2: Build lookup ===
    print("\n  Construiesc lookup stoc...")
    lookup = build_stock_lookup(products)
    print(f"  ✓ {len(lookup)} coduri unice în lookup")

    # Debug: show some lookup entries
    print(f"\n  Sample lookup (primele 5):")
    for i, (cod, info) in enumerate(lookup.items()):
        if i >= 5:
            break
        print(f"    {cod} → sizes: {info['sizes']}")

    # === STEP 3: Update Excel ===
    stats = update_trendyol_excel(args.input, output_path, lookup)

    if not stats:
        sys.exit(1)

    # === STEP 4: Report ===
    print(f"""
═══ SINCRONIZARE COMPLETĂ ═══
  Total rânduri:       {stats['total']}
  Matched:             {stats['matched']} ({stats['matched']*100//max(stats['total'], 1)}%)
  No match:            {stats['no_match']}
  Stoc modificat:      {stats['stock_changed']}
  Stoc zero (epuizat): {stats['stock_zero']}
  Stoc neschimbat:     {stats['stock_same']}

  Match types:""")
    for mt, count in sorted(stats['match_types'].items(), key=lambda x: -x[1]):
        print(f"    {mt}: {count}")

    print(f"\n  📁 Fișier salvat: {output_path}")

    # === STEP 5: Telegram ===
    if args.telegram:
        print(f"\n  📤 Trimit pe Telegram...")
        caption = (
            f"📦 <b>Stoc Trendyol actualizat — {datetime.now().strftime('%d.%m.%Y %H:%M')}</b>\n\n"
            f"✅ Matched: {stats['matched']}/{stats['total']}\n"
            f"✎ Stoc modificat: {stats['stock_changed']}\n"
            f"⚠ Stoc zero: {stats['stock_zero']}\n"
            f"❓ No match: {stats['no_match']}"
        )
        if send_telegram_file(output_path, caption=caption):
            print("  ✓ Trimis pe Telegram!")
        else:
            print("  ✗ Eroare Telegram")


if __name__ == '__main__':
    main()

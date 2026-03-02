#!/usr/bin/env python3
"""
generate_nir.py v2 - Generare automată NIR (Notă de Intrare Recepție) pentru ejolie.ro
Trage date din Extended API, generează Excel și (opțional) trimite pe Telegram.

Utilizare:
    python3 generate_nir.py --start 12356 --end 12415
    python3 generate_nir.py --ids 12356,12358,12360
    python3 generate_nir.py --start 12356 --end 12415 --telegram
    python3 generate_nir.py --start 12356 --end 12415 --output nir_martie.xlsx --telegram
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === CONFIG ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, '..', '.env')
if not os.path.exists(ENV_PATH):
    ENV_PATH = os.path.join(SCRIPT_DIR, '.env')
load_dotenv(ENV_PATH)

API_KEY = os.getenv('EJOLIE_API_KEY')
API_BASE = 'https://ejolie.ro/api/'
HEADERS = {'User-Agent': 'Mozilla/5.0'}

TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
TELEGRAM_CHAT_ID = os.getenv('TELEGRAM_CHAT_ID', '44151343')


# ═══════════════════════════════════════════
#  API FETCH
# ═══════════════════════════════════════════

def fetch_product(product_id):
    """Fetch un singur produs din Extended API."""
    url = f"{API_BASE}?id_produs={product_id}&apikey={API_KEY}"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        if str(product_id) in data:
            return data[str(product_id)]
        return None
    except Exception as e:
        print(f"  ✗ Eroare la produsul {product_id}: {e}")
        return None


# ═══════════════════════════════════════════
#  DATA EXTRACTION
# ═══════════════════════════════════════════

def extract_nir_data(product):
    """Extrage datele necesare pentru NIR dintr-un produs API."""
    pid = product.get('id_produs', '')
    nume_full = product.get('nume', '')

    prefixes = ['Rochie ', 'Bluza ', 'Fusta ', 'Sacou ',
                'Palton ', 'Pantaloni ', 'Pulover ', 'Compleu ']
    nume_scurt = nume_full
    for prefix in prefixes:
        if nume_full.startswith(prefix):
            rest = nume_full[len(prefix):]
            parts = rest.rsplit(' ', 1)
            if len(parts) > 1:
                nume_scurt = parts[0]
            else:
                nume_scurt = rest
            break

    culoare = ''
    for spec in product.get('specificatii', []):
        if spec['nume'] == 'Culoare':
            culoare = ', '.join(spec.get('valoare', []))
            break
    if not culoare:
        parts = nume_full.rsplit(' ', 1)
        if len(parts) > 1:
            culoare = parts[-1]

    cod = product.get('cod_produs', '')

    optiuni = product.get('optiuni', {})
    marimi = []
    total_buc = 0
    for opt in optiuni.values():
        marimi.append(opt.get('nume_optiune', ''))
        total_buc += int(opt.get('stoc_fizic', 0))

    try:
        marimi_sorted = sorted(marimi, key=lambda x: int(x))
    except ValueError:
        marimi_sorted = sorted(marimi)
    marimi_str = f"{marimi_sorted[0]}-{marimi_sorted[-1]}" if marimi_sorted else ''

    pret_iesire = float(product.get('pret', 0))
    pret_discount = float(product.get('pret_discount', 0))
    if pret_discount > 0:
        pret_iesire = pret_discount

    pret_intrare = 0
    furnizori = product.get('furnizor', [])
    if furnizori:
        pret_intrare = float(furnizori[0].get('pret_achizitie', '0'))

    cota_tva = int(product.get('cota_tva', 19))

    return {
        'id': pid, 'nume': nume_scurt, 'culoare': culoare, 'cod': cod,
        'marimi': marimi_str, 'buc': total_buc, 'pret_intrare': pret_intrare,
        'pret_iesire': pret_iesire, 'cota_tva': cota_tva,
    }


# ═══════════════════════════════════════════
#  EXCEL GENERATION
# ═══════════════════════════════════════════

def generate_nir_excel(products_data, output_path):
    """Generează fișierul Excel NIR cu formule."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NIR"

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_align = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(horizontal='center', vertical='center')
    data_align_left = Alignment(horizontal='left', vertical='center')
    money_format = '#,##0.00'
    int_format = '#,##0'
    total_font = Font(name='Arial', bold=True, size=11)
    total_fill = PatternFill('solid', fgColor='D9E2F3')
    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Titlu
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = f"NIR - Notă de Intrare Recepție — {datetime.now().strftime('%d.%m.%Y')}"
    title_cell.font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Header
    headers = ['Id produs', 'Nume', 'Culoare', 'Cod', 'Mărimi', 'Buc', 'Preț intrare', 'Preț ieșire',
               'Valoare intrare', 'Valoare ieșire', 'TVA', 'Adaos comercial', 'Cotă TVA', 'Total intrare']
    col_widths = [12, 20, 16, 16, 12, 8, 14, 14, 16, 16, 16, 16, 10, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 35

    # Date produse
    start_row = 4
    for i, prod in enumerate(products_data):
        row = start_row + i

        # Coloane cu date directe
        for col, key, align in [(1, 'id', data_align), (2, 'nume', data_align_left), (3, 'culoare', data_align),
                                (4, 'cod', data_align), (5, 'marimi', data_align)]:
            val = int(prod[key]) if key == 'id' else prod[key]
            c = ws.cell(row=row, column=col, value=val)
            c.font = data_font
            c.alignment = align

        # Buc
        c = ws.cell(row=row, column=6, value=prod['buc'])
        c.font, c.alignment, c.number_format = data_font, data_align, int_format

        # Preț intrare
        c = ws.cell(row=row, column=7, value=prod['pret_intrare'])
        c.font, c.alignment, c.number_format = data_font, data_align, money_format

        # Preț ieșire
        c = ws.cell(row=row, column=8, value=prod['pret_iesire'])
        c.font, c.alignment, c.number_format = data_font, data_align, money_format

        # Formule Excel
        formulas = {
            9: f'=F{row}*G{row}',                    # Valoare intrare
            10: f'=F{row}*H{row}',                   # Valoare ieșire
            11: f'=J{row}*M{row}/(100+M{row})',      # TVA
            12: f'=J{row}-I{row}-K{row}',            # Adaos comercial
        }
        for col, formula in formulas.items():
            c = ws.cell(row=row, column=col, value=formula)
            c.font, c.alignment, c.number_format = data_font, data_align, money_format

        # Cotă TVA
        c = ws.cell(row=row, column=13, value=prod['cota_tva'])
        c.font, c.alignment = data_font, data_align

        # Total intrare = Valoare intrare
        c = ws.cell(row=row, column=14, value=f'=I{row}')
        c.font, c.alignment, c.number_format = data_font, data_align, money_format

        # Border + alternating colors
        for col in range(1, 15):
            ws.cell(row=row, column=col).border = thin_border
        if i % 2 == 0:
            fill = PatternFill('solid', fgColor='F2F7FB')
            for col in range(1, 15):
                ws.cell(row=row, column=col).fill = fill

    # Totaluri
    total_row = start_row + len(products_data)
    ws.cell(row=total_row, column=1, value='TOTAL').font = total_font
    ws.merge_cells(f'A{total_row}:E{total_row}')
    ws.cell(row=total_row, column=1).alignment = Alignment(
        horizontal='right', vertical='center')

    last_data_row = total_row - 1
    for col, fmt in {6: int_format, 9: money_format, 10: money_format, 11: money_format, 12: money_format, 14: money_format}.items():
        col_letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f'=SUM({col_letter}{start_row}:{col_letter}{last_data_row})')
        c.font, c.alignment, c.number_format = total_font, data_align, fmt

    for col in range(1, 15):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = thin_border

    ws.freeze_panes = 'A4'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(output_path)
    return total_row


# ═══════════════════════════════════════════
#  TELEGRAM
# ═══════════════════════════════════════════

def send_telegram_message(text):
    if not TELEGRAM_BOT_TOKEN:
        print("  ⚠ TELEGRAM_BOT_TOKEN nu e setat în .env")
        return False
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    try:
        resp = requests.post(url, data={
                             'chat_id': TELEGRAM_CHAT_ID, 'text': text, 'parse_mode': 'HTML'}, timeout=15)
        return resp.status_code == 200
    except Exception as e:
        print(f"  ✗ Telegram mesaj eroare: {e}")
        return False


def send_telegram_file(filepath, caption=''):
    if not TELEGRAM_BOT_TOKEN:
        print("  ⚠ TELEGRAM_BOT_TOKEN nu e setat în .env")
        return False
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    try:
        with open(filepath, 'rb') as f:
            resp = requests.post(url, data={'chat_id': TELEGRAM_CHAT_ID, 'caption': caption, 'parse_mode': 'HTML'}, files={
                                 'document': f}, timeout=30)
        return resp.status_code == 200
    except Exception as e:
        print(f"  ✗ Telegram fișier eroare: {e}")
        return False


# ═══════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Generare NIR din Extended API')
    parser.add_argument('--start', type=int, help='ID produs start')
    parser.add_argument('--end', type=int, help='ID produs end')
    parser.add_argument('--ids', type=str,
                        help='Lista de ID-uri separate prin virgulă')
    parser.add_argument('--output', type=str, default=None,
                        help='Nume fișier output')
    parser.add_argument('--telegram', action='store_true',
                        help='Trimite fișierul pe Telegram')
    args = parser.parse_args()

    if not API_KEY:
        print("✗ EJOLIE_API_KEY nu e setat în .env!")
        sys.exit(1)

    if args.ids:
        product_ids = [int(x.strip()) for x in args.ids.split(',')]
    elif args.start and args.end:
        product_ids = list(range(args.start, args.end + 1))
    else:
        print("✗ Specifică --start și --end SAU --ids")
        sys.exit(1)

    if args.output:
        output_path = args.output
    else:
        output_path = os.path.join(
            SCRIPT_DIR, f"nir_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx")

    print(f"═══ Generare NIR ═══")
    print(
        f"  Produse: {len(product_ids)} (ID {product_ids[0]} → {product_ids[-1]})")
    print(f"  Output:  {output_path}")
    print(f"  Telegram: {'Da' if args.telegram else 'Nu'}")
    print()

    products_data = []
    skipped = 0
    warnings = []
    for i, pid in enumerate(product_ids):
        print(f"  [{i+1}/{len(product_ids)}] Fetch produs {pid}...", end=' ')
        product = fetch_product(pid)

        if not product:
            print("✗ Nu există sau eroare API")
            skipped += 1
            continue

        nir_row = extract_nir_data(product)

        if nir_row['buc'] == 0:
            msg = f"Stoc 0 - {nir_row['nume']} {nir_row['culoare']}"
            print(f"⚠ {msg} (skip)")
            warnings.append(msg)
            skipped += 1
            continue

        if nir_row['pret_intrare'] == 0:
            msg = f"Preț achiziție 0 - {nir_row['nume']} {nir_row['culoare']} (ID {pid})"
            print(f"⚠ {msg}")
            warnings.append(msg)

        print(
            f"✓ {nir_row['nume']} {nir_row['culoare']} | {nir_row['buc']} buc | in: {nir_row['pret_intrare']} | out: {nir_row['pret_iesire']}")
        products_data.append(nir_row)

        if i < len(product_ids) - 1:
            time.sleep(0.3)

    print()
    if not products_data:
        msg = "✗ Niciun produs valid găsit!"
        print(msg)
        if args.telegram:
            send_telegram_message(f"❌ NIR EȘUAT\n{msg}")
        sys.exit(1)

    print(f"  Generare Excel cu {len(products_data)} produse...")
    generate_nir_excel(products_data, output_path)

    total_buc = sum(p['buc'] for p in products_data)
    total_intrare = sum(p['buc'] * p['pret_intrare'] for p in products_data)
    total_iesire = sum(p['buc'] * p['pret_iesire'] for p in products_data)
    adaos_brut = total_iesire - total_intrare

    summary = (
        f"═══ NIR GENERAT ═══\n"
        f"  Fișier:          {os.path.basename(output_path)}\n"
        f"  Produse:         {len(products_data)} (skip: {skipped})\n"
        f"  Total bucăți:    {total_buc}\n"
        f"  Valoare intrare: {total_intrare:,.2f} RON\n"
        f"  Valoare ieșire:  {total_iesire:,.2f} RON\n"
        f"  Adaos brut:      {adaos_brut:,.2f} RON"
    )
    print(f"\n{summary}")

    if warnings:
        print(f"\n  ⚠ Avertismente ({len(warnings)}):")
        for w in warnings:
            print(f"    - {w}")

    if args.telegram:
        print(f"\n  📤 Trimit pe Telegram...")
        telegram_caption = (
            f"📋 <b>NIR — {datetime.now().strftime('%d.%m.%Y')}</b>\n\n"
            f"📦 Produse: {len(products_data)} ({total_buc} buc)\n"
            f"💰 Valoare intrare: {total_intrare:,.2f} RON\n"
            f"🏷 Valoare ieșire: {total_iesire:,.2f} RON\n"
            f"📈 Adaos brut: {adaos_brut:,.2f} RON"
        )
        if warnings:
            telegram_caption += f"\n\n⚠️ {len(warnings)} avertismente"

        if send_telegram_file(output_path, caption=telegram_caption):
            print("  ✓ Fișier trimis pe Telegram!")
        else:
            print("  ✗ Eroare la trimiterea pe Telegram")

    print(f"\n  📁 Fișier salvat: {output_path}")


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""
extract_emails.py — Extrage emailuri unice din comenzi ejolie.ro
Folosește Extended API, paginare lunară pentru a evita limita de 2000/request.
"""

import json
import os
import sys
import time
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from calendar import monthrange

# Load .env
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, '..', '.env')
if os.path.exists(ENV_PATH):
    with open(ENV_PATH) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())

API_KEY = os.environ.get('EJOLIE_API_KEY', '')
BASE_URL = 'https://ejolie.ro/api/'

def fetch_orders_month(year, month):
    """Fetch all orders for a given month."""
    last_day = monthrange(year, month)[1]
    data_start = f"01-{month:02d}-{year}"
    data_end = f"{last_day:02d}-{month:02d}-{year}"
    
    params = {
        'comenzi': '',
        'data_start': data_start,
        'data_end': data_end,
        'limit': '5000',
        'apikey': API_KEY,
    }
    
    query = '&'.join(
        f"{k}={urllib.parse.quote(str(v))}" if v else k
        for k, v in params.items()
    )
    url = f"{BASE_URL}?{query}"
    
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Extended API'})
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        
        if isinstance(data, dict):
            return data
        return {}
    except Exception as e:
        print(f"  ✗ Eroare {year}-{month:02d}: {e}")
        return {}


def main():
    if not API_KEY:
        print("✗ EJOLIE_API_KEY lipseste!")
        sys.exit(1)
    
    # Config
    cutoff_date = datetime(2025, 8, 1)  # pana la 01.08.2025
    max_emails = 10000
    
    # Detect earliest orders — start from 2020 (conservative)
    start_year = 2020
    start_month = 1
    
    all_emails = set()
    total_orders = 0
    month_stats = []
    
    print(f"═══ Extragere emailuri comenzi ejolie.ro ═══")
    print(f"  Perioada: {start_year}-01 → 2025-07 (înainte de {cutoff_date.strftime('%d.%m.%Y')})")
    print(f"  Limită: {max_emails} adrese")
    print()
    
    year = start_year
    month = start_month
    
    while True:
        if datetime(year, month, 1) >= cutoff_date:
            break
            
        print(f"  [{year}-{month:02d}] ", end='', flush=True)
        orders = fetch_orders_month(year, month)
        
        month_emails = set()
        for oid, order in orders.items():
            client = order.get('client', {})
            email = client.get('email', '').strip().lower()
            if email and '@' in email and email != '':
                month_emails.add(email)
        
        new_emails = month_emails - all_emails
        all_emails.update(month_emails)
        total_orders += len(orders)
        
        count_str = f"{len(orders)} comenzi, {len(month_emails)} emailuri ({len(new_emails)} noi)"
        print(f"✓ {count_str}")
        month_stats.append((f"{year}-{month:02d}", len(orders), len(month_emails), len(new_emails)))
        
        if len(all_emails) >= max_emails:
            print(f"\n  ⚠ Limită de {max_emails} atinsă!")
            break
        
        # Next month
        month += 1
        if month > 12:
            month = 1
            year += 1
        
        time.sleep(0.5)  # rate limiting
    
    # Save results
    print(f"\n═══ REZULTAT ═══")
    print(f"  Total comenzi scanate: {total_orders}")
    print(f"  Emailuri unice: {len(all_emails)}")
    
    # Sort emails
    sorted_emails = sorted(all_emails)[:max_emails]
    
    # Save CSV
    csv_path = os.path.join(SCRIPT_DIR, 'emailuri_comenzi_ejolie.csv')
    with open(csv_path, 'w') as f:
        f.write('email\n')
        for email in sorted_emails:
            f.write(f"{email}\n")
    print(f"  CSV salvat: {csv_path}")
    
    # Save Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Emailuri Comenzi"
        
        # Header
        header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_align = Alignment(horizontal='center', vertical='center')
        
        headers = ['#', 'Email']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        
        # Data
        for i, email in enumerate(sorted_emails, 1):
            ws.cell(row=i+1, column=1, value=i)
            ws.cell(row=i+1, column=2, value=email)
        
        # Stats sheet
        ws2 = wb.create_sheet("Statistici Lunare")
        stat_headers = ['Luna', 'Comenzi', 'Emailuri', 'Noi']
        for col, h in enumerate(stat_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        for i, (m, orders, emails, new) in enumerate(month_stats, 2):
            ws2.cell(row=i, column=1, value=m)
            ws2.cell(row=i, column=2, value=orders)
            ws2.cell(row=i, column=3, value=emails)
            ws2.cell(row=i, column=4, value=new)
        
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        
        xlsx_path = os.path.join(SCRIPT_DIR, 'emailuri_comenzi_ejolie.xlsx')
        wb.save(xlsx_path)
        print(f"  Excel salvat: {xlsx_path}")
        
    except ImportError:
        print("  ⚠ openpyxl nu e instalat, doar CSV generat")
        xlsx_path = None
    
    return csv_path, xlsx_path if 'xlsx_path' in dir() else None


if __name__ == '__main__':
    csv_path, xlsx_path = main()

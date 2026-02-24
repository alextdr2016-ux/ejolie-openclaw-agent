#!/usr/bin/env python3
"""Export Ejolie sales report to XLSX format."""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from utils import parse_period, fetch_orders, calculate_report, format_number, REPORT_STATUS

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

from collections import Counter
from datetime import datetime


def export_vanzari_xlsx(orders, period_label, output_path):
    """Export sales orders to a nicely formatted XLSX file."""
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Sumar (Summary) ──
    ws_sum = wb.active
    ws_sum.title = "Sumar"
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Calculate metrics
    metrics = calculate_report(orders)
    
    # Title
    ws_sum.merge_cells('A1:D1')
    ws_sum['A1'] = f"📊 RAPORT VÂNZĂRI — {period_label}"
    ws_sum['A1'].font = header_font
    ws_sum['A1'].fill = header_fill
    ws_sum['A1'].alignment = Alignment(horizontal='center')
    
    ws_sum['A2'] = f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_sum['A2'].font = Font(italic=True, size=9, color="888888")
    
    # Summary metrics
    summary_data = [
        ("📦 Total comenzi", metrics['total_comenzi']),
        ("💰 Valoare totală (RON)", metrics['valoare_totala']),
        ("🚚 Transport total (RON)", metrics['transport_total']),
        ("💵 Valoare netă (RON)", metrics['valoare_neta']),
        ("📈 Medie per comandă (RON)", metrics['medie_comanda']),
    ]
    
    row = 4
    for label, value in summary_data:
        ws_sum.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws_sum.cell(row=row, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = '#,##0.00'
        row += 1
    
    # Payment methods
    row += 1
    ws_sum.cell(row=row, column=1, value="💳 Metode plată").font = subheader_font
    ws_sum.cell(row=row, column=1).fill = subheader_fill
    ws_sum.cell(row=row, column=2, value="Comenzi").font = subheader_font
    ws_sum.cell(row=row, column=2).fill = subheader_fill
    row += 1
    for metoda, count in metrics['metode_plata'].most_common():
        ws_sum.cell(row=row, column=1, value=metoda)
        ws_sum.cell(row=row, column=2, value=count)
        row += 1
    
    # Top products
    row += 1
    ws_sum.cell(row=row, column=1, value="🏆 Top produse").font = subheader_font
    ws_sum.cell(row=row, column=1).fill = subheader_fill
    ws_sum.cell(row=row, column=2, value="Cantitate").font = subheader_font
    ws_sum.cell(row=row, column=2).fill = subheader_fill
    row += 1
    for name, qty in metrics['top_produse']:
        ws_sum.cell(row=row, column=1, value=name)
        ws_sum.cell(row=row, column=2, value=qty)
        row += 1
    
    ws_sum.column_dimensions['A'].width = 45
    ws_sum.column_dimensions['B'].width = 20
    ws_sum.column_dimensions['C'].width = 15
    ws_sum.column_dimensions['D'].width = 15
    
    # ── Sheet 2: Comenzi (All Orders) ──
    ws_orders = wb.create_sheet("Comenzi")
    
    order_headers = ["Nr. Comandă", "Data", "Client", "Email", "Telefon", 
                     "Total (RON)", "Transport (RON)", "Metodă plată", "Status"]
    
    for col, h in enumerate(order_headers, 1):
        cell = ws_orders.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 2
    for order_id, order in orders.items():
        if not isinstance(order, dict):
            continue
        
        ws_orders.cell(row=row, column=1, value=order.get("numar_comanda", order_id))
        ws_orders.cell(row=row, column=2, value=order.get("data_comanda", ""))
        
        client_name = f"{order.get('prenume', '')} {order.get('nume', '')}".strip()
        ws_orders.cell(row=row, column=3, value=client_name)
        ws_orders.cell(row=row, column=4, value=order.get("email", ""))
        ws_orders.cell(row=row, column=5, value=order.get("telefon", ""))
        
        try:
            total = float(str(order.get("total_comanda", 0)).replace(",", "."))
        except (ValueError, TypeError):
            total = 0
        cell = ws_orders.cell(row=row, column=6, value=total)
        cell.number_format = '#,##0.00'
        
        try:
            shipping = float(str(order.get("pret_livrare", 0)).replace(",", "."))
        except (ValueError, TypeError):
            shipping = 0
        cell = ws_orders.cell(row=row, column=7, value=shipping)
        cell.number_format = '#,##0.00'
        
        ws_orders.cell(row=row, column=8, value=order.get("metoda_plata", ""))
        ws_orders.cell(row=row, column=9, value=order.get("status_comanda", ""))
        
        for col in range(1, 10):
            ws_orders.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # Auto-width for orders sheet
    for col_letter, width in [('A', 15), ('B', 18), ('C', 25), ('D', 30), 
                               ('E', 15), ('F', 15), ('G', 15), ('H', 30), ('I', 20)]:
        ws_orders.column_dimensions[col_letter].width = width
    
    # Auto-filter
    ws_orders.auto_filter.ref = f"A1:I{row-1}"
    
    # ── Sheet 3: Produse (All Products) ──
    ws_prod = wb.create_sheet("Produse")
    
    prod_headers = ["Nr. Comandă", "Produs", "Brand", "Cantitate", "Preț unitar (RON)", "Total (RON)"]
    
    for col, h in enumerate(prod_headers, 1):
        cell = ws_prod.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 2
    for order_id, order in orders.items():
        if not isinstance(order, dict):
            continue
        
        nr_comanda = order.get("numar_comanda", order_id)
        produse = order.get("produse", {})
        if not isinstance(produse, dict):
            continue
        
        for pid, prod in produse.items():
            if not isinstance(prod, dict):
                continue
            
            nume = prod.get("nume", "Produs necunoscut")
            if "discount" in nume.lower():
                continue
            
            ws_prod.cell(row=row, column=1, value=nr_comanda)
            ws_prod.cell(row=row, column=2, value=nume)
            ws_prod.cell(row=row, column=3, value=prod.get("brand_nume", ""))
            
            try:
                qty = int(float(str(prod.get("cantitate", 1)).replace(",", ".")))
            except (ValueError, TypeError):
                qty = 1
            ws_prod.cell(row=row, column=4, value=qty)
            
            try:
                price = float(str(prod.get("pret_unitar", 0)).replace(",", "."))
            except (ValueError, TypeError):
                price = 0
            cell = ws_prod.cell(row=row, column=5, value=price)
            cell.number_format = '#,##0.00'
            
            cell = ws_prod.cell(row=row, column=6, value=price * qty)
            cell.number_format = '#,##0.00'
            
            for col in range(1, 7):
                ws_prod.cell(row=row, column=col).border = thin_border
            
            row += 1
    
    for col_letter, width in [('A', 15), ('B', 50), ('C', 15), ('D', 12), ('E', 18), ('F', 15)]:
        ws_prod.column_dimensions[col_letter].width = width
    
    ws_prod.auto_filter.ref = f"A1:F{row-1}"
    
    # Save
    wb.save(output_path)
    print(f"✅ Fișier Excel salvat: {output_path}")
    return output_path


def main():
    period_text = "luna februarie"
    data_start, data_end, period_label = parse_period(period_text)
    print(f"📅 Perioadă: {period_label}")
    
    print("📡 Se preiau comenzile din API...", flush=True)
    orders = fetch_orders(data_start, data_end)
    print(f"✅ {len(orders)} comenzi găsite.", flush=True)
    
    output_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        f"raport_vanzari_februarie_2026.xlsx"
    )
    
    export_vanzari_xlsx(orders, period_label, output_path)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
extract_emails_v2.py — Extrage emailuri unice din comenzi ejolie.ro
Perioade de 2 saptamani pentru a evita timeout-uri API.
"""

import json
import os
import sys
import time
import urllib.request
import urllib.parse
from datetime import datetime, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, '..', '.env')
if os.path.exists(ENV_PATH):
    with open(ENV_PATH) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())

API_KEY = os.environ.get('EJOLIE_API_KEY', '')
BASE_URL = 'https://ejolie.ro/api/'
CUTOFF = datetime(2025, 8, 1)
MAX_EMAILS = 10000
CHUNK_DAYS = 7
PROGRESS_FILE = os.path.join(SCRIPT_DIR, 'email_extract_progress.json')

def fetch_chunk(start_date, end_date, retries=2):
    ds = start_date.strftime('%d-%m-%Y')
    de = end_date.strftime('%d-%m-%Y')
    url = f"{BASE_URL}?comenzi&data_start={ds}&data_end={de}&limit=1000&apikey={API_KEY}"
    
    for attempt in range(retries + 1):
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Extended API'})
            with urllib.request.urlopen(req, timeout=90) as resp:
                data = json.loads(resp.read().decode('utf-8'))
            return data if isinstance(data, dict) else {}
        except Exception as e:
            if attempt < retries:
                time.sleep(3)
            else:
                return None  # failed after retries
    return None

def save_progress(emails, current_date, stats):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump({
            'emails': list(emails),
            'current_date': current_date.isoformat(),
            'stats': stats,
        }, f)

def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            data = json.load(f)
        return (
            set(data['emails']),
            datetime.fromisoformat(data['current_date']),
            data['stats']
        )
    return None

def main():
    if not API_KEY:
        print("✗ EJOLIE_API_KEY lipseste!")
        sys.exit(1)
    
    # Try to resume
    progress = load_progress()
    if progress:
        all_emails, current_date, stats = progress
        print(f"▶ Reluare de la {current_date.strftime('%d.%m.%Y')} ({len(all_emails)} emailuri deja)")
    else:
        all_emails = set()
        current_date = datetime(2021, 1, 1)  # start from 2021
        stats = {'total_orders': 0, 'chunks_ok': 0, 'chunks_fail': 0}
    
    print(f"═══ Extragere emailuri comenzi ejolie.ro ═══")
    print(f"  Perioada: {current_date.strftime('%d.%m.%Y')} → {CUTOFF.strftime('%d.%m.%Y')}")
    print(f"  Chunk: {CHUNK_DAYS} zile | Limită: {MAX_EMAILS}")
    print()
    
    while current_date < CUTOFF:
        chunk_end = min(current_date + timedelta(days=CHUNK_DAYS - 1), CUTOFF - timedelta(days=1))
        
        label = f"{current_date.strftime('%d.%m.%Y')}-{chunk_end.strftime('%d.%m.%Y')}"
        print(f"  [{label}] ", end='', flush=True)
        
        orders = fetch_chunk(current_date, chunk_end)
        
        if orders is None:
            print(f"✗ TIMEOUT (skip)")
            stats['chunks_fail'] += 1
            # Try smaller chunk (7 days)
            if CHUNK_DAYS > 7:
                mid = current_date + timedelta(days=6)
                orders1 = fetch_chunk(current_date, mid)
                orders2 = fetch_chunk(mid + timedelta(days=1), chunk_end)
                if orders1 is not None and orders2 is not None:
                    orders = {**orders1, **orders2}
                    stats['chunks_fail'] -= 1
                    print(f"    ↳ Split OK: {len(orders)} comenzi")
                else:
                    current_date = chunk_end + timedelta(days=1)
                    save_progress(all_emails, current_date, stats)
                    time.sleep(2)
                    continue
            else:
                current_date = chunk_end + timedelta(days=1)
                save_progress(all_emails, current_date, stats)
                time.sleep(2)
                continue
        
        chunk_emails = set()
        for oid, order in orders.items():
            client = order.get('client', {})
            if isinstance(client, dict):
                email = client.get('email', '').strip().lower()
                if email and '@' in email and len(email) > 5:
                    chunk_emails.add(email)
        
        new = chunk_emails - all_emails
        all_emails.update(chunk_emails)
        stats['total_orders'] += len(orders)
        stats['chunks_ok'] += 1
        
        print(f"✓ {len(orders)} comenzi, {len(new)} emailuri noi (total: {len(all_emails)})")
        
        current_date = chunk_end + timedelta(days=1)
        save_progress(all_emails, current_date, stats)
        
        if len(all_emails) >= MAX_EMAILS:
            print(f"\n  ⚠ Limită {MAX_EMAILS} atinsă!")
            break
        
        time.sleep(0.8)
    
    # === SAVE RESULTS ===
    sorted_emails = sorted(all_emails)[:MAX_EMAILS]
    
    print(f"\n═══ REZULTAT ═══")
    print(f"  Total comenzi: {stats['total_orders']}")
    print(f"  Emailuri unice: {len(sorted_emails)}")
    print(f"  Chunks OK/Fail: {stats['chunks_ok']}/{stats['chunks_fail']}")
    
    # CSV
    csv_path = os.path.join(SCRIPT_DIR, 'emailuri_comenzi_ejolie.csv')
    with open(csv_path, 'w') as f:
        f.write('email\n')
        for e in sorted_emails:
            f.write(f"{e}\n")
    print(f"  CSV: {csv_path}")
    
    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Emailuri Comenzi"
        
        hfont = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        hfill = PatternFill('solid', fgColor='4472C4')
        halign = Alignment(horizontal='center')
        
        ws.cell(row=1, column=1, value='#').font = hfont
        ws.cell(row=1, column=1).fill = hfill
        ws.cell(row=1, column=2, value='Email').font = hfont
        ws.cell(row=1, column=2).fill = hfill
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        
        for i, email in enumerate(sorted_emails, 1):
            ws.cell(row=i+1, column=1, value=i)
            ws.cell(row=i+1, column=2, value=email)
        
        xlsx_path = os.path.join(SCRIPT_DIR, 'emailuri_comenzi_ejolie.xlsx')
        wb.save(xlsx_path)
        print(f"  Excel: {xlsx_path}")
    except Exception as e:
        print(f"  ⚠ Excel error: {e}")
        xlsx_path = None
    
    # Copy to /tmp for Telegram
    import shutil
    tmp_xlsx = '/tmp/emailuri_comenzi_ejolie.xlsx'
    tmp_csv = '/tmp/emailuri_comenzi_ejolie.csv'
    if xlsx_path:
        shutil.copy2(xlsx_path, tmp_xlsx)
    shutil.copy2(csv_path, tmp_csv)
    print(f"\n  Fișiere în /tmp/ gata de trimis!")
    
    # Cleanup progress file
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)
    
    print(f"\n✅ DONE — {len(sorted_emails)} emailuri extrase")

if __name__ == '__main__':
    main()

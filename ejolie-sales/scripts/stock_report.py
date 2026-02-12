#!/usr/bin/env python3
"""Ejolie.ro Stock Report - reads from local cache (instant)"""

import os, json, argparse

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_FILE = os.path.join(SCRIPT_DIR, "stock_cache.json")


def load_cache():
    if not os.path.exists(CACHE_FILE):
        print("âŒ Cache nu existÄƒ! RuleazÄƒ mai Ã®ntÃ¢i: python3 stock_cache_update.py")
        return None
    with open(CACHE_FILE) as f:
        return json.load(f)


def generate_report(brand=None, only_in_stock=True):
    cache = load_cache()
    if not cache:
        return []
    
    print(f"ðŸ“¦ Raport stoc (cache din {cache['updated']})")
    
    rows = []
    for pid, prod in cache["products"].items():
        if brand and brand.lower() not in prod["brand"].lower():
            continue
        
        sizes = prod.get("sizes", {})
        if not sizes:
            in_stock = "In stoc" in str(prod.get("stoc_general", ""))
            if only_in_stock and not in_stock:
                continue
            rows.append({
                "produs": prod["nume"], "cod": prod["cod"], "brand": prod["brand"],
                "marime": "-", "stoc": prod["stoc_general"],
                "in_stock": in_stock, "pret": float(prod["pret"] or 0),
            })
            continue
        
        for size_name, size_data in sizes.items():
            if only_in_stock and not size_data["in_stock"]:
                continue
            pret_normal = float(size_data["pret"] or 0)
            pret_disc = float(size_data.get("pret_discount") or 0)
            pret_final = pret_disc if pret_disc > 0 and pret_disc < pret_normal else pret_normal
            rows.append({
                "produs": prod["nume"], "cod": prod["cod"], "brand": prod["brand"],
                "marime": size_name, "stoc": size_data["stoc"],
                "stoc_fizic": size_data.get("stoc_fizic", 0),
                "in_stock": size_data["in_stock"],
                "pret": pret_final,
            })
    
    in_stock = sum(1 for r in rows if r["in_stock"])
    print(f"âœ… {in_stock} mÄƒrimi Ã®n stoc")
    return rows


def export_xlsx(rows, brand=None, output=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        os.system("pip3 install openpyxl --break-system-packages -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    if output is None:
        brand_label = f"_{brand}" if brand else ""
        output = f"/home/ubuntu/raport_stoc{brand_label}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stoc pe MÄƒrimi"
    
    headers = ["Produs", "Cod", "Brand", "MÄƒrime", "Stoc", "BucÄƒÈ›i", "PreÈ› (RON)", "Valoare (RON)"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row_idx, r in enumerate(rows, 2):
        data = [r["produs"], r["cod"], r["brand"], r["marime"], r["stoc"], r.get("stoc_fizic", 0), r["pret"], r.get("stoc_fizic", 0) * r["pret"]]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col == 5:
                cell.fill = green_fill if r["in_stock"] else PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            if col in (7, 8):
                cell.number_format = '#,##0.00'
    
    widths = [45, 15, 12, 10, 12, 8, 12, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    
    # Add TOTAL row
    total_row = len(rows) + 2
    total_buc = sum(r.get("stoc_fizic", 0) for r in rows)
    total_val = sum(r.get("stoc_fizic", 0) * r["pret"] for r in rows)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=12)
    ws.cell(row=total_row, column=6, value=total_buc).font = Font(bold=True, size=12)
    ws.cell(row=total_row, column=8, value=total_val).font = Font(bold=True, size=12)
    ws.cell(row=total_row, column=8).number_format = '#,##0.00'
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).border = thin_border
        ws.cell(row=total_row, column=col).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    # Sheet 2: Sumar
    ws2 = wb.create_sheet("Sumar Produse")
    sum_headers = ["Produs", "Cod", "Brand", "MÄƒrimi Ã®n Stoc", "Total MÄƒrimi", "PreÈ›"]
    
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    prod_summary = {}
    for r in rows:
        key = r["cod"] or r["produs"]
        if key not in prod_summary:
            prod_summary[key] = {"produs": r["produs"], "cod": r["cod"], "brand": r["brand"],
                                 "in_stock": 0, "total": 0, "pret": r["pret"]}
        prod_summary[key]["total"] += 1
        if r["in_stock"]:
            prod_summary[key]["in_stock"] += 1
    
    for row_idx, (key, s) in enumerate(sorted(prod_summary.items(), key=lambda x: x[1]["produs"]), 2):
        data = [s["produs"], s["cod"], s["brand"], s["in_stock"], s["total"], s["pret"]]
        for col, val in enumerate(data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col == 6:
                cell.number_format = '#,##0.00'
    
    for col, w in enumerate([45, 15, 12, 15, 15, 12], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    
    cache = load_cache()
    updated = cache["updated"] if cache else "?"
    ws2.cell(row=len(prod_summary) + 3, column=1, value=f"Cache actualizat: {updated}").font = Font(italic=True, color="888888")
    
    wb.save(output)
    print(f"âœ… Excel salvat: {output}")
    return output


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--brand", default=None)
    parser.add_argument("--all", action="store_true")
    parser.add_argument("--output", default=None)
    parser.add_argument("--format", choices=["text", "xlsx"], default="xlsx")
    args = parser.parse_args()
    
    rows = generate_report(brand=args.brand, only_in_stock=not args.all)
    
    if not rows:
        return
    
    if args.format == "xlsx":
        output = export_xlsx(rows, brand=args.brand, output=args.output)
        print(f"XLSX {output}")
    else:
        print(f"\nðŸ“¦ RAPORT STOC{f' - {args.brand.upper()}' if args.brand else ''}")
        print("â”" * 60)
        current_prod = ""
        for r in rows:
            if r["produs"] != current_prod:
                current_prod = r["produs"]
                print(f"\n{r['produs']} ({r['cod']}) - {r['pret']} RON")
            status = "âœ…" if r["in_stock"] else "âŒ"
            print(f"  {status} MÄƒrime {r['marime']}: {r['stoc']} ({r.get('stoc_fizic', 0)} buc)")


if __name__ == "__main__":
    main()
